import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Toplevel, Label, Button, StringVar
from datetime import datetime, timedelta
import sqlite3
import json, os
import hashlib
import uuid
import time
import shutil
import sys
import traceback
import base64
import platform
import functools

try:
    from mobile_order_server import MobileOrderServer, show_mobile_order_popup
    MOBILE_SERVER_AVAILABLE = True
except ImportError:
    MOBILE_SERVER_AVAILABLE = False
 
# ═══════════════════════════════════════════════════════════════
#  PYINSTALLER RESOURCE PATH
#  Resolves asset paths whether running as script or .exe bundle
# ═══════════════════════════════════════════════════════════════
def resource_path(relative_path: str) -> str:
    """Return absolute path to a resource — works for dev and PyInstaller .exe."""
    try:
        # PyInstaller extracts files to a temp folder stored in sys._MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)
 
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.pdfgen import canvas as pdf_canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
 
try:
    import win32print
    WINDOWS_PRINT_AVAILABLE = True
except ImportError:
    WINDOWS_PRINT_AVAILABLE = False
 
try:
    from cryptography.fernet import Fernet
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False
 
# ═══════════════════════════════════════════════════════════════
#  APP PATHS & HIDDEN DIRECTORY
# ═══════════════════════════════════════════════════════════════
def _get_app_base_path():
    try:
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        else:
            return os.path.dirname(os.path.abspath(__file__))
    except Exception:
        return os.getcwd()
 
_APP_HIDDEN_DIR = os.path.join(_get_app_base_path(), ".RestoBillData")
 
def _ensure_hidden_dir(path=None):
    if path is None:
        path = _APP_HIDDEN_DIR
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)
    try:
        import ctypes
        if sys.platform == "win32":
            ctypes.windll.kernel32.SetFileAttributesW(path, 0x02 | 0x04)
    except Exception:
        pass
    return path
 
def _hidden_path(filename):
    _ensure_hidden_dir()
    return os.path.join(_APP_HIDDEN_DIR, filename)


def _demo_mode_enabled() -> bool:
    return os.environ.get("RESTOBILL_DEMO_AUTORUN", "").strip().lower() in {
        "1", "true", "yes", "on"
    }
 
# ═══════════════════════════════════════════════════════════════
#  CENTRAL FILE PATHS
# ═══════════════════════════════════════════════════════════════
DB_FILE          = _hidden_path("restaurant_data.db")
AUTH_DB_FILE     = _hidden_path("users.db")
CONFIG_FILE      = _hidden_path("printer_config.json")
SESSION_FILE     = _hidden_path("sessions.json")
LICENSE_FILE     = _hidden_path("license.dat")
LAST_BILL_FILE   = _hidden_path("last_bill.json")
SALES_EXCEL_FILE = _hidden_path("sales_report.xlsx")
CLOCK_GUARD_FILE = _hidden_path("clock_guard.dat")
 
# ═══════════════════════════════════════════════════════════════
#  ESC/POS THERMAL PRINTER COMMANDS
# ═══════════════════════════════════════════════════════════════
ESC_INIT          = b"\x1b@"
ESC_CUT           = b"\x1dV\x00"
ESC_BOLD_ON       = b"\x1bE\x01"
ESC_BOLD_OFF      = b"\x1bE\x00"
ESC_ALIGN_CENTER  = b"\x1ba\x01"
ESC_ALIGN_LEFT    = b"\x1ba\x00"
ESC_DOUBLE_HEIGHT = b"\x1b!\x10"
ESC_NORMAL_SIZE   = b"\x1b!\x00"
ESC_FEED_LINES    = lambda n: bytes([0x1b, 0x64, n])
 
LOW_STOCK_THRESHOLD = 5
 
DEFAULT_CATEGORIES = [
    "Starters", "Main Course", "Breads", "Rice & Biryani",
    "Soups", "Salads", "Desserts", "Beverages", "Specials"
]
 
 
# ═══════════════════════════════════════════════════════════════
#  CLOCK GUARD  (rollback detection)
# ═══════════════════════════════════════════════════════════════
def _cg_derive_key() -> bytes:
    raw = str(uuid.getnode()) + "CLOCKGUARD_2026"
    return base64.urlsafe_b64encode(hashlib.sha256(raw.encode()).digest())
 
 
def _cg_read_max_ts() -> float:
    try:
        from cryptography.fernet import Fernet as _F
        if not os.path.exists(CLOCK_GUARD_FILE):
            return 0.0
        with open(CLOCK_GUARD_FILE, "rb") as f:
            data = f.read()
        decrypted = _F(_cg_derive_key()).decrypt(data)
        return float(decrypted.decode())
    except Exception:
        return 0.0
 
 
def _cg_write_ts(ts: float):
    try:
        from cryptography.fernet import Fernet as _F
        current = _cg_read_max_ts()
        if ts <= current:
            return
        _ensure_hidden_dir()
        encrypted = _F(_cg_derive_key()).encrypt(str(ts).encode())
        with open(CLOCK_GUARD_FILE, "wb") as f:
            f.write(encrypted)
    except Exception:
        pass
 
 
def _cg_check_rollback() -> bool:
    max_ts = _cg_read_max_ts()
    if max_ts == 0.0:
        return False
    now = time.time()
    TOLERANCE = 300
    return now < (max_ts - TOLERANCE)
 
 
def _cg_update():
    _cg_write_ts(time.time())
 
 
# ═══════════════════════════════════════════════════════════════
#  DATABASE MANAGER
# ═══════════════════════════════════════════════════════════════
class DatabaseManager:
    def __init__(self, db_name=None):
        if db_name is None:
            db_name = DB_FILE
        _ensure_hidden_dir()
        self.conn = sqlite3.connect(db_name)
        self.conn.row_factory = sqlite3.Row
        self.create_tables()
 
    def create_tables(self):
        c = self.conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS menu_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            category TEXT,
            price REAL,
            available INTEGER DEFAULT 1,
            description TEXT DEFAULT ''
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS restaurant_tables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_number TEXT UNIQUE,
            capacity INTEGER DEFAULT 4,
            status TEXT DEFAULT 'free'
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT UNIQUE,
            table_number TEXT,
            order_type TEXT DEFAULT 'dine_in',
            status TEXT DEFAULT 'open',
            created_at TEXT,
            updated_at TEXT,
            waiter TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            token_number INTEGER DEFAULT 0
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            item_name TEXT,
            category TEXT,
            quantity INTEGER,
            price REAL,
            total REAL,
            notes TEXT DEFAULT '',
            kot_printed INTEGER DEFAULT 0,
            FOREIGN KEY(order_id) REFERENCES orders(id)
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_number TEXT UNIQUE,
            order_id INTEGER,
            table_number TEXT,
            order_type TEXT,
            subtotal REAL,
            discount REAL DEFAULT 0,
            tax_percent REAL DEFAULT 5.0,
            tax_amount REAL,
            sgst_percent REAL DEFAULT 0.0,
            sgst_amount REAL DEFAULT 0.0,
            cgst_percent REAL DEFAULT 0.0,
            cgst_amount REAL DEFAULT 0.0,
            service_charge REAL DEFAULT 0,
            total REAL,
            payment_method TEXT DEFAULT 'cash',
            created_at TEXT,
            items_json TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS daily_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            bill_number TEXT,
            table_number TEXT,
            order_type TEXT,
            subtotal REAL,
            discount REAL,
            tax REAL,
            service_charge REAL,
            total REAL,
            payment_method TEXT,
            created_at TEXT
        )""")
        # New table for cancelled orders
        c.execute("""CREATE TABLE IF NOT EXISTS cancelled_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            order_number TEXT,
            table_number TEXT,
            order_type TEXT,
            customer_name TEXT,
            total_items INTEGER,
            subtotal REAL,
            cancelled_by TEXT,
            reason TEXT,
            cancelled_at TEXT,
            items_json TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )""")
        self.conn.commit()
        self._seed_default_tables()
        self._seed_default_settings()
        self._migrate_bills_table()
 
    def _migrate_bills_table(self):
        c = self.conn.cursor()
        c.execute("PRAGMA table_info(bills)")
        cols = [row[1] for row in c.fetchall()]
        if "sgst_percent" not in cols:
            c.execute("ALTER TABLE bills ADD COLUMN sgst_percent REAL DEFAULT 0.0")
        if "sgst_amount" not in cols:
            c.execute("ALTER TABLE bills ADD COLUMN sgst_amount REAL DEFAULT 0.0")
        if "cgst_percent" not in cols:
            c.execute("ALTER TABLE bills ADD COLUMN cgst_percent REAL DEFAULT 0.0")
        if "cgst_amount" not in cols:
            c.execute("ALTER TABLE bills ADD COLUMN cgst_amount REAL DEFAULT 0.0")
        # Migrate orders table — add token_number if missing
        c.execute("PRAGMA table_info(orders)")
        order_cols = [row[1] for row in c.fetchall()]
        if "token_number" not in order_cols:
            c.execute("ALTER TABLE orders ADD COLUMN token_number INTEGER DEFAULT 0")
        # Create cancelled_orders table if not exists (for existing databases)
        c.execute("""CREATE TABLE IF NOT EXISTS cancelled_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            order_number TEXT,
            table_number TEXT,
            order_type TEXT,
            customer_name TEXT,
            total_items INTEGER,
            subtotal REAL,
            cancelled_by TEXT,
            reason TEXT,
            cancelled_at TEXT,
            items_json TEXT
        )""")
        self.conn.commit()
 
    def _get_next_token(self) -> int:
        """Return next daily token number (resets to 1 each day)."""
        today = datetime.now().strftime("%Y-%m-%d")
        c = self.conn.cursor()
        c.execute("""SELECT MAX(token_number) FROM orders
                     WHERE DATE(created_at)=? AND token_number > 0""", (today,))
        row = c.fetchone()
        last = row[0] if row and row[0] else 0
        return last + 1
 
    def _seed_default_tables(self):
        c = self.conn.cursor()
        c.execute("SELECT COUNT(*) FROM restaurant_tables")
        if c.fetchone()[0] == 0:
            for i in range(1, 13):
                c.execute("INSERT OR IGNORE INTO restaurant_tables (table_number, capacity) VALUES (?, ?)",
                          (f"T{i:02d}", 4))
        self.conn.commit()
 
    def _seed_default_settings(self):
        defaults = {
            "restaurant_name": "MY RESTAURANT",
            "address":         "123 Food Street, City",
            "phone":           "+91 98765 43210",
            "gst_number":      "",
            "tax_percent":     "5.0",
            "sgst_percent":    "2.5",
            "cgst_percent":    "2.5",
            "service_charge":  "0.0",
            "currency":        "Rs.",
            "footer_message":  "Thank you! Visit again!",
            "enable_table":    "1",
            "enable_category": "1",
            "mobile_url_override": "",
        }
        c = self.conn.cursor()
        for k, v in defaults.items():
            c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (k, v))
        self.conn.commit()
 
    def get_setting(self, key, default=""):
        c = self.conn.cursor()
        c.execute("SELECT value FROM settings WHERE key=?", (key,))
        row = c.fetchone()
        return row["value"] if row else default
 
    def set_setting(self, key, value):
        c = self.conn.cursor()
        c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
        self.conn.commit()
 
    def add_menu_item(self, name, category, price, description=""):
        c = self.conn.cursor()
        c.execute("INSERT INTO menu_items (name, category, price, description) VALUES (?, ?, ?, ?)",
                  (name, category, price, description))
        self.conn.commit()
 
    def update_menu_item(self, old_name, name, category, price, description=""):
        c = self.conn.cursor()
        c.execute("UPDATE menu_items SET name=?, category=?, price=?, description=? WHERE name=?",
                  (name, category, price, description, old_name))
        self.conn.commit()
 
    def delete_menu_item(self, name):
        c = self.conn.cursor()
        c.execute("DELETE FROM menu_items WHERE name=?", (name,))
        self.conn.commit()
 
    def fetch_menu(self):
        c = self.conn.cursor()
        c.execute("SELECT name, category, price, available, description FROM menu_items ORDER BY category, name")
        return c.fetchall()
 
    def fetch_menu_by_category(self, category):
        c = self.conn.cursor()
        c.execute("SELECT name, price, description FROM menu_items WHERE category=? AND available=1 ORDER BY name",
                  (category,))
        return c.fetchall()
 
    def toggle_item_availability(self, name):
        c = self.conn.cursor()
        c.execute("UPDATE menu_items SET available = 1 - available WHERE name=?", (name,))
        self.conn.commit()
 
    def fetch_tables(self):
        c = self.conn.cursor()
        c.execute("SELECT table_number, capacity, status FROM restaurant_tables ORDER BY table_number")
        return c.fetchall()
 
    def set_table_status(self, table_number, status):
        c = self.conn.cursor()
        c.execute("UPDATE restaurant_tables SET status=? WHERE table_number=?", (status, table_number))
        self.conn.commit()
 
    def add_table(self, table_number, capacity=4):
        c = self.conn.cursor()
        c.execute("INSERT OR IGNORE INTO restaurant_tables (table_number, capacity) VALUES (?, ?)",
                  (table_number, capacity))
        self.conn.commit()
 
    def delete_table(self, table_number):
        c = self.conn.cursor()
        c.execute("DELETE FROM restaurant_tables WHERE table_number=?", (table_number,))
        self.conn.commit()
 
    def create_order(self, table_number, order_type="dine_in", waiter="", customer_name="", notes=""):
        order_number = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
        now   = datetime.now().isoformat()
        token = self._get_next_token()
        c = self.conn.cursor()
        c.execute("""INSERT INTO orders (order_number, table_number, order_type, status, created_at, updated_at, waiter, customer_name, notes, token_number)
                     VALUES (?, ?, ?, 'open', ?, ?, ?, ?, ?, ?)""",
                  (order_number, table_number, order_type, now, now, waiter, customer_name, notes, token))
        self.conn.commit()
        if order_type == "dine_in":
            self.set_table_status(table_number, "occupied")
        return c.lastrowid, order_number
 
    def get_open_order_for_table(self, table_number):
        c = self.conn.cursor()
        c.execute("SELECT id, order_number FROM orders WHERE table_number=? AND status='open' ORDER BY created_at DESC LIMIT 1",
                  (table_number,))
        return c.fetchone()
 
    def get_all_open_orders(self):
        c = self.conn.cursor()
        c.execute("""SELECT id, order_number, table_number, order_type,
                            customer_name, token_number, created_at
                     FROM orders WHERE status='open'
                     ORDER BY token_number ASC, created_at ASC""")
        return c.fetchall()
 
    def add_order_item(self, order_id, item_name, category, quantity, price, notes=""):
        total = quantity * price
        now = datetime.now().isoformat()
        c = self.conn.cursor()
        c.execute("SELECT id, quantity, total FROM order_items WHERE order_id=? AND item_name=? AND notes=?",
                  (order_id, item_name, notes))
        existing = c.fetchone()
        if existing:
            new_qty   = existing["quantity"] + quantity
            new_total = new_qty * price
            c.execute("UPDATE order_items SET quantity=?, total=? WHERE id=?",
                      (new_qty, new_total, existing["id"]))
        else:
            c.execute("INSERT INTO order_items (order_id, item_name, category, quantity, price, total, notes) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (order_id, item_name, category, quantity, price, total, notes))
        c.execute("UPDATE orders SET updated_at=? WHERE id=?", (now, order_id))
        self.conn.commit()
 
    def get_order_items(self, order_id):
        c = self.conn.cursor()
        c.execute("SELECT * FROM order_items WHERE order_id=? ORDER BY id", (order_id,))
        return c.fetchall()
 
    def remove_order_item(self, item_id):
        c = self.conn.cursor()
        c.execute("DELETE FROM order_items WHERE id=?", (item_id,))
        self.conn.commit()
 
    def update_order_item_qty(self, item_id, new_qty):
        c = self.conn.cursor()
        c.execute("SELECT price FROM order_items WHERE id=?", (item_id,))
        row = c.fetchone()
        if row:
            c.execute("UPDATE order_items SET quantity=?, total=? WHERE id=?",
                      (new_qty, new_qty * row["price"], item_id))
            self.conn.commit()
 
    def get_order(self, order_id):
        c = self.conn.cursor()
        c.execute("SELECT * FROM orders WHERE id=?", (order_id,))
        return c.fetchone()
 
    def mark_kot_printed(self, order_id):
        c = self.conn.cursor()
        c.execute("UPDATE order_items SET kot_printed=1 WHERE order_id=?", (order_id,))
        self.conn.commit()
 
    def get_unprinted_kot_items(self, order_id):
        c = self.conn.cursor()
        c.execute("SELECT * FROM order_items WHERE order_id=? AND kot_printed=0", (order_id,))
        return c.fetchall()
 
    def generate_bill_number(self):
        return f"BILL-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
 
    def create_bill(self, order_id, discount=0, payment_method="cash"):
        order = self.get_order(order_id)
        items = self.get_order_items(order_id)
        if not items:
            return None, "No items in order"
 
        subtotal  = sum(i["total"] for i in items)
        tax_pct   = float(self.get_setting("tax_percent",  "5.0"))
        sgst_pct  = float(self.get_setting("sgst_percent", "0.0"))
        cgst_pct  = float(self.get_setting("cgst_percent", "0.0"))
        svc_pct   = float(self.get_setting("service_charge", "0.0"))
        disc_amt  = min(discount, subtotal)
 
        combined_gst = sgst_pct + cgst_pct if (sgst_pct + cgst_pct) > 0 else tax_pct
        taxable_base = round((subtotal - disc_amt) / (1 + combined_gst / 100), 2) if combined_gst > 0 else (subtotal - disc_amt)
 
        sgst_amt = round(taxable_base * sgst_pct / 100, 2) if sgst_pct > 0 else 0.0
        cgst_amt = round(taxable_base * cgst_pct / 100, 2) if cgst_pct > 0 else 0.0
        if sgst_pct + cgst_pct > 0:
            tax_amt = round(sgst_amt + cgst_amt, 2)
        else:
            tax_amt = round(taxable_base * tax_pct / 100, 2)
        svc_amt  = round(taxable_base * svc_pct / 100, 2)
        total    = round(taxable_base + tax_amt + svc_amt, 2)
 
        bill_number = self.generate_bill_number()
        now         = datetime.now().isoformat()
        items_json  = json.dumps([dict(i) for i in items])
 
        c = self.conn.cursor()
        c.execute("""INSERT INTO bills
            (bill_number, order_id, table_number, order_type, subtotal, discount,
             tax_percent, tax_amount, sgst_percent, sgst_amount, cgst_percent, cgst_amount,
             service_charge, total, payment_method, created_at, items_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (bill_number, order_id, order["table_number"], order["order_type"],
             subtotal, disc_amt, tax_pct, tax_amt, sgst_pct, sgst_amt, cgst_pct, cgst_amt,
             svc_amt, total, payment_method, now, items_json))
        bill_id = c.lastrowid

        c.execute("""INSERT INTO daily_sales
            (date, bill_number, table_number, order_type, subtotal, discount, tax,
             service_charge, total, payment_method, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (datetime.now().strftime("%Y-%m-%d"), bill_number, order["table_number"],
             order["order_type"], subtotal, disc_amt, tax_amt, svc_amt, total, payment_method, now))
 
        c.execute("UPDATE orders SET status='closed', updated_at=? WHERE id=?", (now, order_id))
        if order["order_type"] == "dine_in":
            self.set_table_status(order["table_number"], "free")
        self.conn.commit()
 
        c.execute("SELECT * FROM bills WHERE id=?", (bill_id,))
        return c.fetchone(), None
 
    def get_daily_sales(self, date_str):
        c = self.conn.cursor()
        c.execute("SELECT * FROM daily_sales WHERE date=? ORDER BY created_at", (date_str,))
        return c.fetchall()
 
    def get_bill(self, bill_id):
        c = self.conn.cursor()
        c.execute("SELECT * FROM bills WHERE id=?", (bill_id,))
        return c.fetchone()
 
    def backup_database(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("Database Backup", "*.db")],
            title="Save Backup As")
        if not path:
            return
        try:
            shutil.copy(DB_FILE, path)
            messagebox.showinfo("Backup Successful", f"Backup saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Backup Failed", str(e))
 
    def restore_database(self, on_success=None):
        path = filedialog.askopenfilename(
            filetypes=[("Database Backup", "*.db")],
            title="Select Backup File")
        if not path:
            return
        if not messagebox.askyesno("Confirm Restore", "This will overwrite all current data.\n\nContinue?"):
            return
        try:
            self.conn.close()
            shutil.copy(path, DB_FILE)
            self.conn = sqlite3.connect(DB_FILE)
            self.conn.row_factory = sqlite3.Row
            self.create_tables()
            messagebox.showinfo("Restore Successful", "Database restored successfully.")
            if on_success:
                on_success()
        except Exception as e:
            messagebox.showerror("Restore Failed", str(e))
            try:
                self.conn = sqlite3.connect(DB_FILE)
                self.conn.row_factory = sqlite3.Row
            except Exception:
                pass
 
    def cancel_order(self, order_id, cancelled_by, reason):
        """Cancel an order and move it to cancelled_orders table."""
        c = self.conn.cursor()
        
        # Get order details
        order = self.get_order(order_id)
        if not order:
            return False, "Order not found"
        
        if order["status"] != "open":
            return False, "Only open orders can be cancelled"
        
        # Get order items
        items = self.get_order_items(order_id)
        if not items:
            # Allow cancelling even with 0 items (e.g. after clear)
            items = []
                
        # Calculate totals
        total_items = sum(i["quantity"] for i in items) if items else 0
        subtotal = sum(i["total"] for i in items) if items else 0.0
        
        # Insert into cancelled_orders
        now = datetime.now().isoformat()
        items_json = json.dumps([dict(i) for i in items])
        
        c.execute("""INSERT INTO cancelled_orders 
            (order_id, order_number, table_number, order_type, customer_name,
             total_items, subtotal, cancelled_by, reason, cancelled_at, items_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (order_id, order["order_number"], order["table_number"], 
             order["order_type"], order["customer_name"],
             total_items, subtotal, cancelled_by, reason, now, items_json))
        
        # Update order status to cancelled
        c.execute("UPDATE orders SET status='cancelled', updated_at=? WHERE id=?", 
                  (now, order_id))
        
        # Free up the table if it was dine-in
        if order["order_type"] == "dine_in":
            self.set_table_status(order["table_number"], "free")
        
        self.conn.commit()
        return True, "Order cancelled successfully"
    
    def get_cancelled_orders(self, date_str=None):
        """Get cancelled orders, optionally filtered by date."""
        c = self.conn.cursor()
        if date_str:
            c.execute("""SELECT * FROM cancelled_orders 
                         WHERE DATE(cancelled_at)=? 
                         ORDER BY cancelled_at DESC""", (date_str,))
        else:
            c.execute("SELECT * FROM cancelled_orders ORDER BY cancelled_at DESC LIMIT 100")
        return c.fetchall()
    
    def get_cancelled_orders_report(self, start_date=None, end_date=None):
        """Get cancelled orders report with date range."""
        c = self.conn.cursor()
        if start_date and end_date:
            c.execute("""SELECT * FROM cancelled_orders 
                         WHERE DATE(cancelled_at) BETWEEN ? AND ?
                         ORDER BY cancelled_at DESC""", (start_date, end_date))
        else:
            c.execute("SELECT * FROM cancelled_orders ORDER BY cancelled_at DESC LIMIT 200")
        return c.fetchall()
# ═══════════════════════════════════════════════════════════════
#  AUTH DB
# ═══════════════════════════════════════════════════════════════
class AuthDB:
    def __init__(self):
        _ensure_hidden_dir()
        self.conn = sqlite3.connect(AUTH_DB_FILE)
        self.conn.row_factory = sqlite3.Row
        self.conn.cursor().execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT DEFAULT 'cashier'
        )""")
        self.conn.commit()
 
    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()
 
    def signup(self, username, password, role="cashier"):
        try:
            self.conn.cursor().execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                (username, self.hash_password(password), role))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False
 
    def login(self, username, password):
        c = self.conn.cursor()
        c.execute("SELECT * FROM users WHERE username=? AND password=?",
                  (username, self.hash_password(password)))
        return c.fetchone()
 
 
# ═══════════════════════════════════════════════════════════════
#  LICENSE SYSTEM
# ═══════════════════════════════════════════════════════════════
 
def _get_machine_id() -> str:
    raw = str(uuid.getnode())
    return hashlib.sha256(raw.encode()).hexdigest()[:24].upper()
 
 
_KF1 = "Zn9kfm4dA3YWfBBiYx0UBhFrGQwKbnJrHHI4fXt/HRElEwUQLjgyNy1YGAtHAnQeCQoUChMGEAIOfnASe3lwdTgTDBMaFwgnexd/d2dxQEo="
_KF2 = "agVsCSYtDAMcDXJYGCEZXWExBRoNMzp3CwMMEDExMil9aCQ3Ayh3EDMRYhMKJXMEM3p0KXMbbjEwDRcaOXAEKAFuW14IJzoaKhkP"
_KF3 = "aUMaYzYLBTlpAgEBHgQycGEqA3U1Hh0gZgsxdipwKhQbSnYfeQAACzUxD3MTKG1mcDByIC0WNRY5di0UDG43bzQrSUkHDA=="
_KF4 = "AnF3E2EjamcJORRhKzNpXmMDHQl9CTEyPiAsGCIAJWUbdSsBFQEFFQoWPQMJGT4cXV0NLi8rDCEhDhoOH0sBNhhnWwY4IwAAIxw2JQYxGwQ2Okh0AA=="
_KF5 = "Mj0ZIHAyBCw9AzQ4Vw0RJTtwGzUzFictKwQJPg16TDMzNAx3HCMdFXBHAQ4zPgFTNAoXDB0YHSk6CDVlAiQHfW84HzALBwMB"
_KF6 = "fAMZNDcoDy8RFhE4FgwAfE43KzVtBBQZK2ciL24CLR5bQVkvNBwWBA4HEEtqcn5kdXV4F2UTABAJFgVyCgIGfmR1HRtZ"
 
_KS = [
    b'KRISC_F1_2026_XOR',
    b'ABU_DHABI_FRAG_02',
    b'BILLSOFT_KEY_SEC3',
    b'LICENSE_GUARD_004',
    b'ANTI_TAMPER_05_KS',
    b'SECURE_FRAG_SIX06',
]
_KO = [2, 0, 4, 1, 5, 3]
_KI = "0DFF0AC1CBB2F0F6FC1163A66F78D438254A21958F1A1C6BE7A2CFFECEC702E3"
_KB = [_KF1, _KF2, _KF3, _KF4, _KF5, _KF6]
 
 
def _bill_pub_pem() -> bytes:
    raw = []
    for o in _KO:
        xored = base64.b64decode(_KB[o])
        salt  = _KS[o]
        orig  = bytes(xored[j] ^ salt[j % len(salt)] for j in range(len(xored)))
        raw.append((o, orig))
    raw.sort(key=lambda x: x[0])
    pem = b"".join(p for _, p in raw)
    if hashlib.sha256(pem).hexdigest().upper() != _KI:
        raise RuntimeError("Security: public key integrity check failed.")
    return pem
 
 
def _load_public_key():
    from cryptography.hazmat.primitives.serialization import load_pem_public_key
    return load_pem_public_key(_bill_pub_pem())
 
 
def _derive_fernet_key() -> bytes:
    from cryptography.fernet import Fernet
    machine_id = _get_machine_id()
    raw = hashlib.sha256(f"KRISC_BILL_FERNET::{machine_id}".encode()).digest()
    return base64.urlsafe_b64encode(raw)
 
 
def _parse_activation_key(activation_key: str):
    activation_key = activation_key.strip()
    if ":" not in activation_key:
        raise ValueError("Invalid activation key format.")
    expiry_date, _, sig_b64 = activation_key.partition(":")
    expiry_date = expiry_date.strip()
    sig_b64     = sig_b64.strip()
    try:
        datetime.strptime(expiry_date, "%Y-%m-%d")
    except ValueError:
        raise ValueError("Activation key contains invalid date.")
    return expiry_date, sig_b64
 
 
def _verify_activation_key(machine_id: str, activation_key: str):
    try:
        from cryptography.hazmat.primitives.asymmetric import padding as _p
        from cryptography.hazmat.primitives import hashes as _h
        expiry_date, sig_b64 = _parse_activation_key(activation_key)
        payload   = f"{machine_id.upper()}|{expiry_date}".encode("utf-8")
        signature = base64.b64decode(sig_b64)
        pub_key   = _load_public_key()
        pub_key.verify(
            signature, payload,
            _p.PSS(mgf=_p.MGF1(_h.SHA256()), salt_length=_p.PSS.MAX_LENGTH),
            _h.SHA256()
        )
        return True, expiry_date
    except Exception:
        return False, ""
 
 
def _verify_renewal_key(serial: str, machine_id: str, expiry_date: str, renewal_key: str) -> bool:
    try:
        from cryptography.hazmat.primitives.asymmetric import padding as _p
        from cryptography.hazmat.primitives import hashes as _h
        payload   = f"{serial.upper()}|{machine_id.upper()}|{expiry_date}".encode("utf-8")
        signature = base64.b64decode(renewal_key.strip())
        pub_key   = _load_public_key()
        pub_key.verify(
            signature, payload,
            _p.PSS(mgf=_p.MGF1(_h.SHA256()), salt_length=_p.PSS.MAX_LENGTH),
            _h.SHA256()
        )
        return True
    except Exception:
        return False
 
 
def _build_preact_secret_key() -> str:
    machine_id = _get_machine_id()
    salt   = b"KRISC_ACT_2026"
    raw    = machine_id.encode("utf-8")
    salted = bytes(raw[i] ^ salt[i % len(salt)] for i in range(len(raw)))
    return base64.b64encode(salted).decode()
 
 
# ── LicenseDB (Fernet-encrypted JSON) ─────────────────────────
class LicenseDB:
    def __init__(self):
        self._path   = LICENSE_FILE
        self._fernet = None
 
    def _get_fernet(self):
        if self._fernet is None:
            from cryptography.fernet import Fernet
            self._fernet = Fernet(_derive_fernet_key())
        return self._fernet
 
    def save(self, data: dict):
        _ensure_hidden_dir()
        data["last_seen"] = datetime.now().isoformat()
        clean     = {k: v for k, v in data.items() if k != "id"}
        encrypted = self._get_fernet().encrypt(json.dumps(clean).encode("utf-8"))
        with open(self._path, "wb") as f:
            f.write(encrypted)
 
    def load(self):
        if not os.path.exists(self._path):
            return None
        try:
            with open(self._path, "rb") as f:
                encrypted = f.read()
            decrypted = self._get_fernet().decrypt(encrypted)
            return json.loads(decrypted.decode("utf-8"))
        except Exception as e:
            print(f"[LicenseDB] load error (tampered?): {e}")
            return None
 
 
# ── LicenseManager ────────────────────────────────────────────
class LicenseManager:
    def __init__(self):
        self._db = LicenseDB()
 
    def get_machine_id(self) -> str:
        return _get_machine_id()
 
    def get_serial(self) -> str:
        data = self._db.load()
        return data["serial"] if data else ""
 
    def get_expiry_date(self) -> str:
        data = self._db.load()
        return data["expiry_date"] if data else ""
 
    def get_secret_key(self) -> str:
        serial = self.get_serial()
        mid    = _get_machine_id()
        if not serial:
            return _build_preact_secret_key()
        combined = f"{serial}:{mid}"
        return base64.b64encode(combined.encode()).decode()
 
    @staticmethod
    def decode_secret_key(secret_key: str):
        try:
            decoded = base64.b64decode(secret_key.strip()).decode()
            parts   = decoded.split(":")
            if len(parts) != 2:
                raise ValueError("Invalid format")
            return parts[0], parts[1]
        except Exception:
            raise ValueError("Invalid Secret Key.")
 
    def is_activated(self) -> bool:
        return os.path.exists(self._db._path) and self._db.load() is not None
    
    def check_license(self):
        if not os.path.exists(self._db._path):
            return "not_activated", 0

        data = self._db.load()
        if data is None:
            return "blocked", 0

        if data.get("machine_id", "") != _get_machine_id():
            return "blocked", 0

        # Only check clock guard (rollback detection), remove the last_seen check
        if _cg_check_rollback():
            return "blocked", 0

        try:
            expiry    = datetime.fromisoformat(data["expiry_date"]).date()
            today     = datetime.now().date()
            remaining = (expiry - today).days
            if remaining < 0:
                return "expired", 0
            elif remaining <= 7:
                return "warning", remaining
            else:
                return "ok", remaining
        except Exception:
            return "expired", 0
    
    def activate_license(self, activation_key: str):
        machine_id = _get_machine_id()
        is_valid, expiry_date = _verify_activation_key(machine_id, activation_key)
        if not is_valid:
            return False, "Invalid activation key."
 
        today  = datetime.now()
        serial = uuid.uuid4().hex[:16].upper()
        record = {
            "serial":      serial,
            "machine_id":  machine_id,
            "first_login": today.isoformat(),
            "expiry_date": expiry_date,
            "last_seen":   today.isoformat(),
            "status":      "active",
        }
        try:
            self._db.save(record)
            _cg_update()
            print(f"[License] Activated: serial={serial[:8]}... expiry={expiry_date}")
            return True, f"License activated!\nExpiry: {expiry_date}"
        except Exception as e:
            return False, f"Activation failed: {e}"
 
    def renew_license(self, renewal_key: str):
        data = self._db.load()
        if not data:
            return False, "License record not found. Cannot renew."
 
        serial     = data["serial"]
        machine_id = data["machine_id"]
        new_expiry = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
 
        if not _verify_renewal_key(serial, machine_id, new_expiry, renewal_key):
            return False, (
                "Invalid renewal key.\n"
                "Please ensure you entered the correct key provided by KRISC support."
            )
        today = datetime.now()
        data["expiry_date"] = new_expiry
        data["last_seen"]   = today.isoformat()
        data["status"]      = "active"
        self._db.save(data)
        _cg_update()
        return True, f"License renewed!\nNew expiry: {new_expiry}"
 
    @staticmethod
    def is_activated_static() -> bool:
        return LicenseManager().is_activated()
 
    @staticmethod
    def days_remaining() -> int:
        _, days = LicenseManager().check_license()
        return days
 
 
# ═══════════════════════════════════════════════════════════════
#  ROUNDED POPUP HELPER
#  Creates a Toplevel with a thick accent border and no visible
#  square OS chrome. The window bg matches the card color so
#  no raw square background shows through.
# ═══════════════════════════════════════════════════════════════
def _make_rounded_popup(parent, width, height, accent="#FFA500", card="#141414", bg="#0a0a0a"):
    """
    Returns a Toplevel configured as a borderless popup.
    The outer frame provides a thick accent-colored border.
    The window background matches the card so no square is visible.
    Returns (win, inner_frame) where inner_frame is where you pack content.
    """
    win = tk.Toplevel(parent) if parent else tk.Tk()
    win.overrideredirect(True)
    win.configure(bg=accent)          # accent = border color shown at edges
    win.resizable(False, False)
 
    # Center on screen
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x  = (sw - width)  // 2
    y  = (sh - height) // 2
    win.geometry(f"{width}x{height}+{x}+{y}")
 
    # Outer border frame (3px accent border)
    border = tk.Frame(win, bg=accent, padx=3, pady=3)
    border.place(relx=0, rely=0, relwidth=1, relheight=1)
 
    # Inner content frame (card background)
    inner = tk.Frame(border, bg=card)
    inner.pack(fill="both", expand=True)
 
    return win, inner
 
 
# ═══════════════════════════════════════════════════════════════
#  BLOCKED SCREEN
# ═══════════════════════════════════════════════════════════════
def _show_blocked_window(reason: str = "expired"):
    C = {
        "bg":     "#0a0a0a",
        "card":   "#141414",
        "accent": "#FFA500",
        "red":    "#E53935",
        "green":  "#4CAF50",
        "text":   "#F5F5F5",
        "muted":  "#555555",
    }
 
    root = tk.Tk()
    root.overrideredirect(True)
    root.state("zoomed")
    root.configure(bg=C["bg"])
 
    messages = {
        "expired": (
            "⛔  LICENSE EXPIRED",
            "Your software license has expired.\n"
            "Please renew your license to continue using this application."
        ),
        "blocked": (
            "🔒  ACCESS BLOCKED",
            "A security violation was detected:\n"
            "• License file may have been tampered with, or\n"
            "• System clock has been rolled back.\n\n"
            "Please contact KRISC support to resolve this."
        ),
        "not_activated": (
            "⚠  NOT ACTIVATED",
            "This software has not been activated.\n"
            "Please enter your activation key to continue."
        ),
    }
 
    title_txt, body_txt = messages.get(reason, messages["blocked"])
 
    frame = tk.Frame(root, bg=C["bg"])
    frame.place(relx=0.5, rely=0.5, anchor="center")
 
    icon = "⛔" if reason == "expired" else ("🔒" if reason == "blocked" else "⚠")
    tk.Label(frame, text=icon,
             font=("Segoe UI Emoji", 72),
             bg=C["bg"], fg=C["red"]).pack(pady=(0, 10))
 
    tk.Label(frame, text=title_txt,
             font=("Georgia", 28, "bold"),
             bg=C["bg"], fg=C["red"]).pack(pady=(0, 6))
 
    tk.Label(frame, text=body_txt,
             font=("Arial", 13),
             bg=C["bg"], fg=C["text"],
             justify="center",
             wraplength=540).pack(pady=(0, 30))
 
    lm = LicenseManager()
    mid = lm.get_machine_id()
    expiry = lm.get_expiry_date()
    info_parts = [f"Machine ID : {mid}"]
    if expiry:
        info_parts.append(f"Expiry     : {expiry}")
    tk.Label(frame,
             text="\n".join(info_parts),
             font=("Courier New", 9),
             bg=C["bg"], fg=C["muted"]).pack(pady=(0, 24))
 
    btn_frame = tk.Frame(frame, bg=C["bg"])
    btn_frame.pack()
 
    renewed = [False]
 
    def open_renew():
        if reason == "not_activated":
            root.destroy()
            LicenseWindow()
            return
 
        ren_win, inner = _make_rounded_popup(root, 560, 580, accent=C["accent"], card=C["card"], bg=C["bg"])
        ren_win.grab_set()
        _build_embedded_renewal(inner, lm, C,
                                 on_success=lambda: _on_renewal_success(ren_win))
 
    def _on_renewal_success(ren_win):
        renewed[0] = True
        ren_win.destroy()
        root.destroy()
        show_splash()
 
    tk.Button(btn_frame,
              text="  ↻  Renew License  " if reason != "not_activated" else "  🔑  Activate  ",
              command=open_renew,
              bg=C["accent"], fg="black",
              font=("Arial", 14, "bold"),
              bd=0, padx=28, pady=14,
              cursor="hand2").pack(side="left", padx=(0, 16))
 
    tk.Button(btn_frame,
              text="  ✕  Exit  ",
              command=root.destroy,
              bg=C["muted"], fg=C["text"],
              font=("Arial", 14, "bold"),
              bd=0, padx=28, pady=14,
              cursor="hand2").pack(side="left")
 
    tk.Label(root,
             text="Krisc_soft © 2026  RestoBill  All Rigths Reserved  |  Contact support for assistance",
             font=("Arial", 9),
             bg=C["bg"], fg=C["muted"]).place(relx=0.5, rely=0.97, anchor="center")
 
    root.mainloop()
 
 
def _build_embedded_renewal(container, lm, C, on_success=None):
    """Build renewal UI into an existing container frame."""
    hdr = tk.Frame(container, bg=C["accent"], pady=16)
    hdr.pack(fill="x")
    tk.Label(hdr, text="LICENSE RENEWAL",
             font=("Georgia", 16, "bold"),
             bg=C["accent"], fg="black").pack()
    tk.Label(hdr, text="RestoBill-  Krisc_soft 2026",
             font=("Arial", 9),
             bg=C["accent"], fg="#332200").pack(pady=(2, 0))
 
    cbar = tk.Frame(container, bg="#1a0d00", pady=8)
    cbar.pack(fill="x")
    tk.Label(cbar,
             text="📞  (+91) 75 98 70 90 83     💬  WhatsApp: (+91) 75 98 70 90 83",
             font=("Arial", 9, "bold"),
             bg="#1a0d00", fg="#FFA500").pack()
    tk.Label(cbar,
             text="Send your Secret Key via WhatsApp or call to get your renewal key",
             font=("Arial", 8),
             bg="#1a0d00", fg="#886633").pack(pady=(2, 0))
 
    body = tk.Frame(container, bg=C["card"], padx=40, pady=14)
    body.pack(fill="both", expand=True)
 
    status, days = lm.check_license()
    expiry = lm.get_expiry_date()
    if status in ("ok", "warning"):
        msg = f"License expires in {days} day(s) — {expiry}"
        badge_fg = "#fbbf24" if days <= 14 else C["green"]
    else:
        msg = f"License {status}  ({expiry or 'N/A'})"
        badge_fg = C["red"]
    info_row = tk.Frame(body, bg="#1e1e1e",
                        highlightbackground=C["accent"],
                        highlightthickness=1)
    info_row.pack(fill="x", pady=(0, 12))
    tk.Label(info_row, text=msg, font=("Arial", 9, "bold"),
             bg="#1e1e1e", fg=badge_fg, padx=10, pady=7).pack(anchor="w")
 
    sk     = lm.get_secret_key()
    serial = lm.get_serial()
    mid    = _get_machine_id()
 
    tk.Label(body, text="Secret Key  (send to KRISC to get renewal key)",
             font=("Arial", 10),
             bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(0, 4))
    sk_row = tk.Frame(body, bg=C["card"])
    sk_row.pack(fill="x", pady=(0, 4))
    sk_entry = tk.Entry(sk_row, font=("Courier New", 9),
                        bg="#1e1e1e", fg=C["accent"],
                        insertbackground=C["accent"],
                        bd=0, highlightbackground=C["accent"],
                        highlightthickness=1,
                        readonlybackground="#1e1e1e", state="readonly")
    sk_entry.pack(side="left", fill="x", expand=True, ipady=6)
    sk_entry.config(state="normal"); sk_entry.insert(0, sk); sk_entry.config(state="readonly")
 
    def copy_sk():
        container.winfo_toplevel().clipboard_clear()
        container.winfo_toplevel().clipboard_append(sk)
        cb.config(text="Copied!")
        container.winfo_toplevel().after(1500, lambda: cb.config(text="Copy"))
    cb = tk.Button(sk_row, text="Copy", command=copy_sk,
                   bg=C["accent"], fg="black",
                   font=("Arial", 9, "bold"), bd=0, padx=10, cursor="hand2")
    cb.pack(side="left", padx=(6, 0), ipady=6)
 
    tk.Label(body,
             text=f"Serial: {serial or 'N/A'}   |   Machine: {mid}",
             font=("Arial", 8), bg=C["card"],
             fg="#444455").pack(anchor="w", pady=(2, 10))
 
    tk.Label(body, text="Renewal Key",
             font=("Arial", 10),
             bg=C["card"], fg=C["muted"]).pack(anchor="w")
    rk_text = tk.Text(body, font=("Courier New", 9),
                      bg="#1e1e1e", fg=C["accent"],
                      insertbackground=C["accent"],
                      bd=0, height=4, wrap="word",
                      highlightbackground=C["accent"],
                      highlightthickness=2)
    rk_text.pack(fill="x", pady=(6, 4))
    rk_text.focus_set()
 
    tk.Label(body, text="Renewal key covers +365 days from today automatically.",
             font=("Arial", 8), bg=C["card"], fg="#555").pack(anchor="w", pady=(0, 8))
 
    status_lbl = tk.Label(body, text="",
                           font=("Arial", 10, "bold"),
                           bg=C["card"], fg=C["red"], wraplength=440)
    status_lbl.pack(pady=(0, 10))
 
    top = container.winfo_toplevel()
 
    def do_renew():
        rk = rk_text.get("1.0", "end").strip()
        if not rk:
            status_lbl.config(text="Please paste your renewal key."); return
        success, message = lm.renew_license(rk)
        if not success:
            status_lbl.config(text=f"  {message.split(chr(10))[0]}"); return
        expiry2 = lm.get_expiry_date()
        days2   = LicenseManager().days_remaining()
        status_lbl.config(text=f"  Renewed!  Valid until {expiry2}  ({days2} days)",
                          fg=C["green"])
        if on_success:
            top.after(1800, on_success)
 
    btn_row = tk.Frame(body, bg=C["card"])
    btn_row.pack(fill="x")
    tk.Button(btn_row, text="↻  Renew License", command=do_renew,
              bg=C["accent"], fg="black",
              font=("Arial", 12, "bold"), bd=0, pady=10,
              cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
    tk.Button(btn_row, text="✕  Close", command=top.destroy,
              bg=C["red"], fg="white",
              font=("Arial", 12, "bold"), bd=0, pady=10,
              cursor="hand2").pack(side="left", expand=True, fill="x")
 
 
# ═══════════════════════════════════════════════════════════════
#  LICENSE GATE
# ═══════════════════════════════════════════════════════════════
def _check_license_gate() -> bool:
    lm = LicenseManager()
    status, days = lm.check_license()
 
    if status in ("ok", "warning"):
        _cg_update()
        data = lm._db.load()
        if data:
            # Only update last_seen if it's been more than 60 seconds since last save
            try:
                last = datetime.fromisoformat(data.get("last_seen", "2000-01-01"))
                if (datetime.now() - last).total_seconds() > 60:
                    lm._db.save(data)
            except Exception:
                lm._db.save(data)
        return True
 
 
# ═══════════════════════════════════════════════════════════════
#  LICENSE ACTIVATION WINDOW
# ═══════════════════════════════════════════════════════════════
class LicenseWindow:
    C = {
        "bg":     "#0a0a0a",
        "card":   "#141414",
        "accent": "#FFA500",
        "text":   "#F5F5F5",
        "muted":  "#666666",
        "red":    "#E53935",
        "green":  "#4CAF50",
    }
 
    def __init__(self):
        self.lm   = LicenseManager()
        self.root = tk.Tk()
        self.root.title("Activate - Hotel Bill Soft")
        self.root.configure(bg=self.C["accent"])   # accent = visible border
        self.root.resizable(False, False)
        W, H = 560, 560
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
        self.root.overrideredirect(True)
        self._build()
        self.root.mainloop()
 
    def _build(self):
        border = tk.Frame(self.root, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        card = tk.Frame(border, bg=self.C["card"])
        card.pack(fill="both", expand=True)
 
        hdr = tk.Frame(card, bg=self.C["accent"], pady=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="SOFTWARE ACTIVATION",
                 font=("Georgia", 16, "bold"),
                 bg=self.C["accent"], fg="black").pack()
        tk.Label(hdr, text="RestoBill  -  Krisc_soft 2026 ",
                 font=("Arial", 9),
                 bg=self.C["accent"], fg="#332200").pack(pady=(2, 0))
 
        cbar = tk.Frame(card, bg="#1a0d00", pady=8)
        cbar.pack(fill="x")
        tk.Label(cbar,
                 text="📞  (+91) 75 98 70 90 83     💬  WhatsApp: (+91) 75 98 70 90 83",
                 font=("Arial", 9, "bold"),
                 bg="#1a0d00", fg="#FFA500").pack()
        tk.Label(cbar,
                 text="Send your Secret Key via WhatsApp or call to get your activation key",
                 font=("Arial", 8),
                 bg="#1a0d00", fg="#886633").pack(pady=(2, 0))
 
        body = tk.Frame(card, bg=self.C["card"], padx=40, pady=16)
        body.pack(fill="both", expand=True)
 
        secret_key = _build_preact_secret_key()
        mid        = _get_machine_id()
 
        tk.Label(body,
                text="To activate your software, please provide the following Secret Key to KRISC support.",
                 font=("Arial", 9), bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(0, 4))
 
        tk.Label(body, text="Secret Key", font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
        sk_entry_row = tk.Frame(body, bg=self.C["card"])
        sk_entry_row.pack(fill="x", pady=(4, 12))
        sk_entry = tk.Entry(sk_entry_row, font=("Courier New", 9),
                            bg="#1e1e1e", fg=self.C["accent"],
                            insertbackground=self.C["accent"],
                            bd=0, highlightbackground=self.C["accent"],
                            highlightthickness=1,
                            readonlybackground="#1e1e1e", state="readonly")
        sk_entry.pack(side="left", fill="x", expand=True, ipady=6)
        sk_entry.config(state="normal"); sk_entry.insert(0, secret_key); sk_entry.config(state="readonly")
        copy_btn = tk.Button(sk_entry_row, text="Copy",
                             command=lambda: self._copy(secret_key, copy_btn),
                             bg=self.C["accent"], fg="black",
                             font=("Arial", 9, "bold"), bd=0, padx=10,
                             cursor="hand2")
        copy_btn.pack(side="left", padx=(6, 0), ipady=6)
 
        tk.Label(body, text="License Key", font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
 
        self.key_text = tk.Text(body,
                                font=("Courier New", 9),
                                bg="#1e1e1e", fg=self.C["accent"],
                                insertbackground=self.C["accent"],
                                bd=0, height=4, wrap="word",
                                highlightbackground=self.C["accent"],
                                highlightthickness=2)
        self.key_text.pack(fill="x", pady=(6, 4))
        self.key_text.focus_set()
 
        tk.Label(body, text="Format:  YYYY-MM-DD:<RSA-PSS-BASE64-SIGNATURE>",
                 font=("Arial", 8), bg=self.C["card"], fg="#555").pack(pady=(0, 8))
 
        self.status_lbl = tk.Label(body, text="",
                                   font=("Arial", 10, "bold"),
                                   bg=self.C["card"], fg=self.C["red"],
                                   wraplength=440)
        self.status_lbl.pack(pady=(0, 10))
 
        btn_row = tk.Frame(body, bg=self.C["card"])
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="✔  Activate",
                  command=self._activate,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Button(btn_row, text="✕  Exit",
                  command=self.root.destroy,
                  bg=self.C["red"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
    def _copy(self, text, btn):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        btn.config(text="Copied!")
        self.root.after(1500, lambda: btn.config(text="Copy"))
 
    def _activate(self):
        lic = self.key_text.get("1.0", "end").strip()
        if not lic:
            self.status_lbl.config(text="Please paste your license key.", fg=self.C["red"])
            return
        success, message = self.lm.activate_license(lic)
        if not success:
            self.status_lbl.config(text=f"  {message}", fg=self.C["red"])
            return
        expiry = self.lm.get_expiry_date()
        days   = LicenseManager().days_remaining()
        self.status_lbl.config(
            text=f"  Activated!  Valid until {expiry}  ({days} days)",
            fg=self.C["green"])
        self.root.after(1800, self._proceed)
 
    def _proceed(self):
        self.root.destroy()
        show_splash()
 
 
# ═══════════════════════════════════════════════════════════════
#  LICENSE RENEWAL WINDOW
# ═══════════════════════════════════════════════════════════════
class LicenseRenewalWindow:
    C = {
        "bg":     "#0a0a0a",
        "card":   "#141414",
        "accent": "#FFA500",
        "text":   "#F5F5F5",
        "muted":  "#666666",
        "red":    "#E53935",
        "green":  "#4CAF50",
    }
 
    def __init__(self, parent=None):
        self.lm   = LicenseManager()
        W, H = 560, 580
        if parent:
            self.root = tk.Toplevel(parent)
            self.root.overrideredirect(True)
            self.root.configure(bg=self.C["accent"])
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            self.root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
            self.root.grab_set()
        else:
            self.root = tk.Tk()
            self.root.overrideredirect(True)
            self.root.configure(bg=self.C["accent"])
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            self.root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
        self.root.title("Renew License — KRISC BillSoft")
        self.root.resizable(False, False)
        self._build()
        if not parent:
            self.root.mainloop()
 
    def _build(self):
        border = tk.Frame(self.root, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        card = tk.Frame(border, bg=self.C["card"])
        card.pack(fill="both", expand=True)
 
        hdr = tk.Frame(card, bg=self.C["accent"], pady=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="LICENSE RENEWAL",
                 font=("Georgia", 16, "bold"),
                 bg=self.C["accent"], fg="black").pack()
        tk.Label(hdr, text="Hotel Bill Soft  -  Krisc_soft 2026  |  RSA-Secured",
                 font=("Arial", 9),
                 bg=self.C["accent"], fg="#332200").pack(pady=(2, 0))
 
        cbar = tk.Frame(card, bg="#1a0d00", pady=8)
        cbar.pack(fill="x")
        tk.Label(cbar,
                 text="📞  (+91) 75 98 70 90 83     💬  WhatsApp: (+91) 75 98 70 90 83",
                 font=("Arial", 9, "bold"),
                 bg="#1a0d00", fg="#FFA500").pack()
        tk.Label(cbar,
                 text="Send your Secret Key via WhatsApp or call to get your renewal key",
                 font=("Arial", 8),
                 bg="#1a0d00", fg="#886633").pack(pady=(2, 0))
 
        body = tk.Frame(card, bg=self.C["card"], padx=40, pady=14)
        body.pack(fill="both", expand=True)
 
        status, days = self.lm.check_license()
        expiry = self.lm.get_expiry_date()
        if status in ("ok", "warning"):
            msg = f"License expires in {days} day(s) — {expiry}"
            badge_fg = "#fbbf24" if days <= 14 else self.C["green"]
        else:
            msg = f"License {status}  ({expiry or 'N/A'})"
            badge_fg = self.C["red"]
        info_row = tk.Frame(body, bg="#1e1e1e",
                            highlightbackground=self.C["accent"],
                            highlightthickness=1)
        info_row.pack(fill="x", pady=(0, 12))
        tk.Label(info_row, text=msg, font=("Arial", 9, "bold"),
                 bg="#1e1e1e", fg=badge_fg, padx=10, pady=7).pack(anchor="w")
 
        sk = self.lm.get_secret_key()
        serial = self.lm.get_serial()
        mid    = _get_machine_id()
 
        tk.Label(body, text="Secret Key  (send to KRISC to get renewal key)",
                 font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(0, 4))
        sk_row = tk.Frame(body, bg=self.C["card"])
        sk_row.pack(fill="x", pady=(0, 4))
        sk_entry = tk.Entry(sk_row, font=("Courier New", 9),
                            bg="#1e1e1e", fg=self.C["accent"],
                            insertbackground=self.C["accent"],
                            bd=0, highlightbackground=self.C["accent"],
                            highlightthickness=1,
                            readonlybackground="#1e1e1e", state="readonly")
        sk_entry.pack(side="left", fill="x", expand=True, ipady=6)
        sk_entry.config(state="normal"); sk_entry.insert(0, sk); sk_entry.config(state="readonly")
        copy_btn = tk.Button(sk_row, text="Copy",
                             command=lambda: self._copy(sk, copy_btn),
                             bg=self.C["accent"], fg="black",
                             font=("Arial", 9, "bold"), bd=0, padx=10,
                             cursor="hand2")
        copy_btn.pack(side="left", padx=(6, 0), ipady=6)
 
        tk.Label(body,
                 text=f"Serial: {serial or 'N/A'}   |   Machine: {mid}",
                 font=("Arial", 8), bg=self.C["card"],
                 fg="#444455").pack(anchor="w", pady=(2, 10))
 
        tk.Label(body, text="Renewal Key",
                 font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
        self._rk_text = tk.Text(body,
                                font=("Courier New", 9),
                                bg="#1e1e1e", fg=self.C["accent"],
                                insertbackground=self.C["accent"],
                                bd=0, height=4, wrap="word",
                                highlightbackground=self.C["accent"],
                                highlightthickness=2)
        self._rk_text.pack(fill="x", pady=(6, 4))
        self._rk_text.focus_set()
 
        tk.Label(body, text="Renewal key covers +365 days from today automatically.",
                 font=("Arial", 8), bg=self.C["card"], fg="#555").pack(anchor="w", pady=(0, 8))
 
        self._status_lbl = tk.Label(body, text="",
                                    font=("Arial", 10, "bold"),
                                    bg=self.C["card"], fg=self.C["red"],
                                    wraplength=440)
        self._status_lbl.pack(pady=(0, 10))
 
        btn_row = tk.Frame(body, bg=self.C["card"])
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="↻  Renew License",
                  command=self._renew,
                  bg=self.C["accent"], fg="black",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Button(btn_row, text="✕  Close",
                  command=self.root.destroy,
                  bg=self.C["red"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
    def _copy(self, text, btn):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        btn.config(text="Copied!")
        self.root.after(1500, lambda: btn.config(text="Copy"))
 
    def _renew(self):
        rk = self._rk_text.get("1.0", "end").strip()
        if not rk:
            self._status_lbl.config(text="Please paste your renewal key.", fg=self.C["red"]); return
        success, message = self.lm.renew_license(rk)
        if not success:
            self._status_lbl.config(text=f"  {message.split(chr(10))[0]}", fg=self.C["red"]); return
        expiry = self.lm.get_expiry_date()
        days   = LicenseManager().days_remaining()
        self._status_lbl.config(
            text=f"  Renewed!  Valid until {expiry}  ({days} days)",
            fg=self.C["green"])
        self.root.after(2000, self.root.destroy)
 
 
# ═══════════════════════════════════════════════════════════════
#  SESSION TRACKER
# ═══════════════════════════════════════════════════════════════
class SessionTracker:
    def __init__(self):
        self.session_file = SESSION_FILE
        self.current_session = {
            "start_time": datetime.now().isoformat(),
            "user": None,
            "status": "in_progress"
        }
 
    def set_user(self, username):
        self.current_session["user"] = username
 
    def end_session(self):
        self.current_session["end_time"] = datetime.now().isoformat()
        self.current_session["status"] = "completed"
        self._save_session()
        # Temporary debug — remove after confirming it works:
        print(f"[Session] Saved to: {self.session_file}")
 
    def _save_session(self):
        try:
            sessions = []
            if os.path.exists(self.session_file):
                with open(self.session_file) as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        sessions = data
            sessions.append(self.current_session)
            with open(self.session_file, "w") as f:
                json.dump(sessions, f, indent=2)
        except Exception as e:
            print(f"[SessionTracker] save error: {e}")
 
    def get_last_successful_run(self):
        try:
            if not os.path.exists(self.session_file):
                return None
            with open(self.session_file) as f:
                data = json.load(f)
            if not isinstance(data, list):
                return None
            completed = [s for s in data if s.get("status") == "completed"]
            if not completed:
                return None
            last     = completed[-1]
            end_time = last.get("end_time", "")
            user     = last.get("user", "unknown")
            if end_time:
                dt = datetime.fromisoformat(end_time)
                return f"{dt.strftime('%d %b %Y  %I:%M %p')}"
        except Exception as e:
            print(f"[SessionTracker] read error: {e}")
        return None
 
 
# ═══════════════════════════════════════════════════════════════
#  SPLASH SCREEN
# ═══════════════════════════════════════════════════════════════
def show_splash():
    _ensure_hidden_dir()
 
    if not _check_license_gate():
        return
 
    splash = tk.Tk()
    splash.overrideredirect(True)
    sw = splash.winfo_screenwidth()
    sh = splash.winfo_screenheight()
    splash.geometry(f"{sw}x{sh}+0+0")
    splash.configure(bg="#0a0a0a")
 
    _lm = LicenseManager()
    _lic_status, days = _lm.check_license()
 
    frame = tk.Frame(splash, bg="#0a0a0a")
    frame.place(relx=0.5, rely=0.5, anchor="center")
 
    try:
        from PIL import Image, ImageTk
        import io, base64
        _logo_data = base64.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/4gIoSUNDX1BST0ZJTEUAAQEAAAIYAAAAAAIQAABtbnRyUkdCIFhZWiAAAAAAAAAAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAAHRyWFlaAAABZAAAABRnWFlaAAABeAAAABRiWFlaAAABjAAAABRyVFJDAAABoAAAAChnVFJDAAABoAAAAChiVFJDAAABoAAAACh3dHB0AAAByAAAABRjcHJ0AAAB3AAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAFgAAAAcAHMAUgBHAEIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFhZWiAAAAAAAABvogAAOPUAAAOQWFlaIAAAAAAAAGKZAAC3hQAAGNpYWVogAAAAAAAAJKAAAA+EAAC2z3BhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABYWVogAAAAAAAA9tYAAQAAAADTLW1sdWMAAAAAAAAAAQAAAAxlblVTAAAAIAAAABwARwBvAG8AZwBsAGUAIABJAG4AYwAuACAAMgAwADEANv/bAEMACAYGBwYFCAcHBwkJCAoMFA0MCwsMGRITDxQdGh8eHRocHCAkLicgIiwjHBwoNyksMDE0NDQfJzk9ODI8LjM0Mv/bAEMBCQkJDAsMGA0NGDIhHCEyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMv/AABEIAWYBZgMBIgACEQEDEQH/xAAcAAEBAAIDAQEAAAAAAAAAAAAAAQYHAgQFAwj/xABWEAABAwMCAwUDBAsKCwgDAAABAAIDBAURBiESMUEHE1FhcTKBkRQiobEVIzZCUmJzssHR8BYXJDU3Q3KEk+EzNERTY3R1ksLS8QglJlZXgpSVRVWi/8QAGwEBAAIDAQEAAAAAAAAAAAAAAAEDAgQFBgf/xAA1EQACAgIABQIDBAkFAAAAAAAAAQIDBBEFEiExQSIyE1GBBhQzYRUjJDRCU5GxwVJxoeHw/9oADAMBAAIRAxEAPwDRqiIumAqoiAIqogKoqiAIoqgIqoiAIiICqIqgCiIgKoiIAiqiAqiIgKiiqAiIiAIqogKoiIAiKoCKqKoCKoogKoiICqKqIAiKoCIqiAKKogCIiAiKqIAqipa4M4+Elp2B6ZUEbIoqikkiIqgIqmwwSQB59fpULmg4BHxH61G0AicTfEfR+tMt/CH0frTaBUTIxzb8Rv8ASnx96bARRVSCIiICqKqIAiIgCqbbA9Rtn/qmw24h8R+tRsEVU2/CHxH61cj8JvxH61ICJkfhN+I/WmW+Iz6j9aAKK/H3jCiAIqiAiqiqAiKqIAiqIAiiqAiKogIiqiAIiqAiIigHIDbYZJ2Hr+2FkeqLc20U9poQPnfJzLIfF7ic/AYHuXmWGnbW32hpnjiY+VoLfHdZX2nxFt3oZSNnwuGfR2/1rJLps0Lb9ZddPh7ZgZ2OPBRXmMnmuTGOeHODSWs3cR0GQMn4rE3kcFVOm4weqKfJJszSVXS2Tspu18dZbVcaqK5shb8vpRIA0sHI810P3zZOmitGEf7Ld/zLlbv5B78cDP2Yh6fiNWAuA4nbDmVTGCbewZ3++ZJ/5J0Z/wDVO/5k/fMk/wDJOjf/AKp3/MsDwPAfBTA8B8Fl8OINyaD1PTau1GbTXaS0vBC6lleX0tuLXgtaSMZJWnN+ox5Yx9C2D2M/d9/UKn8xa+xjIHLKiC1JgIiqtBEREAREQFX1pqaWsqIqemjdLPK4MjjaMlzydh6L5bEc8EfD3rZWlKeDQumDrS5RNddKoOhs1K7mDyMpHkc4ysJS0gdi41Vi7NqOksbrFar5e+HvrjLVs4mwuI2jHoCF5n749tAwez7Sx8xAsFqaqoq6uapqJXSTzPL5Xn75xOT9K+OT4rFVrXUGwf3yLZ/6e6W/sFP3yLX/AOn2l/7Ba/RZfDh8gZ+e0i1Dc9n+mP7D6Vls14s9B2fOv110bYKOurCY7ZSxUwLnjcGQ55YKwbQGloLzVz3a7uMOn7WO/rJjtxEbiMeq8zWGqKjVV/krnt7qmYO6pIByiiAw0D1G/vVXKnLSB4LiS8khoJOSGjAB8guKqi2AFVFUBERVARERAEVUQBERAEREBUURAVFEQFURUISu57mjQDq+2Z/zqzftPohLaKWuxnuZSw+QcAsD0vMINUW2Q7ATtHx2W5NQW0XWwVtHjJfGSwfjN3CzivSzyvFbvgcRosfb/s0L09NlmHZtQRXnUNZZZGtzcrdUU8ZPR4bxtPxaPgsPPE1xDhhwOCPAr3dEXUWTW9luTvYiqmB/9F2Wn6CqZe16PUp7WzxJ4JKaolgmaWyxPLHg+IOCuC2Z226WFk1iblTMxR3Qd6COTZPvhnz5+9az6nH/AEUwlzJMGfW7+Qe/f7Zh/MasBd7TvUrPrd/IPfv9sw/mNWAu9p3qVjX5AURFYDYHY19339QqfzFgHU+q2B2Nfd9/UKn8xa+6n1Krj7n9AVFEVgKiiICqIu/ZbNWagvFNa7fF3lVUPDWjo0dXHyChvlWwZFoLS1PfK2e43VxhsNsZ39bNy4sbiMevx3Xn6x1RPqq/PrC3uqSJoio6ccoYgMADzPPx3Xv67vFHa6CDRFikD7fb3cVbOw/41Uffb9QCSMctlr8+0T5quK5nsBVFFaAvU0/YqzUl6pbVQRl1RUP4c42Y3q8+i8xoLnAAEknAaObj0AWz58dmejfkbMDVV6i/hDg7/E6c9PIkYPjusJS10B5mu75SUlJBouwPP2Ltrj8plb/lVR98SeZAcSMctlgZyXEnGc9FXEk5znrnx81xUxjyrQKiiLIFRREBVEVQBFEQFRREBUURAVFEQBFVEAVUVQERVRAfWnlfBNHNGcPjeHA+BGD+hfoqKVsscUzQMPaHD37r84bnPzsDG/0reei7h8v0pQyvPFI1ndP9WkgfRhZ199HlvtTTumFq8P8Auaz11Z/sTqSfgZwwVH26LPmNx8crGcHhIHPO3ly/uW6de2P7M2IyxtzUUf2xviW9QtLkA7+O6xktPR0+DZiycRSk/UujP0PN3faZ2T2yne9vy+SDggkP3tXCDkH+m0fSvzzMx0Uz43sLHscWvYRgtcOY9x2Wy+ya7SzG4aYbOIp6rFZbnnA4aqMZx/7mjC63aTZo61kGs7fD3dLXuMNfEBg01W3Z4I6AkH4rWg+SWmdY+Nu/kGv3+2YfzGrAXe071Kz63fyD37x+zMOf9xqwF3tO9SrK+7BERVWA2B2Nfd9/UKn8xa+6n1K2D2Nfd9/UKn8xa+6n1Vcfc/oAqiKwERFceRUADz+jmtnR47MtGceW/urvcR7sY3o6Y9fJxG/juvO0HZaKho6jW9+jP2JtrsU0J/yuo+9aPIHcrE77ea2/3uqulwkL6md+Tnk0dGgdABt7lW/XLp2B57nHiJ4s5++8fNcVVFaAqByz+wRZLonSj9VXoxSP7i3Ure/r6p3sxRDc+8rFvS2D3dDWmksVnl17fYw+kpTwW2neD/CajofQH3bLCbvdKy9XaruVwk7yqqZC+Qnp+KPADljyXv651WNR3eKGhjEFloG9xQU/DgBg24yPE88nfdYp7sLGC36pAKKqKwBFUQEVREBFVEQBERAERVARVRVARFUQBRFUBFURARERAXphbJ7K7l8+stj3bYE0YPwIWtl6+mbn9idQ0dUT9ra/hkH4p2KJ6Zo8SxvvGLOvzo34W5aW42IOR4rSGtLCbFf5Wxt/gkx7yE+XUe45W8A7iAcDkEZBHUcwvA1fYRf7JJCwD5TF9shPXPUe9WzW0eF4HnvDyeWftfRmk6GuqLbcKeupHmOop5BLE4c8tIK3dVXO21JiutQP/Cerou4rsf5HWgYEniNxnoFomRjo3ujeC1zThwPMEc1n3Z1c6SuhrdGXiRrbbeBiCRw2gqQPmkeuy07I76o+kJ77Hr3XT1Xpbsn1Laq0AmO9QujlHsyxljeFwPmMFapPM55rZ+oNU1jezu46Iv5c272urjbE5w3miDts+4jHlhawPM+qmnswERRWg2D2Nfd//UKn8xa+6n1K2D2Nfd9/UKn8xa/6n1Kwj7n9AFEVWYIsg0fpap1bfYrfETHTtBlq6g+zBCPacfPmvFpqaWrnjp6eN0k0rxHExo3c88gti6kmj0BpX9x1BIx94rQJbzURu3YObYh4bEZwsJy8IHja91RT3eup7RZ2d1YbUzuKOIcnkbOlPiTjqsO5806AcsITusox5VoBRVD/AHk+A/Wsgdm32+qulfT0FFC6WqqHiOKMDdzjt8BzWe6zr6TSNgboWzTB8nEJrxVRu3mlx/gx5DYYHh6r72mIdmukxfapmNTXZhZbYHc6aE85T5np7lrR73ySukke58jiXF7tySeqq98tg4nOd+n0eSKFVWgKKqICqKogIqoqgIqoiAIiICqIiAqKKoAiIgIiqiAIiqAiIiAKjx6DOVFQoBu/Ql4+y+m4eN4M9P8AaZMf/wAn4YWS7H0ytMdn98+xN/bBK77RV4jdnkHdCtzgFux3I2z4q6D2j5tx3D+7ZbaXpl1X+TVPaPpr5HVi80zf4PM7EzQPZd/esCBc0gtJa4EEO6tPPIX6MraSC4UUtHUs44pW8DgfA9QtD3+yT2C6y0UzSQ0kxu/DafBYTiel+z/E/vFPwbH6o/8AKM3vDB2i6JZqGBjXais0YiuUbQMzwAbSjxIGFrM4ztuOh8Qvb0nqWp0pf4LnTgPYz5k8B5TRH2mkHyXs6+0xS2+aDUFkcJbBdcyQPGSIH/fRk+RyPcqF6Xo9IYUipG+Pj6qKwGwOxr7vv6hU/mLX/U+q2D2Nfd9/UKn8xa+6n1Kwj7n9AFRjG5x1OeWFFmGgtKwX6tmr7o7ubDbW9/XTnbONxGPX4rKT5Qe1pOmg0Pph2trlGDc6hroLNSvG5cdjMQfDBAWuqqpnrauaqqZXS1Ezy+R7ubnZ3Xu6y1VNqu/vrC3uqOICKjp+kUQGAB5nn71jpWMI+WAqoisByDc7DPPwWb6B07RyMqdVX4cNitR4yD/lM43bG3xHivC0npqp1ZfYLXS4bxHinmI2hiHtPP0hbAv1LSapoobPYqw0ths8pgha2In5RIAC6V2++5KiMJ3S5Idyuy2NUeafY11qbUVZqm+1N0rDh0pxHGOUUY9lo8MDC8jKzz97fP8A+UP/AMc/rT97c/8A7M//ABj+tbUcG5L2mt+kMf8A1GBqrL6vs9r4ml1LUxT4HJzSw/BYtVUdRRTmCpidDIPvX9fRV2U2V+5GxVfXb7Hs+CKgZ9fRZvT9nff08Uv2TLe8aHY7gnmB5pVTO32Ii7Irp1zvRhCi7l0ofsbc6ij4+PuncPFw8OfcuoBkgAEknAAG5Vbi09MtUk1tdiKrKbZoS4VsbZqp7aWIjIBGXFe23s7ouHDq6cnxDQAtmGHdJbSNSedRB6bNdIs3ruzydrS6gq2ykDPBI3BPvWH1dHUUNQ6nqYnRSt5hw5+iqsosr9yLqsiu32M+CIvrT081XO2GCN0kruTGDcqpLfRFzaXc+SLNKDs9qZWNfXVLYSR7DBl3vXons6oizDa2oDvEtGFtRwrpLejSlxDHT1zGulVldz0HXUcbpaSVtUxo3bjDx6BYqWlri0jDhzadiFTZVOt6kjYquhatweziiqKotCKIpBURRAVREQFRFEBya5zHAtdg5BB8MFb10lfBfbBBMXD5RGBFMOocNgffjPvWifBZRoW//YS+MbI4fJqkiKTi5A9Cpi9PqcbjmB97xm4r1R6r/KN2A7rwNW6bj1HaixoDayIZgf1J/BXu5GARuDyPj5q5GVa47R88ousxrVbDo1/7R+cZ4Zaed8MrSyVhw5p6HwWZaF1LSQxVOlr+7i09czwOc7/JJTykHUdPJe92gaUNZE670DB3zB/CGAe038JaqJzkdCMbrXlHa0fTsDOhmUq2Pfyvkz29VaarNLXua3VYDx7cEzd2zxHcOafTHvXiLZOmrhS67sMWjr1MIrnBk2aud48+5cfA+awC42+qtdxqKCuhdBVQOLJI3bcJ/V4LGD/hZvGbdjX3ff1Cp/MWv+p9VsDsZ+78Zz/iFTz/AKC1+0ZGTn9aR9zB37NZ6u/XeltVDH3lVUvDGDo3xcfILMNeXeitNug0RYpQ6goXcVdOw/4zUdd+oBJGOWy78Ib2ZaN788I1Xe4ftQxvR055uPgTz8d1rJziXHJ4id+I9fP3qF62CH2iiiKwFXOKGSeVkULDJK9wYxgGS9x5ALgPPl4+AWydI0dNovTjtcXaE/LJSYbPSuG73n+dI8jke5YyloHO+Sx9nOlDpeikadQXFokutQznDERlsQ8NiDtvuuPZ7g6fmOAP4S7p+K1a7raupr66errJXS1Mzy+V7zuXft0WxOz37n5/9ad+a1b3DY6tOfxP8Bnc1NqJ1gNNimZOZ+P2nYxjH614De0V4522M+OJsfoX07RyQLZgn+d6/wBBYJkkcz8VflZNsLXGL6FOHiU2UxlJdTatj1bRXmX5N3ZgqDu1jiCD6FffU1liu9rlHCPlMTS6N2N8jmFq+0ue28URYTx98zGPVbq5nfHmtnGteRW1M1culYtsZVmisb8sHOfRbtoQPkFLsP8ABM6eQWlqkj5ZNw+z3hx8Vuqg/i+l/JM+oKjhy1KRfxV7jBmptTfdJXflFkGg7PHOZrnO0O7t3dxNI68yVj2p/ulr/wAp+hZxoKVj9OujB+cyZwcPUbKvHipZL3+ZdlTlDEXL8kepfr7TWOlE87S+R5wyMHcrDpO0O4cfzKOmDc+y4uJ+OV6Wv7bU1EVNVwsc+OEOD2t3Iz1WvR13yAs8zIthZyroivAxaZ1c7W2bOsGsYbtUCkqYBBUO9gg/Nd8V3NUWSO8WuTDQKmIF8bwN9ui1XR1Bpa2Co3+1SB+WjfC2Mdf2l38zV7/iDCzpyY21uNzK8jElVbGdCNZ8JJwAeInZvn4LbGlrBHZ7fG97QayUBz3EbtBHILA9O0sdx1ZAwNPdd6ZMOHIDdbRr6oUduqao/wA3GX/qWGBVFKVj8GfEbpPlqj5PEv8Aq6ms0hp4o+/qfvhnZvqsdb2hV/eAuo6QtzybnKxKWZ9RK+aVxc+QlzieuVwJ6dFRZm2yluL0jZq4fTGCUltm4LFf6e+07pIWmOaP2487hY1ryyRsYLtTtAOeCYAc/ArwdIXFltv8L5ZGxwSAskLjgYxss5vN1tFZZayn+X073vhPCA774DIW2rY5FD5+5oypli5Kda6M1T5Ih58x7kXIO718BRVEBFURAFFVEBUREBFc7eHn4eaIoBuTQmoReLOKWd4NXSgA/js6H3DCysnwWgrFd5rJdYq2Fx+Yfnt6FvULedJWwV9HDV0zuKKVvECD8R7itit76Hz7j3Dfu93xIL0y/ud0O+bghpHUHqFqXXmj/sXUOuVE1xo5TmQD+bctrF2dh6L5zRxVEL6eZjZIpBhzXdQkobNDhmfPBuUl2fdH54ZI+IhzHPYWuDmvbnIPQjC2fAYO1extpp+7h1lQQ/apD80XCIdD04hyzz2WJ6v0tJp+uD4gTRSuzE/8H8UrH6asqKCqiqqSV0NRC8PjkbsWkeHktWcD6VTkQvrVkHtMyPRd+Gi9YMrq2lke2IPp6qHk9gcMOwPJZpRaK0/pyol1s+4x3DTFMwTUMYPz5p/vYXDpg59cLq1VNTdqtqdcqFkdPrCkjzVUrfmiujH37B48sjxytZvlqYoTRSSSsjZIS6EkhrXjY5b4qtrmf5lx279e63UN7qrpXycU9Q/JHRjejB4ADA9y8w81UVyWloERF9IY2yzRxue1jXuDXSOPzWAnGT9anwDK9BaVh1BXy1t0f3Fhtre/rqg7bDcRjzK6ms9VTarvjqnh7migHc0VO3lHEBge88/esq7SHO0xZLXpC2ROjtJibVyVYORXyEb4PgCTty2WsySeePcqormewcTz55Wy+z37n5v9ad+a1a1wtldn33Pzf60781q6WB+Mc/iX4DPrq+wVl8FH8kMY7nvOLjdjnw4+pYx+4G8fhUoHnIs5vWoKWw9waqOd/fcXD3bQeWPH1Xm0+u7VPURw91Vxl7g3LwAAT4rduqx5WPnfU0Me7KhUlWuiPlp7RjLXVMrKuVs07PYaz2WnxXq6iu8dmtckhcO/kHBC3xPivWk4yxwicO8IIYTyzjbK01eaiuqLlMLhIXzRuLSOWB5BTfKONXqC7mOPCWZbzWPseeOe/PK3fQfxfS/kmfUFpELd1D/F9L+SZ9QVHDe8jZ4t0UTU2pvukrvyn6Fz03fpLFXF5bx08gxKzy8Vw1N90lf+UPL0Xk891pTnKFrlHwzoQhGymMZdmjdlDcKS6U4kpJWzNcN25GR6hedcNKWi4lz5KfupDzfEeE59OS1TT1M9LJ3lPK+J/iw4WSUOvbpTFrakR1TRtlw4XfQt6OdVYtWo5s+H21vmokdq49n1TEHSW+ds4G/A8Yd8QsPqKeWkmdDURvjkGxa8bra1m1VQXh/cjihqP83J19Fy1LYYbzbpMMAq4wXRvxucdPNLcOuyHPSyac62ufw70YZoAf8AiM5HKB2PiFmmqyRpevI2+YB8SsK0I7g1KGu2c6J7cHx5rONTxmXTNe0DJ7vPwKzxf3aX1K8z97j9DUCio3GUXGO7oiAAdAuTI3SvDI2Oc92zWtGSV93UFYAXGknAAyfmHZZKLfVGLlFd2dYor6oo2ZEREQBFUQERVRAEVRAEURAXJCzjs/1IKGp+xdW/FLMftTjyY89PQrBlyaXNGRkeY6eaKTi9mvl40MmqVU+zP0SSQSfqU4ljOir3PdrEDUgGWB3dl4+/GNlkQdsMrfj6ls+ZZWO8e6VUvBxraSnuNFJSVUQkikG4P1haZ1LpufT1cWOBdSuOYpfELdPHthdS40FNc6B9JVs44nDbxafEKudO10Ohwjic8Ken7X4NHUNwq7VXQ1tBPJBUwO445GnBBHT08uS2PVUNF2p22S52uGOm1bTszWUQIArWj7+MeOwzjG+Vg+otO1VhrTHKC+F5+1S/euC86hr6u2VsVZQzvp6qB3HHKw4II8PLy5LRnCW9+T6DXbC2Ksg9pnykjkjldHIxzHsdwva4YLXeBC+efULaj4bb2sUvyimZT0GsomjvIXHhjuAH3w8HemN1rKrpZ6KplpamF8NRC7hkjeMFp8MJGSfRlh8FeI4x08DyU92EWYNiaOvlHf7QNE6mmxSzP/7srnnJpJujc/gnwO26w6/2Ou03eai13GF0c8Jxy2e3o5p65GD715oJA26/t8Vs6zVcHaVYGabuk7Y9R0bSbXWSbd+wD/AvP0A89lTJOL2uwNYnn09y2X2ffc/N/rLvzWrXdbRT2+vmoqqF8FTC/u5InDBY4c/d+hbE7Pvufm2x/CnfmtXS4c92nP4l+Azzu0jB+xmw/nf+BYJ08P1rPO0jlbPWX/gWB9Fjmr9eyzh/7vE2to+8fZSztZI7NRTgMfg8x0PwwvF19ZtmXaJu2zJsfQVjWmru6z3mKYn7S/5krehaVtqaKKrpXxPAkikbyPJ3UfQt+prJo5H3Rzb08TJVkezNHA77HO63fQfxfS/kmfUFrDV1hFhvAbFvSTsE8BPUcnN9xz8Fs+hGKCmHhEwZ8dgquGpqU0/BbxSSlCEl5NTan+6Sv/KH6lsSxUVurLDQz/IaZ5dE0Oc6Jpy4DB6eIWvNTfdJX/lP0L1tH6mjtZNBWOIpXuy1/wDmyq6LIwyJKXkuyapzxouD6pHV1rbfkF8dJHE2OnmYHMDW4aMAA8vMLHfX6VuyopKO60nBNHHUQO5EH9PMLwH6Csz3lzX1LBn2WvGPp3Wd2DJycq30ZXj8ShGChZ3Rr21iU3akEGe971vDw9N91uvcc+eV5Vq05bbS8yUsDjKdi954nD9C6GqNSQWqjkp4ZBJWSDhDW78A8Stiir7rW3Nmtk2/fLUq12MFoK9tv1W2qbtE2ocD/RccLbM0MdVTSQuIMcrCzI6tPVaO9ea2DpDVEUtOy218gZK0YikdycPArWwr47cJdmbXEMaTSsh3RhFyoJrXXy0k7S1zHEDPUdCF1M9M781ue52ahu8YFZBxEey8HBHoRzXjM0DZmP4nPqnD8FzwB9AUT4fPm9HYmvilfL6+5jeg6B1Te/lhae6p2k8XTiPILN9S1goNPVkjjhxZwNHiTsu5T01FaqLgiZHTUzdySdveeq1zq7UQu9Synpj/AAWE5z+G7xV7UcWhxb6mtHmzMhSS9KMY5bZzhERcdHeRVERAVFFUBEREAREQBERAFRsoqofYGzezdxFmqvy4H0LM8rCuzk4s9T+X/QsxDjjmujStwR874yv22Z9MrkHHC+PF5q8ZxjKt5Tl6PlcKGmudG+lq4hJG4bE8x5grUeodN1NhqiH5fTH/AAcw5ehW4OI8sr4VdNT11K+nqoxJE4YLSOXmqrKOZb8nY4VxOeHPlfWL8Gj4p5qWdk9PJJFLG4Fr2EhzTvuCN1smG42ntQpYaC+Sw2/VUbQymuJbiOrwNmS9AeW/PksT1JpaeySulizLRu3bIB7PkVjuSBgZH7c/71zrK9Pr3PdU3wvjz1voz0L1Zbhp+6S2+6Ur6apiO7HDZ48WnqOq88gdDkLYVl1lbtQW2DTuuA6Sma3gpLqzeamPQOP3w9c7Lw9V6KuOlpY5pOCpts5zT18HzonjpkjkfHzWEZ9dSLjGVzillhlZJDI5kjHBzXDILT4jC4HYqdcrP8gbSqoou1XTzrhBHGzWNthzUwtaB8viG3G38Yft0WA0F+utnifS0dQYWcZLmGNpIdyPMeS+FqulbZrnT3G3zGKpp3B0bh5b4PkfDks81LaaLWthl1pp+ER1cX8b0DecZ/zrR4EYOBssITlU+jMZRjJaktmC3O9XC7918un73us8HzGtxnGeQHgF0FTjOxBHQhRWSk5PbexGKitJBe3TarvdLTsghrSGRjDAYmHHhuRleIuzQ0VRca2Gio4ny1NQ8Rxxt5ucTge5SrJQ6p6IlCM/ctm27jZX6m7BaC+yYkudFLNO5wGDJEJSx2AOgAb06LW7NYX2JjY2V3zGANb9pZyH/tW2YbvSWDWOlNEOma+lp6GS33EMI4TLOMkZ8nYWmL1bZLPfa+2S5L6SokgJPXhcRn34yq6rpqT6vqQ6oNaaXQ69TUzVlTJUVD+OWQ5c7AGT7l8s9OioaScD4LjzWb6vbM1pdEd+hvFxtpBpKuSMdG5y34HZew3Xd6aMF0Lz4mPf6FjCqsjdZHomVSoqn1lFHuVmr71WMLX1fdxnm2Nob/evELnPcXvcXOO5Ljkrj1RROyU/czOFcIe1aCZP05RFgZnsUOqLxb2COGrc6P8ABkHEPpXfdr29ObhroGnxbHv9KxhFar7EtKRQ8aqT24o79ferjcz/AAurkkH4Psj4BdBEVcpOT3J7LoxjFaitBERQSEVUQBVEQBREQFRREAVUVQEVUVUA2R2dnFoqfy/6FmAdsFhnZ8cWiq/Lj81Zdxcl1sdbrR8/4wv2yZ9cpxL5cScSu5Tmcp9OJOIj3r58RynEQmmSo6OUjGSxOikY17XD5zHciPRa71JoySjL6u3MdJT83Rcyz08Vn5dunGcnfmsLKIzR0cLNtxJbg+nlGj+WRnyKyvSuuq7T0TrfURNuVklz39unwWnPMsJ3afTG69rUGkYa7jqaENiqeZZya/8AUsAqaaWkmMNRG5j28wRuPRcm2hweme1xM6rKjzRfX5Gd3LQttvlvkvWhKk1lM0cU9rlP8Jp9+TR98P0LX8kbo3vY5rmvacFjhhzT4ELs2y519nro6221c1LURnLZYyR8cc/etgjUOmNetZBq2BlovWOGO80rR3cp/wBK3kPXC1+sfzNw1oefLHkva0tqet0rfIrhRFr/AL2eB3szRnm1wOxXe1JoK96ab8omhbWW5wzHX0ZL4XDxJ6LFs+YI+IVi1JA2PNQ9meo6iSoprxcNPVExLzBUwB8LHHfAI3x71139m1vl3o9f6Xc08u/qjCf90gkfFYCSTzJPqVMA8wD6rDka6Jg2Ezs5stMA+66+sEcY3cKKQzE+h/uX1Gq9NaMgkj0VTT1lyewsdeK1oHAP9Gw7Dwzha3wPwR8FyJLt3EuP426nkb7sHYjrp4riyvdK51Qyds5kPMuDs5WadsVMyLtGqqiIAMrKeGq/3mAfWFgjWOle1jRxOcQ0DqSSBhZ/2xuH7toKcn59NbaaJ/kQ3iI+lO01oGAxvdE9rwNwcj3fsV2rlTNpax3CPtMgEkZ/FIz9C6eCevVZFVQfLdK0lSBmWnBDvNoJC26q/iRkvK6lNtnw5Rfz6GOKqcseiKhdi4KqIgKoiICoiiAKqKoAiiICoiiAKoiAKIiAIqogKiiIAqoqnkGw9AHFpqfy/wDwrLOJYhoI4tVT+XH1LK+JdnFW6ong+LLeXM+nEhOy+eU4tithxOcl1PoXBjS5xw0DJPkvnDUx1MLZonh0bxs5q4yuJhd5ArWNj1JUWibhdmSlPtx9R5ha91yrklLszpYfD3lVTcfcjaRKnEutR11PcKZtRTSB7D58vVfUnzWxHTW0aMqpRbjJdUci8rz7paaW7xGOdmHj2ZRzC7nEuJcplCM1posqnKp80HpmtLxpystDy4tMtPzEzBsPVeNyJG2PDmCtwuIc0scA5p5tcMg+5Y1dNI01XxS0R7mbmWk/NK51+A11genw+MqSUb+/zPG05re+6ZPd0lV39G7aSiqftkLx4Frsge7CyPOgdaO4pHSaWuz/AAHFSPd788PuwsDrbdU2+XgqYi3zI+aV1snruPB24XLlU0/kzuRlGa5ovaMwvXZlqS1xOqoaRtyoQMtq7c/vWEeJHMLEJGd2/hd809Q5pafpXp2jUl70/KJbVdKukPgx5Lf907H3hZY3tVnrQGX/AE7ZLwTs6R8AjlPnxNAWPqRJr7Hl8DlDgdWt/pH9ithO1L2cT4dPoSqif1EFwc1vuGVRrfSFrAksehIBMOUtxqHTAeeM7qeeWuwPnoHSoEg1XfmGksFtd33FMC11TK3drGDqM9ViOobzNqHUNfd6gYkq5nScP4LTyHuGAu3qPWF41VMx1xqmmCPaKngbwQxjyaMfErwkhFrqyShZtp6IS6ejjeMtk42n3khYT4Y3OcLYlrpvklqp4Xc2s+vc/WuvwuHNZKX5HK4tPlrivOzAaundS1csDhuxxC+K97VVOIrgyYDHfM3PmNvqC8HOdxyWlkV/Dtcfkb+PZ8SqM/mRVRVUlwUREBURRAVFFUARREARVRAFVEQBERAEVUQBEVQEVUVQGe6Fdi11A/0w+pZTxrE9EHFsqPyw+pZOTuu7hx/UxPD8UW8uZ9eNXi2K+OVQ7otnlNDlOT3faX+i03nYLb8riIn4/BK1AuXxFe36npvs/wBIz+n+Tv2u61VqqRLTv2Ozo+jh4FbDtV8pbvADEeGYe3E4/OHp4rVi+sE0tNK2aF7mPbyc3mFrY+TKp/kdHN4dVkrfZm3ifBcS5YxZ9WR1HDT3AiOQ8pRjhd6+CyQuHC08QIO4x1Hku5VZG1bj3PJ5GJZRLlmikqZBGCuJK4FyvSKkhPGyZhjmY2Rvg4ZWOV+kqaYl9HJ3LzvwO9n4lZCXFcS4rGzGrtXqRt0ZNtL3Bmu6yy19CT3tO4sH37N10DnkSSByH9y2gXHl0XRqrVQVme9pYy/8IHhP0LnWcJ/lyOxVxj+Yv6GuyP2xhUbbjmswm0lSyZMUssXhkcQXTdo+YHDauP3sK05cPyY9FHZvx4ljS/i0Y2SSckknxJTGxJWRjSE2fn1bAPJhXo0mmKKnDTMXVDwc9QPgphw+9vqtCfEcdLaezyNP2d9VUMqZmkU7OWR7ZWZE596ABjQ1o4WjkB0TOy7uNjRohpd/JwMvLlk2c0u3g8LVcPHbGy43ifz9Vhx5rPNQM7yx1Q8Gg/SsDPM48VxuKw1fteUdvhM3KjT8M4oiLmnTCKqIAiqiAKqKoCIiICqIiAKqKoCIiIAiqiAqiIgCIiLuDOdFH/uyo/LD6lkvEsY0X/F04/0v6FkfEvRYa/URPGcTW8qZ9MplfPiTPmtrlNDlOUp+1P8A6J+pajW2ZD9pfv8AelamXI4otOP1PR8DWlP6BXooquUd8H3L2LTqKrtmGF3fQZz3bjy9F4ydcrKE5VvcWV2VQsXLJbRsuhvFJcmcUEnzzu5jtnD0XaLlq2OR8Tw+N7muHIg7rIKDVc0WGVre8YP5wD530Ls4/EovUbTg5PCHHrT1MwJXAkrr0lfS17M08zSfAn53wXY2G3VdeEoyW0zkyrlB6ktMio5JhMLPx1MdfIYQ80RRpEkyrlTKKexBSU6KJuUB0b07hstYf9Hj4kLXw5LNtUVAhtLoifnSuA9w3WFZB5LzvFZp3JL5HpuEwapbfzIqoi5h1AiIgKoiIAiKoCIiICqKqIAqoqgCiKoAiiqAiqiIAiIgfYyTTt7pbbSyRTiXic/iHA3O2F641ZbcezUf2awTO2FdvALbrzra4qK8Ghdw6i2bnIzv91lt/BqP7NT91lszyqP7P+9YIis/SVxWuE435/1M5k1XbSx7WifJGBlmAsHx5Y8lFVr35M79Ofg2sfErx01DyFEVVBshREQBXJxzURAcmSPjeHscWuHVpwvao9T1cADZwJwNt9iF4aZVld1lXsZVbRXatTWzPKTUNvqgGmUxP/Bf+teoxzXjLSHjxacrV+c8919oKyopiDDM9mPArpV8VkvetnMt4RB+x6NlHyU9VhlPqi4Rkd6Y5h+MMH6F6MWrYiR31K8eJY7K34cSon50c6zhmRDolsyJF5LNTWt+MyPYfBzV9vs9aiM/K2/7pWysqlrfMjWeJeu8WegoXcDS4kBrdyTyAXkT6nt0bT3Rkld0AGAseud9q7gO6/wUP+bbzPqVr38QqrXpe2bOPw26x+paRL7cRcqzLM90wcLB9ZXlqcuR+CLzdk5WSc5dz01darioR7FURVYmZEREBURRAFVEQFRREBVERAVFEQBFVEBVFUQBFFUBFVFUBEREBVEVQEVURAFVEQBERAVREQFUVUQBFVFACfBEU/QdfJRnxPxUPPmURRobGT4n4oCRyJRFJIVUVQgiIqgIiqIAoiIAqoqgIiqICIqiAiKogIqiICIhIABPXx/QqRjx96gERFVLBEQ4wd1cZAxjz3BUAiKqKQEVUPIdPVAVEPu+KKARFVFICqIoAUVRSCImMg9MJ+3MKAEVGDyBTGOaAiKqHAxkjdAEVwikERXbZTbBP1FQAqm3iEUgiqioA8fcgIibevvCoGSP1hRsBRVFIIiKoCIqiAKIiAIiqAiyPROkqrWeo4bXTksi9ueYfzbVji3L2B8LhqiKNw+Vvpo+5bncj5/F/wAKrsbUegO9UTdjul6l1knoX18zDwVFSA9+HYwd84HuWI9pOibJYIaK8adr2TWyuOG0/eBz4/DG+T71gFQJmVM7KgubM17hKHfhZ3yPMpJT1EUUUksEjGSN+1OkYQ0jf2T6hVwj52D4nmiHA5HKLY8gzrsksNt1JrllBdqVtTSmmlf3ZcW/OGMHYg9Ss3rajsYprvUWmosdRFNHM6B72OkABBI58SxrsJ/lKj8qOZ31LN6jsZst+1VcKt+pe9L6l881LE1pc3LskHqFqWy9QNadpuh4dE6hihpJXy2+ri76nL+bQCMtJWDrZHbLquj1FqaCjt2XUtrjdB3hGA9xxxbc9sY9y1ueavq3y9QFs3sX0xZtTXq6RXqibVwwUzZGNLnDhPFgn5pHTK1mtyf9njA1HeieQo2E+nH/ANVFu+XoDl9nOxIHDrBWOcDg8IkI/PWrtSTWio1FWS2GB8FrLm9xG8EFo4RnmSeeVtE1/Ydx/Pt9Y5wO5zLjPrxrUNeac3KqdRtLaUzPMIIxhmTw/RhYVaB1kRFsA+1LTS1tXBS07S+ad4jY3HNzjgfDms/7T+zwaLFpmpgTT1FO2KZ+SQKgD5x38fgu72K2GGW71eqLg3Fvs8bnhx5d5jP1LLqC+jte0jqazVIb8tp5n1FFgYJZn5n6vetec/V07A/P+f8Ap4Iucsb4pnxSt4ZGOLHjwcDg/SuK2H4B6Wn5bZBqCgmvUBmtjZR8oYAcub1Gxyt62C0dlOpLLdbpbrBJ8ntjC+fvDIDjhLtvneS/PC3L2S/yZ6+dzIpXnf8AIvVFy11QMO13XaGr4qEaPt81I9rnmpMnF84YHCNyeuVhW6NJ4QM9EVsVpALZ3YzpezaouV4hvNC2qZBTNfGC9zeFxJGfmkLWK3N/2ef44v8A/qbPzisbXqINOSgNmka0YAcQB71wX0m/xiX+mfrXBZrsCbnAwDv16ra970nYrt2QW/VWn7e2CspSBXsa97i7BLXZBJA3GdvFapW2uxG8wy11z0lcHcVHdYXFjSduLGD8RhV2bS2gdHsn0dar626XrUFOJrVb4iC0uc0OfgOO4I6LXdxmp57lUy0cDYKV0rjFE0khrM7DJ35Ldev2Q9nnZXRaPo5c1dxkc+aQbO4OLiOfob7loz3YHQeSirb3IEQ+yUQ+yf28Va+wN+SWTs303oXT131BY3ySV9LEXSRGQlzywOJxxY6rEtUXbssqtNVsWnrNUw3VzW9xI9r/AJpyM83EcsrYt0l0bD2YaQdrGCSWn+SQdxwcWQ7um59kjotW63qezaWyRN0jTVEdwE4LjJ3h+14OR85xHgtWHfrsGvnbOIxhRAMDBGEW0AiKqQRERAFURAEUVQBe3pTU1dpG/wAF2oAHSRjhfGfZkYeYK8NPcPeFDSa0wbxqdUdkWpapl4vNsnhuOOKWIMe3iPnwkNKxDtG7Q4tXMprbbKIUVnozxRNc0B7nYx7hudlr3JPM59d0VaqSYCqKKz/YGadl2pbdpLWTbpdHStpm08jMxt4iHHH6lxotbSWPtLq9R2x0jqaarkc+Nw3khc4nGFh2VMnby69Vg609sGbdpN10xqC/tu+ne+ZJUj+FwyRcIDwMZb7h8VhJVyTzJPqosorS0CrYnZHrK0aMu9yqLw6dsdRTtjYYouLcOytdKglJR5loG5jdOw8uybTWuBPhN/zrVupJLPLqGrfYInxWsub3DHgggcIzzJPPK8ryTrlYxrUWCpg9Bk9NtlEVgNmV2trLbuyWm0rp+Wd9bUO4q6V8fDnO7h9IHuWM6D1O/R+rqK6Ev+Sg93UNaN3Rnn8FjJ35+GE65VarSTXzBk+va6xXPVtXcbA+U0dViVzHx8PBIfaA9SM+9Yzy2TJ8VFmlpaAWxdA6ztOndGaptde+YVFyhLIRHHkHMbm7npzWukyfH9v2KiUVJaYINgAei5KE5OVVkCLYvZLrK06Nr7tNdXTNbU07Y4zGzi3BPP4ha6VHJYyjzLQOUjg+V72+y5xI9MrgnJFkCrt2y4T2i60dxp3FstLM2VpHPY/3LqJlRpPuDKe0LVrtZ6rmuTWubTBjYoI3c2NAGc/+7J96xY8ypkqpFaWgRU8gB18lE+CA3nHrzs1u2jrHZ9Rw1tS+3U8beFsbwGvDA0kFrhnksb1RX9lU+na2PTtuq47oWD5O+Qy4BzvzcRy8VrDKZPiq1Uk97BTzPqiKK0FRREBUREAUVUQBVRVARFVEARVRAERVARVREAREQFURVARFUQERFUBEREBVFVEBVERAEREARFUBEVUQFUVRARVRVARVREAREQFUREAVREBEREBVEVQEVUVQERVRAEVUQBEVQEVURAVRVRAVFEQBERAVRFUBEREARVRAEVUQBEVQEVURAEVUQFREQEVUVQEREQBFVEBVFUQERFUAREQEREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAf//Z")
        _logo_img  = Image.open(io.BytesIO(_logo_data)).resize((110, 110), Image.LANCZOS)
        _logo_ph   = ImageTk.PhotoImage(_logo_img)
        tk.Label(frame, image=_logo_ph, bg="#0a0a0a").pack(pady=(10, 0))
        frame._logo_ph = _logo_ph
    except Exception:
        tk.Label(frame, text="🍽️", font=("Segoe UI Emoji", 80), bg="#0a0a0a", fg="#FF6B35").pack(pady=10)
    tk.Label(frame, text="RESTO BILL", font=("Georgia", 72, "bold"),
             fg="#FF6B35", bg="#0a0a0a").pack(pady=5)
    tk.Label(frame, text="Restaurant Billing System", font=("Georgia", 20),
             fg="#888888", bg="#0a0a0a").pack(pady=5)
 
    if 0 < days <= 14:
        tk.Label(frame,
                 text=f"⚠  License expires in {days} day(s)  —  {_lm.get_expiry_date()}",
                 font=("Arial", 12, "bold"),
                 fg="#fbbf24", bg="#0a0a0a").pack(pady=(8, 0))
 
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("splash.Horizontal.TProgressbar",
                    troughcolor="#1a1a1a", background="#FF6B35",
                    bordercolor="#1a1a1a", thickness=8)
    progress = ttk.Progressbar(frame, style="splash.Horizontal.TProgressbar",
                                orient='horizontal', length=400, mode='determinate')
    progress.pack(pady=20)
    pct_lbl = tk.Label(frame, text="0%", font=("Georgia", 16), fg="#666666", bg="#0a0a0a")
    pct_lbl.pack()
 
    for i in range(101):
        progress['value'] = i
        pct_lbl.config(text=f"{i}%")
        splash.update()
        time.sleep(0.015)
 
    splash.destroy()
    LoginWindow()
 
 
# ═══════════════════════════════════════════════════════════════
#  RECEIPT / KOT FORMATTER
# ═══════════════════════════════════════════════════════════════
class ReceiptFormatter:
    def __init__(self, db: DatabaseManager, width=42):
        self.db    = db
        self.width = width
 
    def _center(self, text):
        return text.center(self.width)
 
    def _line(self, char="-"):
        return char * self.width
 
    def format_kot(self, order_id, items, order_info):
        W = self.width
        type_labels = {"dine_in": "Dine-In", "takeaway": "Takeaway", "delivery": "Delivery"}
        token = order_info.get("token_number", 0) or 0
        lines = []
        lines.append("=" * W)
        lines.append(self._center("KOT - KITCHEN ORDER TICKET"))
        lines.append("=" * W)
        # Token number — big and bold (KFC style)
        if token:
            lines.append(self._center(f"TOKEN  #{token:03d}"))
            lines.append(self._center("* * * * * * * * * *"))
        lines.append(f"Order  : {order_info.get('order_number', '')}")
        tbl_kot = order_info.get("table_number", "") or ""
        is_real_tbl = (
            tbl_kot
            and not tbl_kot.startswith("TKWY-")
            and not tbl_kot.startswith("DEL-")
            and not tbl_kot.startswith("BILL-")
        )
        if self.db.get_setting("enable_table", "1") == "1" and is_real_tbl:
            lines.append(f"Table  : {tbl_kot}")
        raw_type = order_info.get('order_type', 'dine_in')
        lines.append(f"Type   : {type_labels.get(raw_type, raw_type.replace('_', ' ').title())}")
        if order_info.get("waiter"):
            lines.append(f"Waiter : {order_info['waiter']}")
        lines.append(f"Time   : {datetime.now().strftime('%H:%M:%S  %d/%m/%Y')}")
        lines.append(self._line("-"))
        lines.append(f"{'ITEM':<28} {'QTY':>5}")
        lines.append(self._line("-"))
        for item in items:
            item = dict(item)
            name = item["item_name"][:28]
            lines.append(f"{name:<28} {item['quantity']:>5}")
            if item.get("notes"):
                lines.append(f"  ** {item['notes'][:W-5]}")
        lines.append("=" * W)
        lines.append("")
        return lines
 
    def format_bill(self, bill, items_data=None, extra_info=None):
        W      = self.width
        rname  = self.db.get_setting("restaurant_name", "RESTAURANT")
        addr   = self.db.get_setting("address", "")
        phone  = self.db.get_setting("phone", "")
        gst    = self.db.get_setting("gst_number", "")
        footer = self.db.get_setting("footer_message", "Thank you!")
        curr   = self.db.get_setting("currency", "Rs.")
        type_labels = {"dine_in": "Dine-In", "takeaway": "Takeaway", "delivery": "Delivery"}
 
        token = (extra_info or {}).get("token_number", 0) or 0
        lines = []
        lines.append("=" * W)
        lines.append(self._center(rname.upper()))
        if addr:
            lines.append(self._center(addr))
        if phone:
            lines.append(self._center(f"Ph: {phone}"))
        if gst:
            lines.append(self._center(f"GST: {gst}"))
        lines.append("=" * W)
        # Token number — KFC style, big and centred
        if token:
            lines.append(self._center(f"*** TOKEN  #{token:03d} ***"))
            lines.append("=" * W)
        lines.append(f"Bill No : {bill['bill_number']}")
        tbl_num = bill.get("table_number", "") or ""
        # Only show table if enabled AND it's a real dine-in table (not an auto-generated ID)
        is_real_table = (
            tbl_num
            and not tbl_num.startswith("TKWY-")
            and not tbl_num.startswith("DEL-")
            and not tbl_num.startswith("BILL-")
        )
        if self.db.get_setting("enable_table", "1") == "1" and is_real_table:
            lines.append(f"Table   : {tbl_num}")
        lines.append(f"Type    : {type_labels.get(bill['order_type'], bill['order_type'].replace('_',' ').title())}")
        lines.append(f"Date    : {datetime.now().strftime('%d/%m/%Y  %H:%M')}")
        lines.append(f"Payment : {bill['payment_method'].upper()}")
 
        # ── Customer details for Delivery / Takeaway ──────────
        ei         = extra_info or {}
        cust_name  = ei.get("customer_name", "") or ""
        cust_notes = ei.get("notes", "") or ""
        if cust_name and cust_name not in ("Walk-in", ""):
            lines.append(self._line("-"))
            lines.append(f"Customer: {cust_name}")
            if cust_notes:
                for part in cust_notes.split(" | "):
                    part = part.strip()
                    if part:
                        lines.append(f"  {part}")
        lines.append(self._line("-"))
        lines.append(f"{'ITEM':<22} {'QTY':>3} {'RATE':>6} {'AMT':>7}")
        lines.append(self._line("-"))
 
        if items_data:
            for item in items_data:
                item = dict(item) if not isinstance(item, dict) else item
                name = item["item_name"][:22]
                lines.append(f"{name:<22} {item['quantity']:>3} {item['price']:>6.0f} {item['total']:>7.0f}")
                if item.get("notes"):
                    lines.append(f"  ({item['notes'][:W-4]})")
 
        lines.append(self._line("-"))
 
        sgst_pct = float(bill.get('sgst_percent') or 0)
        cgst_pct = float(bill.get('cgst_percent') or 0)
        sgst_amt = float(bill.get('sgst_amount')  or 0)
        cgst_amt = float(bill.get('cgst_amount')  or 0)
        tax_pct  = float(bill.get('tax_percent')  or 0)
        tax_amt  = float(bill.get('tax_amount')   or 0)
        svc_amt  = float(bill.get('service_charge') or 0)
        combined_gst = sgst_pct + cgst_pct if (sgst_pct + cgst_pct) > 0 else tax_pct
 
        disc = float(bill.get('discount') or 0)
        if combined_gst > 0:
            pre_tax = round((bill['subtotal'] - disc) / (1 + combined_gst / 100), 2)
        else:
            pre_tax = bill['subtotal'] - disc
 
        
        if bill['discount'] > 0:
            lines.append(f"{'Discount':<30}-{curr}{bill['discount']:>6.2f}")
 
        lines.append(self._line("-"))
        lines.append(f"{'Taxable Amount':<30} {curr}{pre_tax:>7.2f}")
 
        if tax_amt > 0:
            lines.append(f"{'GST (' + str(tax_pct) + '%)':<30} {curr}{tax_amt:>7.2f}")
 
        if svc_amt > 0:
            lines.append(f"{'Service Charge':<30} {curr}{svc_amt:>7.2f}")
 
        lines.append("=" * W)
        lines.append(f"{'TOTAL':<30} {curr}{bill['total']:>7.2f}")
        lines.append("=" * W)
        lines.append("")
        lines.append(self._center(footer))
        lines.append("")
        lines.append(self._line("-"))
        lines.append("")
        return lines
 
 
# ═══════════════════════════════════════════════════════════════
#  PRINTING MANAGER
# ═══════════════════════════════════════════════════════════════
class PrintManager:
    def __init__(self, db: DatabaseManager, root=None):
        self.db   = db
        self._root_ref = root
 
    def get_config(self):
        if not os.path.exists(CONFIG_FILE):
            return {"method": "Browser", "printer_name": ""}
        try:
            with open(CONFIG_FILE) as f:
                return json.load(f)
        except Exception:
            return {"method": "Browser", "printer_name": ""}
 
    def save_config(self, cfg):
        with open(CONFIG_FILE, "w") as f:
            json.dump(cfg, f, indent=2)
 
    def print_lines(self, lines, title="Document"):
        cfg    = self.get_config()
        method = cfg.get("method", "Browser")
        if method == "Browser":
            self._print_browser(lines, title)
        elif method == "Thermal Printer" and WINDOWS_PRINT_AVAILABLE:
            self._print_thermal(cfg.get("printer_name", ""), lines)
        else:
            self._print_browser(lines, title)
 
    def _print_browser(self, lines, title="Receipt"):
        import tempfile, webbrowser
        html = (
            f"<html><head><title>{title}</title>"
            "<style>body{font-family:'Courier New',monospace;max-width:400px;"
            "margin:20px auto;padding:20px;}"
            "pre{font-size:12px;line-height:1.4;}"
            "@media print{body{margin:0;padding:5px;}}</style></head>"
            f"<body><pre>{chr(10).join(lines)}</pre>"
            "<script>window.onload=function(){window.print();}</script></body></html>"
        )
        tf = tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8")
        tf.write(html); tf.close()
 
        # Temporarily minimize the fullscreen app so browser print dialog is visible
        try:
            root = self._get_root_window()
            if root:
                root.overrideredirect(False)
                root.iconify()
                def _restore():
                    try:
                        root.deiconify()
                        root.overrideredirect(True)
                        root.state("zoomed")
                        root.lift()
                        root.focus_force()
                    except Exception:
                        pass
                # Restore after 10 seconds — enough time to finish print dialog
                root.after(10000, _restore)
        except Exception:
            pass
 
        webbrowser.open(f"file://{tf.name}")
 
    def _get_root_window(self):
        """Return the main RestaurantApp root window."""
        return self._root_ref
 
    def _print_thermal(self, printer_name, lines):
        if not printer_name:
            messagebox.showerror("No Printer", "No printer configured."); return
        try:
            h = win32print.OpenPrinter(printer_name)
            win32print.StartDocPrinter(h, 1, ("Document", None, "RAW"))
            win32print.StartPagePrinter(h)
            win32print.WritePrinter(h, ESC_INIT)
            for line in lines:
                win32print.WritePrinter(h, (line + "\n").encode("cp437", errors="replace"))
            win32print.WritePrinter(h, ESC_FEED_LINES(4) + ESC_CUT)
            win32print.EndPagePrinter(h)
            win32print.EndDocPrinter(h)
            win32print.ClosePrinter(h)
        except Exception as e:
            messagebox.showerror("Print Error", str(e))
 
 
# ═══════════════════════════════════════════════════════════════
#  MAIN RESTAURANT APP
# ═══════════════════════════════════════════════════════════════
class RestaurantApp:
    C = {
        "bg":      "#1e1e2e",
        "panel":   "#252535",
        "card":    "#2a2a3a",
        "border":  "#3a3a4a",
        "accent":  "#FFA500",
        "green":   "#4CAF50",
        "red":     "#E53935",
        "blue":    "#2196F3",
        "text":    "#F0F0F0",
        "muted":   "#888899",
        "header":  "#12121f",
        "listbg":  "#1a1a2a",
    }
 
    def __init__(self, root, username=""):
        self.root     = root
        self.username = username
        self.root.title("HOTEL BILL SOFT")
        self.root.state("zoomed")
        self.root.overrideredirect(True)
        self.root.configure(bg=self.C["bg"])
 
        self.db        = DatabaseManager()
        self.printer   = PrintManager(self.db, root=self.root)
        self.formatter = ReceiptFormatter(self.db)
        self.session   = SessionTracker()
        self.session.set_user(username)

        if MOBILE_SERVER_AVAILABLE:
            self._mobile_server = MobileOrderServer(self.db, self._on_mobile_order)
            self._mobile_server.start()
            self._mobile_server.poll_orders(self.root)
        else:
            self._mobile_server = None
        

        self.active_order_id     = None
        self.active_order_number = None
        self.active_table        = None
        self.active_order_type   = "Dine-In"
        self._order_item_ids     = []
        self._last_bill_lines    = []
        self._last_bill_obj      = None
        try:
            if os.path.exists(LAST_BILL_FILE):
                with open(LAST_BILL_FILE, "r", encoding="utf-8") as f:
                    _saved = json.load(f)
                self._last_bill_lines = _saved.get("lines", [])
                self._last_bill_obj   = _saved.get("bill", None)
        except Exception:
            pass
 
        self._menu_items_cache = []
        self._session_count    = self._load_session_count()
 
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self._set_window_icon(self.root)
        # Style the dropdown popup listbox (the floating list when combobox is clicked)
        self.root.option_add("*TCombobox*Listbox.background", "#000000")
        self.root.option_add("*TCombobox*Listbox.foreground", "#F0F0F0")
        self.root.option_add("*TCombobox*Listbox.selectBackground", "#FFA500")
        self.root.option_add("*TCombobox*Listbox.selectForeground", "black")
        self.root.option_add("*TCombobox*Listbox.font", "Arial 11")
        # Global combobox dark styling — applied once at startup
        _cb = ttk.Style()
        _cb.theme_use("clam")
        _cb.configure("TCombobox",
                      fieldbackground="#000000",
                      background="#000000",
                      foreground="#F0F0F0",
                      selectbackground="#FFA500",
                      selectforeground="black",
                      arrowcolor="#FFA500",
                      bordercolor="#3a3a4a",
                      darkcolor="#000000",
                      lightcolor="#000000")
        _cb.map("TCombobox",
                fieldbackground=[("readonly", "#000000"), ("active", "#000000")],
                background=[("active", "#FFA500"), ("!active", "#000000")],
                foreground=[("readonly", "#F0F0F0")])
        self._build_ui()
 
        self.root.bind("<F2>", lambda e: self.add_selected_item())
        self.root.bind("<F5>", lambda e: self.print_kot())
        self.root.bind("<F8>", lambda e: self.reprint_last_receipt())
        self.root.bind("<F9>", lambda e: self.generate_bill())

        self._renewal_prompted = False

        if _demo_mode_enabled():
            self.root.after(1500, self._run_demo_tour)
        else:
            self.root.after(1500, self._check_license_expiry)
            # Re-check license every hour so expiry is enforced mid-session
            self.root.after(3_600_000, self._periodic_license_check)
 
    def _check_license_expiry(self):
        _lm = LicenseManager()
        _status, days = _lm.check_license()
 
        if _status in ("expired", "blocked"):
            self.root.withdraw()
            _show_blocked_window(reason=_status)
            _status2, _ = _lm.check_license()
            if _status2 not in ("ok", "warning"):
                self.root.destroy()
                return
            self.root.deiconify()
            return
 
        if 0 < days <= 14 and not self._renewal_prompted:
            self._renewal_prompted = True
            if messagebox.askyesno(
                "License Expiring Soon",
                f"Your license expires in {days} day(s) "
                f"({_lm.get_expiry_date()}).\n\n"
                "Would you like to renew now?"
            ):
                LicenseRenewalWindow(parent=self.root)
 
    def _periodic_license_check(self):
        """Re-check license status every hour. Blocks the app if expired."""
        try:
            _lm = LicenseManager()
            _status, days = _lm.check_license()
            if _status in ("expired", "blocked"):
                self.root.withdraw()
                _show_blocked_window(reason=_status)
                _status2, _ = _lm.check_license()
                if _status2 not in ("ok", "warning"):
                    self.root.destroy()
                    return
                self.root.deiconify()
            elif 0 < days <= 3:
                # Urgent warning when only 3 or fewer days remain
                messagebox.showwarning(
                    "License Expiring",
                    f"Your license expires in {days} day(s).\n"
                    "Please renew to avoid interruption."
                )
        except Exception:
            pass
        # Schedule the next check in 1 hour
        self.root.after(3_600_000, self._periodic_license_check)

    def _load_session_count(self):
        try:
            if os.path.exists(SESSION_FILE):
                with open(SESSION_FILE) as f:
                    data = json.load(f)
                if isinstance(data, list):
                    return len([s for s in data if s.get("status") == "completed"])
        except Exception:
            pass
        return 0
 
    # ═══════════════════════════════════════════════════════════
    #  NAV BUTTON IMAGE LOADER
    # ═══════════════════════════════════════════════════════════
    def _load_nav_image(self, filename, size=(28, 28)):
        """Load an image for a nav button. Returns PhotoImage or None."""
        try:
            path = resource_path(filename)
            if not os.path.exists(path):
                return None
            from PIL import Image, ImageTk
            img = Image.open(path).resize(size, Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            try:
                # Fallback: try tk.PhotoImage for .png without Pillow
                path = resource_path(filename)
                if os.path.exists(path):
                    return tk.PhotoImage(file=path).subsample(1)
            except Exception:
                pass
            return None
 
    def _make_nav_btn(self, parent, image_file, fallback_text, command, **kw):
        """Create a nav button using image if available, else fallback text."""
        img = self._load_nav_image(image_file)
        if img:
            btn = tk.Button(parent, image=img, command=command, **kw)
            btn.image = img   # keep reference to prevent GC
        else:
            btn = tk.Button(parent, text=fallback_text, command=command, **kw)
        return btn
    
    def cancel_order_popup(self):
        """Show popup to cancel the current active order."""
        if not self.active_order_id:
            messagebox.showwarning("No Order", "No active order to cancel.")
            return
        
        # Check if there are KOT-printed items
        items = self.db.get_order_items(self.active_order_id)
        kot_printed = [i for i in items if i["kot_printed"] == 1]

        if not items and not kot_printed:
            messagebox.showwarning("No Items", "This order has no items to cancel.")
            return
                
        dlg = self._make_dialog(420, 460, "Cancel Order")
        win = dlg["win"]
        f = dlg["frame"]
        
        order = self.db.get_order(self.active_order_id)
        order_type = order["order_type"] if order else "unknown"
        token = order["token_number"] if order else 0
        
        # Header
        hdr = tk.Frame(f, bg=self.C["red"], pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚠️  CANCEL ORDER", 
                 font=("Arial", 14, "bold"),
                 bg=self.C["red"], fg="white").pack()
        
        if token:
            tk.Label(hdr, text=f"Token #{token:03d}",
                     font=("Arial", 11),
                     bg=self.C["red"], fg="#ffcccc").pack()
        
        body = tk.Frame(f, bg=self.C["card"], padx=24, pady=16)
        body.pack(fill="both", expand=True)
        
        # Warning if KOT printed
        if kot_printed:
            warn = tk.Frame(body, bg="#4a2020", padx=10, pady=8)
            warn.pack(fill="x", pady=(0, 12))
            tk.Label(warn, 
                     text=f"⚠️ {len(kot_printed)} item(s) already sent to kitchen!",
                     font=("Arial", 10, "bold"),
                     bg="#4a2020", fg="#ff9999").pack()
            tk.Label(warn,
                     text="These items may already be prepared.",
                     font=("Arial", 9),
                     bg="#4a2020", fg="#cc7777").pack()
        
        # Order info
        info_text = f"Order: {self.active_order_number}\n"
        info_text += f"Type: {order_type.replace('_', ' ').title()}\n"
        info_text += f"Items: {len(items)}  |  Total: Rs.{sum(i['total'] for i in items):.2f}"
        
        tk.Label(body, text=info_text,
                 font=("Courier New", 10),
                 bg=self.C["card"], fg=self.C["text"],
                 justify="left").pack(anchor="w", pady=(0, 12))
        
        # Reason selection
        tk.Label(body, text="Cancellation Reason *",
                 font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(0, 4))
        
        reason_var = tk.StringVar()
        reasons = [
            "Customer changed mind",
            "Customer left without ordering",
            "Wrong order placed",
            "Kitchen cannot prepare",
            "Out of stock",
            "Duplicate order",
            "Other"
        ]
        
        reason_combo = ttk.Combobox(body, textvariable=reason_var,
                                    values=reasons,
                                    font=("Arial", 11), state="readonly")
        reason_combo.pack(fill="x", pady=(0, 10))
        reason_combo.set(reasons[0])
        
        # Custom reason entry (shown when "Other" selected)
        custom_frame = tk.Frame(body, bg=self.C["card"])
        custom_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(custom_frame, text="Specify reason:",
                 font=("Arial", 9),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
        custom_reason = tk.Entry(custom_frame, font=("Arial", 11),
                                 bg=self.C["bg"], fg=self.C["text"],
                                 insertbackground=self.C["accent"], bd=0,
                                 highlightbackground=self.C["accent"],
                                 highlightthickness=1)
        custom_reason.pack(fill="x", ipady=6)
        custom_frame.pack_forget()  # Hidden by default
        
        def on_reason_change(*_):
            if reason_var.get() == "Other":
                custom_frame.pack(fill="x", pady=(0, 10))
                custom_reason.focus_set()
            else:
                custom_frame.pack_forget()
        
        reason_var.trace("w", on_reason_change)
        
        status_lbl = tk.Label(body, text="",
                              font=("Arial", 9, "bold"),
                              bg=self.C["card"], fg=self.C["red"])
        status_lbl.pack(anchor="w", pady=(4, 0))
        
        # Buttons
        btn_row = tk.Frame(f, bg=self.C["card"], padx=24, pady=12)
        btn_row.pack(fill="x")
        
        def do_cancel():
            reason = reason_var.get()
            if reason == "Other":
                reason = custom_reason.get().strip()
                if not reason:
                    status_lbl.config(text="Please specify a reason.")
                    return
            elif not reason:
                status_lbl.config(text="Please select a reason.")
                return
            
            # Confirm cancellation
            if not messagebox.askyesno("Confirm Cancellation",
                                       f"Are you sure you want to cancel this order?\n\n"
                                       f"Order: {self.active_order_number}\n"
                                       f"Reason: {reason}",
                                       parent=win):
                return
            
            # Perform cancellation
            success, msg = self.db.cancel_order(
                self.active_order_id, 
                self.username,
                reason
            )
            
            if success:
                dlg["close"]()
                messagebox.showinfo("Cancelled", f"Order {self.active_order_number} has been cancelled.")
                
                # Reset the billing screen
                self.active_order_id = None
                self.active_order_number = None
                self.active_table = None
                self.table_no_var.set("")
                self._block_discount_trace = True
                self.discount_var.set("0")
                self._block_discount_trace = False
                self._order_item_ids = []
                self.cart_listbox.delete(0, "end")
                curr = self.db.get_setting("currency", "Rs.")
                self.total_lbl.config(text=f"Total: {curr} 0.00")
                self._update_status()
            else:
                status_lbl.config(text=f"Error: {msg}")
        
        tk.Button(btn_row, text="❌  Cancel Order",
                  command=do_cancel,
                  bg=self.C["red"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        
        tk.Button(btn_row, text="Keep Order",
                  command=dlg["close"],
                  bg=self.C["green"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
        
        win.bind("<Escape>", lambda e: dlg["close"]())

    # ═══════════════════════════════════════════════════════════
    #  WINDOW ICON
    # ═══════════════════════════════════════════════════════════
    def _set_window_icon(self, window):
        """Set window icon from icon.ico / icon.png if available."""
        for name in ("icon.ico", "icon.png"):
            path = resource_path(name)
            if os.path.exists(path):
                try:
                    img = tk.PhotoImage(file=path)
                    window.iconphoto(True, img)
                    return
                except Exception:
                    pass
        try:
            # Fallback: try .ico directly via iconbitmap (Windows)
            ico = resource_path("icon.ico")
            if os.path.exists(ico):
                window.iconbitmap(ico)
        except Exception:
            pass
 
    # ═══════════════════════════════════════════════════════════
    #  TOP-LEVEL SHELL
    # ═══════════════════════════════════════════════════════════
    def _build_ui(self):
        nav = tk.Frame(self.root, bg=self.C["header"], height=62)
        nav.pack(fill="x")
        nav.pack_propagate(False)
 
        try:
            from PIL import Image, ImageTk
            import io, base64
            _nav_logo_data = base64.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/4gIoSUNDX1BST0ZJTEUAAQEAAAIYAAAAAAIQAABtbnRyUkdCIFhZWiAAAAAAAAAAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAAHRyWFlaAAABZAAAABRnWFlaAAABeAAAABRiWFlaAAABjAAAABRyVFJDAAABoAAAAChnVFJDAAABoAAAAChiVFJDAAABoAAAACh3dHB0AAAByAAAABRjcHJ0AAAB3AAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAFgAAAAcAHMAUgBHAEIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFhZWiAAAAAAAABvogAAOPUAAAOQWFlaIAAAAAAAAGKZAAC3hQAAGNpYWVogAAAAAAAAJKAAAA+EAAC2z3BhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABYWVogAAAAAAAA9tYAAQAAAADTLW1sdWMAAAAAAAAAAQAAAAxlblVTAAAAIAAAABwARwBvAG8AZwBsAGUAIABJAG4AYwAuACAAMgAwADEANv/bAEMACAYGBwYFCAcHBwkJCAoMFA0MCwsMGRITDxQdGh8eHRocHCAkLicgIiwjHBwoNyksMDE0NDQfJzk9ODI8LjM0Mv/bAEMBCQkJDAsMGA0NGDIhHCEyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMv/AABEIAWYBZgMBIgACEQEDEQH/xAAcAAEBAAIDAQEAAAAAAAAAAAAAAQYHAgQFAwj/xABWEAABAwMCAwUDBAsKCwgDAAABAAIDBAURBiESMUEHE1FhcTKBkRQiobEVIzZCUmJzssHR8BYXJDU3Q3KEk+EzNERTY3R1ksLS8QglJlZXgpSVRVWi/8QAGwEBAAIDAQEAAAAAAAAAAAAAAAEDAgQFBgf/xAA1EQACAgIABQIDBAkFAAAAAAAAAQIDBBEFEiExQSIyE1GBBhQzYRUjJDRCU5GxwVJxoeHw/9oADAMBAAIRAxEAPwDRqiIumAqoiAIqogKoqiAIoqgIqoiAIiICqIqgCiIgKoiIAiqiAqiIgKiiqAiIiAIqogKoiIAiKoCKqKoCKoogKoiICqKqIAiKoCIqiAKKogCIiAiKqIAqipa4M4+Elp2B6ZUEbIoqikkiIqgIqmwwSQB59fpULmg4BHxH61G0AicTfEfR+tMt/CH0frTaBUTIxzb8Rv8ASnx96bARRVSCIiICqKqIAiIgCqbbA9Rtn/qmw24h8R+tRsEVU2/CHxH61cj8JvxH61ICJkfhN+I/WmW+Iz6j9aAKK/H3jCiAIqiAiqiqAiKqIAiqIAiiqAiKogIiqiAIiqAiIigHIDbYZJ2Hr+2FkeqLc20U9poQPnfJzLIfF7ic/AYHuXmWGnbW32hpnjiY+VoLfHdZX2nxFt3oZSNnwuGfR2/1rJLps0Lb9ZddPh7ZgZ2OPBRXmMnmuTGOeHODSWs3cR0GQMn4rE3kcFVOm4weqKfJJszSVXS2Tspu18dZbVcaqK5shb8vpRIA0sHI810P3zZOmitGEf7Ld/zLlbv5B78cDP2Yh6fiNWAuA4nbDmVTGCbewZ3++ZJ/5J0Z/wDVO/5k/fMk/wDJOjf/AKp3/MsDwPAfBTA8B8Fl8OINyaD1PTau1GbTXaS0vBC6lleX0tuLXgtaSMZJWnN+ox5Yx9C2D2M/d9/UKn8xa+xjIHLKiC1JgIiqtBEREAREQFX1pqaWsqIqemjdLPK4MjjaMlzydh6L5bEc8EfD3rZWlKeDQumDrS5RNddKoOhs1K7mDyMpHkc4ysJS0gdi41Vi7NqOksbrFar5e+HvrjLVs4mwuI2jHoCF5n749tAwez7Sx8xAsFqaqoq6uapqJXSTzPL5Xn75xOT9K+OT4rFVrXUGwf3yLZ/6e6W/sFP3yLX/AOn2l/7Ba/RZfDh8gZ+e0i1Dc9n+mP7D6Vls14s9B2fOv110bYKOurCY7ZSxUwLnjcGQ55YKwbQGloLzVz3a7uMOn7WO/rJjtxEbiMeq8zWGqKjVV/krnt7qmYO6pIByiiAw0D1G/vVXKnLSB4LiS8khoJOSGjAB8guKqi2AFVFUBERVARERAEVUQBERAEREBUURAVFEQFURUISu57mjQDq+2Z/zqzftPohLaKWuxnuZSw+QcAsD0vMINUW2Q7ATtHx2W5NQW0XWwVtHjJfGSwfjN3CzivSzyvFbvgcRosfb/s0L09NlmHZtQRXnUNZZZGtzcrdUU8ZPR4bxtPxaPgsPPE1xDhhwOCPAr3dEXUWTW9luTvYiqmB/9F2Wn6CqZe16PUp7WzxJ4JKaolgmaWyxPLHg+IOCuC2Z226WFk1iblTMxR3Qd6COTZPvhnz5+9az6nH/AEUwlzJMGfW7+Qe/f7Zh/MasBd7TvUrPrd/IPfv9sw/mNWAu9p3qVjX5AURFYDYHY19339QqfzFgHU+q2B2Nfd9/UKn8xa+6n1Krj7n9AVFEVgKiiICqIu/ZbNWagvFNa7fF3lVUPDWjo0dXHyChvlWwZFoLS1PfK2e43VxhsNsZ39bNy4sbiMevx3Xn6x1RPqq/PrC3uqSJoio6ccoYgMADzPPx3Xv67vFHa6CDRFikD7fb3cVbOw/41Uffb9QCSMctlr8+0T5quK5nsBVFFaAvU0/YqzUl6pbVQRl1RUP4c42Y3q8+i8xoLnAAEknAaObj0AWz58dmejfkbMDVV6i/hDg7/E6c9PIkYPjusJS10B5mu75SUlJBouwPP2Ltrj8plb/lVR98SeZAcSMctlgZyXEnGc9FXEk5znrnx81xUxjyrQKiiLIFRREBVEVQBFEQFRREBUURAVFEQBFVEAVUVQERVRAfWnlfBNHNGcPjeHA+BGD+hfoqKVsscUzQMPaHD37r84bnPzsDG/0reei7h8v0pQyvPFI1ndP9WkgfRhZ199HlvtTTumFq8P8Auaz11Z/sTqSfgZwwVH26LPmNx8crGcHhIHPO3ly/uW6de2P7M2IyxtzUUf2xviW9QtLkA7+O6xktPR0+DZiycRSk/UujP0PN3faZ2T2yne9vy+SDggkP3tXCDkH+m0fSvzzMx0Uz43sLHscWvYRgtcOY9x2Wy+ya7SzG4aYbOIp6rFZbnnA4aqMZx/7mjC63aTZo61kGs7fD3dLXuMNfEBg01W3Z4I6AkH4rWg+SWmdY+Nu/kGv3+2YfzGrAXe071Kz63fyD37x+zMOf9xqwF3tO9SrK+7BERVWA2B2Nfd9/UKn8xa+6n1K2D2Nfd9/UKn8xa+6n1Vcfc/oAqiKwERFceRUADz+jmtnR47MtGceW/urvcR7sY3o6Y9fJxG/juvO0HZaKho6jW9+jP2JtrsU0J/yuo+9aPIHcrE77ea2/3uqulwkL6md+Tnk0dGgdABt7lW/XLp2B57nHiJ4s5++8fNcVVFaAqByz+wRZLonSj9VXoxSP7i3Ure/r6p3sxRDc+8rFvS2D3dDWmksVnl17fYw+kpTwW2neD/CajofQH3bLCbvdKy9XaruVwk7yqqZC+Qnp+KPADljyXv651WNR3eKGhjEFloG9xQU/DgBg24yPE88nfdYp7sLGC36pAKKqKwBFUQEVREBFVEQBERAERVARVRVARFUQBRFUBFURARERAXphbJ7K7l8+stj3bYE0YPwIWtl6+mbn9idQ0dUT9ra/hkH4p2KJ6Zo8SxvvGLOvzo34W5aW42IOR4rSGtLCbFf5Wxt/gkx7yE+XUe45W8A7iAcDkEZBHUcwvA1fYRf7JJCwD5TF9shPXPUe9WzW0eF4HnvDyeWftfRmk6GuqLbcKeupHmOop5BLE4c8tIK3dVXO21JiutQP/Cerou4rsf5HWgYEniNxnoFomRjo3ujeC1zThwPMEc1n3Z1c6SuhrdGXiRrbbeBiCRw2gqQPmkeuy07I76o+kJ77Hr3XT1Xpbsn1Laq0AmO9QujlHsyxljeFwPmMFapPM55rZ+oNU1jezu46Iv5c272urjbE5w3miDts+4jHlhawPM+qmnswERRWg2D2Nfd//UKn8xa+6n1K2D2Nfd9/UKn8xa/6n1Kwj7n9AFEVWYIsg0fpap1bfYrfETHTtBlq6g+zBCPacfPmvFpqaWrnjp6eN0k0rxHExo3c88gti6kmj0BpX9x1BIx94rQJbzURu3YObYh4bEZwsJy8IHja91RT3eup7RZ2d1YbUzuKOIcnkbOlPiTjqsO5806AcsITusox5VoBRVD/AHk+A/Wsgdm32+qulfT0FFC6WqqHiOKMDdzjt8BzWe6zr6TSNgboWzTB8nEJrxVRu3mlx/gx5DYYHh6r72mIdmukxfapmNTXZhZbYHc6aE85T5np7lrR73ySukke58jiXF7tySeqq98tg4nOd+n0eSKFVWgKKqICqKogIqoqgIqoiAIiICqIiAqKKoAiIgIiqiAIiqAiIiAKjx6DOVFQoBu/Ql4+y+m4eN4M9P8AaZMf/wAn4YWS7H0ytMdn98+xN/bBK77RV4jdnkHdCtzgFux3I2z4q6D2j5tx3D+7ZbaXpl1X+TVPaPpr5HVi80zf4PM7EzQPZd/esCBc0gtJa4EEO6tPPIX6MraSC4UUtHUs44pW8DgfA9QtD3+yT2C6y0UzSQ0kxu/DafBYTiel+z/E/vFPwbH6o/8AKM3vDB2i6JZqGBjXais0YiuUbQMzwAbSjxIGFrM4ztuOh8Qvb0nqWp0pf4LnTgPYz5k8B5TRH2mkHyXs6+0xS2+aDUFkcJbBdcyQPGSIH/fRk+RyPcqF6Xo9IYUipG+Pj6qKwGwOxr7vv6hU/mLX/U+q2D2Nfd9/UKn8xa+6n1Kwj7n9AFRjG5x1OeWFFmGgtKwX6tmr7o7ubDbW9/XTnbONxGPX4rKT5Qe1pOmg0Pph2trlGDc6hroLNSvG5cdjMQfDBAWuqqpnrauaqqZXS1Ezy+R7ubnZ3Xu6y1VNqu/vrC3uqOICKjp+kUQGAB5nn71jpWMI+WAqoisByDc7DPPwWb6B07RyMqdVX4cNitR4yD/lM43bG3xHivC0npqp1ZfYLXS4bxHinmI2hiHtPP0hbAv1LSapoobPYqw0ths8pgha2In5RIAC6V2++5KiMJ3S5Idyuy2NUeafY11qbUVZqm+1N0rDh0pxHGOUUY9lo8MDC8jKzz97fP8A+UP/AMc/rT97c/8A7M//ABj+tbUcG5L2mt+kMf8A1GBqrL6vs9r4ml1LUxT4HJzSw/BYtVUdRRTmCpidDIPvX9fRV2U2V+5GxVfXb7Hs+CKgZ9fRZvT9nff08Uv2TLe8aHY7gnmB5pVTO32Ii7Irp1zvRhCi7l0ofsbc6ij4+PuncPFw8OfcuoBkgAEknAAG5Vbi09MtUk1tdiKrKbZoS4VsbZqp7aWIjIBGXFe23s7ouHDq6cnxDQAtmGHdJbSNSedRB6bNdIs3ruzydrS6gq2ykDPBI3BPvWH1dHUUNQ6nqYnRSt5hw5+iqsosr9yLqsiu32M+CIvrT081XO2GCN0kruTGDcqpLfRFzaXc+SLNKDs9qZWNfXVLYSR7DBl3vXons6oizDa2oDvEtGFtRwrpLejSlxDHT1zGulVldz0HXUcbpaSVtUxo3bjDx6BYqWlri0jDhzadiFTZVOt6kjYquhatweziiqKotCKIpBURRAVREQFRFEBya5zHAtdg5BB8MFb10lfBfbBBMXD5RGBFMOocNgffjPvWifBZRoW//YS+MbI4fJqkiKTi5A9Cpi9PqcbjmB97xm4r1R6r/KN2A7rwNW6bj1HaixoDayIZgf1J/BXu5GARuDyPj5q5GVa47R88ousxrVbDo1/7R+cZ4Zaed8MrSyVhw5p6HwWZaF1LSQxVOlr+7i09czwOc7/JJTykHUdPJe92gaUNZE670DB3zB/CGAe038JaqJzkdCMbrXlHa0fTsDOhmUq2Pfyvkz29VaarNLXua3VYDx7cEzd2zxHcOafTHvXiLZOmrhS67sMWjr1MIrnBk2aud48+5cfA+awC42+qtdxqKCuhdBVQOLJI3bcJ/V4LGD/hZvGbdjX3ff1Cp/MWv+p9VsDsZ+78Zz/iFTz/AKC1+0ZGTn9aR9zB37NZ6u/XeltVDH3lVUvDGDo3xcfILMNeXeitNug0RYpQ6goXcVdOw/4zUdd+oBJGOWy78Ib2ZaN788I1Xe4ftQxvR055uPgTz8d1rJziXHJ4id+I9fP3qF62CH2iiiKwFXOKGSeVkULDJK9wYxgGS9x5ALgPPl4+AWydI0dNovTjtcXaE/LJSYbPSuG73n+dI8jke5YyloHO+Sx9nOlDpeikadQXFokutQznDERlsQ8NiDtvuuPZ7g6fmOAP4S7p+K1a7raupr66errJXS1Mzy+V7zuXft0WxOz37n5/9ad+a1b3DY6tOfxP8Bnc1NqJ1gNNimZOZ+P2nYxjH614De0V4522M+OJsfoX07RyQLZgn+d6/wBBYJkkcz8VflZNsLXGL6FOHiU2UxlJdTatj1bRXmX5N3ZgqDu1jiCD6FffU1liu9rlHCPlMTS6N2N8jmFq+0ue28URYTx98zGPVbq5nfHmtnGteRW1M1culYtsZVmisb8sHOfRbtoQPkFLsP8ABM6eQWlqkj5ZNw+z3hx8Vuqg/i+l/JM+oKjhy1KRfxV7jBmptTfdJXflFkGg7PHOZrnO0O7t3dxNI68yVj2p/ulr/wAp+hZxoKVj9OujB+cyZwcPUbKvHipZL3+ZdlTlDEXL8kepfr7TWOlE87S+R5wyMHcrDpO0O4cfzKOmDc+y4uJ+OV6Wv7bU1EVNVwsc+OEOD2t3Iz1WvR13yAs8zIthZyroivAxaZ1c7W2bOsGsYbtUCkqYBBUO9gg/Nd8V3NUWSO8WuTDQKmIF8bwN9ui1XR1Bpa2Co3+1SB+WjfC2Mdf2l38zV7/iDCzpyY21uNzK8jElVbGdCNZ8JJwAeInZvn4LbGlrBHZ7fG97QayUBz3EbtBHILA9O0sdx1ZAwNPdd6ZMOHIDdbRr6oUduqao/wA3GX/qWGBVFKVj8GfEbpPlqj5PEv8Aq6ms0hp4o+/qfvhnZvqsdb2hV/eAuo6QtzybnKxKWZ9RK+aVxc+QlzieuVwJ6dFRZm2yluL0jZq4fTGCUltm4LFf6e+07pIWmOaP2487hY1ryyRsYLtTtAOeCYAc/ArwdIXFltv8L5ZGxwSAskLjgYxss5vN1tFZZayn+X073vhPCA774DIW2rY5FD5+5oypli5Kda6M1T5Ih58x7kXIO718BRVEBFURAFFVEBUREBFc7eHn4eaIoBuTQmoReLOKWd4NXSgA/js6H3DCysnwWgrFd5rJdYq2Fx+Yfnt6FvULedJWwV9HDV0zuKKVvECD8R7itit76Hz7j3Dfu93xIL0y/ud0O+bghpHUHqFqXXmj/sXUOuVE1xo5TmQD+bctrF2dh6L5zRxVEL6eZjZIpBhzXdQkobNDhmfPBuUl2fdH54ZI+IhzHPYWuDmvbnIPQjC2fAYO1extpp+7h1lQQ/apD80XCIdD04hyzz2WJ6v0tJp+uD4gTRSuzE/8H8UrH6asqKCqiqqSV0NRC8PjkbsWkeHktWcD6VTkQvrVkHtMyPRd+Gi9YMrq2lke2IPp6qHk9gcMOwPJZpRaK0/pyol1s+4x3DTFMwTUMYPz5p/vYXDpg59cLq1VNTdqtqdcqFkdPrCkjzVUrfmiujH37B48sjxytZvlqYoTRSSSsjZIS6EkhrXjY5b4qtrmf5lx279e63UN7qrpXycU9Q/JHRjejB4ADA9y8w81UVyWloERF9IY2yzRxue1jXuDXSOPzWAnGT9anwDK9BaVh1BXy1t0f3Fhtre/rqg7bDcRjzK6ms9VTarvjqnh7migHc0VO3lHEBge88/esq7SHO0xZLXpC2ROjtJibVyVYORXyEb4PgCTty2WsySeePcqormewcTz55Wy+z37n5v9ad+a1a1wtldn33Pzf60781q6WB+Mc/iX4DPrq+wVl8FH8kMY7nvOLjdjnw4+pYx+4G8fhUoHnIs5vWoKWw9waqOd/fcXD3bQeWPH1Xm0+u7VPURw91Vxl7g3LwAAT4rduqx5WPnfU0Me7KhUlWuiPlp7RjLXVMrKuVs07PYaz2WnxXq6iu8dmtckhcO/kHBC3xPivWk4yxwicO8IIYTyzjbK01eaiuqLlMLhIXzRuLSOWB5BTfKONXqC7mOPCWZbzWPseeOe/PK3fQfxfS/kmfUFpELd1D/F9L+SZ9QVHDe8jZ4t0UTU2pvukrvyn6Fz03fpLFXF5bx08gxKzy8Vw1N90lf+UPL0Xk891pTnKFrlHwzoQhGymMZdmjdlDcKS6U4kpJWzNcN25GR6hedcNKWi4lz5KfupDzfEeE59OS1TT1M9LJ3lPK+J/iw4WSUOvbpTFrakR1TRtlw4XfQt6OdVYtWo5s+H21vmokdq49n1TEHSW+ds4G/A8Yd8QsPqKeWkmdDURvjkGxa8bra1m1VQXh/cjihqP83J19Fy1LYYbzbpMMAq4wXRvxucdPNLcOuyHPSyac62ufw70YZoAf8AiM5HKB2PiFmmqyRpevI2+YB8SsK0I7g1KGu2c6J7cHx5rONTxmXTNe0DJ7vPwKzxf3aX1K8z97j9DUCio3GUXGO7oiAAdAuTI3SvDI2Oc92zWtGSV93UFYAXGknAAyfmHZZKLfVGLlFd2dYor6oo2ZEREQBFUQERVRAEVRAEURAXJCzjs/1IKGp+xdW/FLMftTjyY89PQrBlyaXNGRkeY6eaKTi9mvl40MmqVU+zP0SSQSfqU4ljOir3PdrEDUgGWB3dl4+/GNlkQdsMrfj6ls+ZZWO8e6VUvBxraSnuNFJSVUQkikG4P1haZ1LpufT1cWOBdSuOYpfELdPHthdS40FNc6B9JVs44nDbxafEKudO10Ohwjic8Ken7X4NHUNwq7VXQ1tBPJBUwO445GnBBHT08uS2PVUNF2p22S52uGOm1bTszWUQIArWj7+MeOwzjG+Vg+otO1VhrTHKC+F5+1S/euC86hr6u2VsVZQzvp6qB3HHKw4II8PLy5LRnCW9+T6DXbC2Ksg9pnykjkjldHIxzHsdwva4YLXeBC+efULaj4bb2sUvyimZT0GsomjvIXHhjuAH3w8HemN1rKrpZ6KplpamF8NRC7hkjeMFp8MJGSfRlh8FeI4x08DyU92EWYNiaOvlHf7QNE6mmxSzP/7srnnJpJujc/gnwO26w6/2Ou03eai13GF0c8Jxy2e3o5p65GD715oJA26/t8Vs6zVcHaVYGabuk7Y9R0bSbXWSbd+wD/AvP0A89lTJOL2uwNYnn09y2X2ffc/N/rLvzWrXdbRT2+vmoqqF8FTC/u5InDBY4c/d+hbE7Pvufm2x/CnfmtXS4c92nP4l+Azzu0jB+xmw/nf+BYJ08P1rPO0jlbPWX/gWB9Fjmr9eyzh/7vE2to+8fZSztZI7NRTgMfg8x0PwwvF19ZtmXaJu2zJsfQVjWmru6z3mKYn7S/5krehaVtqaKKrpXxPAkikbyPJ3UfQt+prJo5H3Rzb08TJVkezNHA77HO63fQfxfS/kmfUFrDV1hFhvAbFvSTsE8BPUcnN9xz8Fs+hGKCmHhEwZ8dgquGpqU0/BbxSSlCEl5NTan+6Sv/KH6lsSxUVurLDQz/IaZ5dE0Oc6Jpy4DB6eIWvNTfdJX/lP0L1tH6mjtZNBWOIpXuy1/wDmyq6LIwyJKXkuyapzxouD6pHV1rbfkF8dJHE2OnmYHMDW4aMAA8vMLHfX6VuyopKO60nBNHHUQO5EH9PMLwH6Csz3lzX1LBn2WvGPp3Wd2DJycq30ZXj8ShGChZ3Rr21iU3akEGe971vDw9N91uvcc+eV5Vq05bbS8yUsDjKdi954nD9C6GqNSQWqjkp4ZBJWSDhDW78A8Stiir7rW3Nmtk2/fLUq12MFoK9tv1W2qbtE2ocD/RccLbM0MdVTSQuIMcrCzI6tPVaO9ea2DpDVEUtOy218gZK0YikdycPArWwr47cJdmbXEMaTSsh3RhFyoJrXXy0k7S1zHEDPUdCF1M9M781ue52ahu8YFZBxEey8HBHoRzXjM0DZmP4nPqnD8FzwB9AUT4fPm9HYmvilfL6+5jeg6B1Te/lhae6p2k8XTiPILN9S1goNPVkjjhxZwNHiTsu5T01FaqLgiZHTUzdySdveeq1zq7UQu9Synpj/AAWE5z+G7xV7UcWhxb6mtHmzMhSS9KMY5bZzhERcdHeRVERAVFFUBEREAREQBERAFRsoqofYGzezdxFmqvy4H0LM8rCuzk4s9T+X/QsxDjjmujStwR874yv22Z9MrkHHC+PF5q8ZxjKt5Tl6PlcKGmudG+lq4hJG4bE8x5grUeodN1NhqiH5fTH/AAcw5ehW4OI8sr4VdNT11K+nqoxJE4YLSOXmqrKOZb8nY4VxOeHPlfWL8Gj4p5qWdk9PJJFLG4Fr2EhzTvuCN1smG42ntQpYaC+Sw2/VUbQymuJbiOrwNmS9AeW/PksT1JpaeySulizLRu3bIB7PkVjuSBgZH7c/71zrK9Pr3PdU3wvjz1voz0L1Zbhp+6S2+6Ur6apiO7HDZ48WnqOq88gdDkLYVl1lbtQW2DTuuA6Sma3gpLqzeamPQOP3w9c7Lw9V6KuOlpY5pOCpts5zT18HzonjpkjkfHzWEZ9dSLjGVzillhlZJDI5kjHBzXDILT4jC4HYqdcrP8gbSqoou1XTzrhBHGzWNthzUwtaB8viG3G38Yft0WA0F+utnifS0dQYWcZLmGNpIdyPMeS+FqulbZrnT3G3zGKpp3B0bh5b4PkfDks81LaaLWthl1pp+ER1cX8b0DecZ/zrR4EYOBssITlU+jMZRjJaktmC3O9XC7918un73us8HzGtxnGeQHgF0FTjOxBHQhRWSk5PbexGKitJBe3TarvdLTsghrSGRjDAYmHHhuRleIuzQ0VRca2Gio4ny1NQ8Rxxt5ucTge5SrJQ6p6IlCM/ctm27jZX6m7BaC+yYkudFLNO5wGDJEJSx2AOgAb06LW7NYX2JjY2V3zGANb9pZyH/tW2YbvSWDWOlNEOma+lp6GS33EMI4TLOMkZ8nYWmL1bZLPfa+2S5L6SokgJPXhcRn34yq6rpqT6vqQ6oNaaXQ69TUzVlTJUVD+OWQ5c7AGT7l8s9OioaScD4LjzWb6vbM1pdEd+hvFxtpBpKuSMdG5y34HZew3Xd6aMF0Lz4mPf6FjCqsjdZHomVSoqn1lFHuVmr71WMLX1fdxnm2Nob/evELnPcXvcXOO5Ljkrj1RROyU/czOFcIe1aCZP05RFgZnsUOqLxb2COGrc6P8ABkHEPpXfdr29ObhroGnxbHv9KxhFar7EtKRQ8aqT24o79ferjcz/AAurkkH4Psj4BdBEVcpOT3J7LoxjFaitBERQSEVUQBVEQBREQFRREAVUVQEVUVUA2R2dnFoqfy/6FmAdsFhnZ8cWiq/Lj81Zdxcl1sdbrR8/4wv2yZ9cpxL5cScSu5Tmcp9OJOIj3r58RynEQmmSo6OUjGSxOikY17XD5zHciPRa71JoySjL6u3MdJT83Rcyz08Vn5dunGcnfmsLKIzR0cLNtxJbg+nlGj+WRnyKyvSuuq7T0TrfURNuVklz39unwWnPMsJ3afTG69rUGkYa7jqaENiqeZZya/8AUsAqaaWkmMNRG5j28wRuPRcm2hweme1xM6rKjzRfX5Gd3LQttvlvkvWhKk1lM0cU9rlP8Jp9+TR98P0LX8kbo3vY5rmvacFjhhzT4ELs2y519nro6221c1LURnLZYyR8cc/etgjUOmNetZBq2BlovWOGO80rR3cp/wBK3kPXC1+sfzNw1oefLHkva0tqet0rfIrhRFr/AL2eB3szRnm1wOxXe1JoK96ab8omhbWW5wzHX0ZL4XDxJ6LFs+YI+IVi1JA2PNQ9meo6iSoprxcNPVExLzBUwB8LHHfAI3x71139m1vl3o9f6Xc08u/qjCf90gkfFYCSTzJPqVMA8wD6rDka6Jg2Ezs5stMA+66+sEcY3cKKQzE+h/uX1Gq9NaMgkj0VTT1lyewsdeK1oHAP9Gw7Dwzha3wPwR8FyJLt3EuP426nkb7sHYjrp4riyvdK51Qyds5kPMuDs5WadsVMyLtGqqiIAMrKeGq/3mAfWFgjWOle1jRxOcQ0DqSSBhZ/2xuH7toKcn59NbaaJ/kQ3iI+lO01oGAxvdE9rwNwcj3fsV2rlTNpax3CPtMgEkZ/FIz9C6eCevVZFVQfLdK0lSBmWnBDvNoJC26q/iRkvK6lNtnw5Rfz6GOKqcseiKhdi4KqIgKoiICoiiAKqKoAiiICoiiAKoiAKIiAIqogKiiIAqoqnkGw9AHFpqfy/wDwrLOJYhoI4tVT+XH1LK+JdnFW6ong+LLeXM+nEhOy+eU4tithxOcl1PoXBjS5xw0DJPkvnDUx1MLZonh0bxs5q4yuJhd5ArWNj1JUWibhdmSlPtx9R5ha91yrklLszpYfD3lVTcfcjaRKnEutR11PcKZtRTSB7D58vVfUnzWxHTW0aMqpRbjJdUci8rz7paaW7xGOdmHj2ZRzC7nEuJcplCM1posqnKp80HpmtLxpystDy4tMtPzEzBsPVeNyJG2PDmCtwuIc0scA5p5tcMg+5Y1dNI01XxS0R7mbmWk/NK51+A11genw+MqSUb+/zPG05re+6ZPd0lV39G7aSiqftkLx4Frsge7CyPOgdaO4pHSaWuz/AAHFSPd788PuwsDrbdU2+XgqYi3zI+aV1snruPB24XLlU0/kzuRlGa5ovaMwvXZlqS1xOqoaRtyoQMtq7c/vWEeJHMLEJGd2/hd809Q5pafpXp2jUl70/KJbVdKukPgx5Lf907H3hZY3tVnrQGX/AE7ZLwTs6R8AjlPnxNAWPqRJr7Hl8DlDgdWt/pH9ithO1L2cT4dPoSqif1EFwc1vuGVRrfSFrAksehIBMOUtxqHTAeeM7qeeWuwPnoHSoEg1XfmGksFtd33FMC11TK3drGDqM9ViOobzNqHUNfd6gYkq5nScP4LTyHuGAu3qPWF41VMx1xqmmCPaKngbwQxjyaMfErwkhFrqyShZtp6IS6ejjeMtk42n3khYT4Y3OcLYlrpvklqp4Xc2s+vc/WuvwuHNZKX5HK4tPlrivOzAaundS1csDhuxxC+K97VVOIrgyYDHfM3PmNvqC8HOdxyWlkV/Dtcfkb+PZ8SqM/mRVRVUlwUREBURRAVFFUARREARVRAFVEQBERAEVUQBEVQEVUVQGe6Fdi11A/0w+pZTxrE9EHFsqPyw+pZOTuu7hx/UxPD8UW8uZ9eNXi2K+OVQ7otnlNDlOT3faX+i03nYLb8riIn4/BK1AuXxFe36npvs/wBIz+n+Tv2u61VqqRLTv2Ozo+jh4FbDtV8pbvADEeGYe3E4/OHp4rVi+sE0tNK2aF7mPbyc3mFrY+TKp/kdHN4dVkrfZm3ifBcS5YxZ9WR1HDT3AiOQ8pRjhd6+CyQuHC08QIO4x1Hku5VZG1bj3PJ5GJZRLlmikqZBGCuJK4FyvSKkhPGyZhjmY2Rvg4ZWOV+kqaYl9HJ3LzvwO9n4lZCXFcS4rGzGrtXqRt0ZNtL3Bmu6yy19CT3tO4sH37N10DnkSSByH9y2gXHl0XRqrVQVme9pYy/8IHhP0LnWcJ/lyOxVxj+Yv6GuyP2xhUbbjmswm0lSyZMUssXhkcQXTdo+YHDauP3sK05cPyY9FHZvx4ljS/i0Y2SSckknxJTGxJWRjSE2fn1bAPJhXo0mmKKnDTMXVDwc9QPgphw+9vqtCfEcdLaezyNP2d9VUMqZmkU7OWR7ZWZE596ABjQ1o4WjkB0TOy7uNjRohpd/JwMvLlk2c0u3g8LVcPHbGy43ifz9Vhx5rPNQM7yx1Q8Gg/SsDPM48VxuKw1fteUdvhM3KjT8M4oiLmnTCKqIAiqiAKqKoCIiICqIiAKqKoCIiIAiqiAqiIgCIiLuDOdFH/uyo/LD6lkvEsY0X/F04/0v6FkfEvRYa/URPGcTW8qZ9MplfPiTPmtrlNDlOUp+1P8A6J+pajW2ZD9pfv8AelamXI4otOP1PR8DWlP6BXooquUd8H3L2LTqKrtmGF3fQZz3bjy9F4ydcrKE5VvcWV2VQsXLJbRsuhvFJcmcUEnzzu5jtnD0XaLlq2OR8Tw+N7muHIg7rIKDVc0WGVre8YP5wD530Ls4/EovUbTg5PCHHrT1MwJXAkrr0lfS17M08zSfAn53wXY2G3VdeEoyW0zkyrlB6ktMio5JhMLPx1MdfIYQ80RRpEkyrlTKKexBSU6KJuUB0b07hstYf9Hj4kLXw5LNtUVAhtLoifnSuA9w3WFZB5LzvFZp3JL5HpuEwapbfzIqoi5h1AiIgKoiIAiKoCIiICqKqIAqoqgCiKoAiiqAiqiIAiIgfYyTTt7pbbSyRTiXic/iHA3O2F641ZbcezUf2awTO2FdvALbrzra4qK8Ghdw6i2bnIzv91lt/BqP7NT91lszyqP7P+9YIis/SVxWuE435/1M5k1XbSx7WifJGBlmAsHx5Y8lFVr35M79Ofg2sfErx01DyFEVVBshREQBXJxzURAcmSPjeHscWuHVpwvao9T1cADZwJwNt9iF4aZVld1lXsZVbRXatTWzPKTUNvqgGmUxP/Bf+teoxzXjLSHjxacrV+c8919oKyopiDDM9mPArpV8VkvetnMt4RB+x6NlHyU9VhlPqi4Rkd6Y5h+MMH6F6MWrYiR31K8eJY7K34cSon50c6zhmRDolsyJF5LNTWt+MyPYfBzV9vs9aiM/K2/7pWysqlrfMjWeJeu8WegoXcDS4kBrdyTyAXkT6nt0bT3Rkld0AGAseud9q7gO6/wUP+bbzPqVr38QqrXpe2bOPw26x+paRL7cRcqzLM90wcLB9ZXlqcuR+CLzdk5WSc5dz01darioR7FURVYmZEREBURRAFVEQFRREBVERAVFEQBFVEBVFUQBFFUBFVFUBEREBVEVQEVURAFVEQBERAVREQFUVUQBFVFACfBEU/QdfJRnxPxUPPmURRobGT4n4oCRyJRFJIVUVQgiIqgIiqIAoiIAqoqgIiqICIqiAiKogIqiICIhIABPXx/QqRjx96gERFVLBEQ4wd1cZAxjz3BUAiKqKQEVUPIdPVAVEPu+KKARFVFICqIoAUVRSCImMg9MJ+3MKAEVGDyBTGOaAiKqHAxkjdAEVwikERXbZTbBP1FQAqm3iEUgiqioA8fcgIibevvCoGSP1hRsBRVFIIiKoCIqiAKIiAIiqAiyPROkqrWeo4bXTksi9ueYfzbVji3L2B8LhqiKNw+Vvpo+5bncj5/F/wAKrsbUegO9UTdjul6l1knoX18zDwVFSA9+HYwd84HuWI9pOibJYIaK8adr2TWyuOG0/eBz4/DG+T71gFQJmVM7KgubM17hKHfhZ3yPMpJT1EUUUksEjGSN+1OkYQ0jf2T6hVwj52D4nmiHA5HKLY8gzrsksNt1JrllBdqVtTSmmlf3ZcW/OGMHYg9Ss3rajsYprvUWmosdRFNHM6B72OkABBI58SxrsJ/lKj8qOZ31LN6jsZst+1VcKt+pe9L6l881LE1pc3LskHqFqWy9QNadpuh4dE6hihpJXy2+ri76nL+bQCMtJWDrZHbLquj1FqaCjt2XUtrjdB3hGA9xxxbc9sY9y1ueavq3y9QFs3sX0xZtTXq6RXqibVwwUzZGNLnDhPFgn5pHTK1mtyf9njA1HeieQo2E+nH/ANVFu+XoDl9nOxIHDrBWOcDg8IkI/PWrtSTWio1FWS2GB8FrLm9xG8EFo4RnmSeeVtE1/Ydx/Pt9Y5wO5zLjPrxrUNeac3KqdRtLaUzPMIIxhmTw/RhYVaB1kRFsA+1LTS1tXBS07S+ad4jY3HNzjgfDms/7T+zwaLFpmpgTT1FO2KZ+SQKgD5x38fgu72K2GGW71eqLg3Fvs8bnhx5d5jP1LLqC+jte0jqazVIb8tp5n1FFgYJZn5n6vetec/V07A/P+f8Ap4Iucsb4pnxSt4ZGOLHjwcDg/SuK2H4B6Wn5bZBqCgmvUBmtjZR8oYAcub1Gxyt62C0dlOpLLdbpbrBJ8ntjC+fvDIDjhLtvneS/PC3L2S/yZ6+dzIpXnf8AIvVFy11QMO13XaGr4qEaPt81I9rnmpMnF84YHCNyeuVhW6NJ4QM9EVsVpALZ3YzpezaouV4hvNC2qZBTNfGC9zeFxJGfmkLWK3N/2ef44v8A/qbPzisbXqINOSgNmka0YAcQB71wX0m/xiX+mfrXBZrsCbnAwDv16ra970nYrt2QW/VWn7e2CspSBXsa97i7BLXZBJA3GdvFapW2uxG8wy11z0lcHcVHdYXFjSduLGD8RhV2bS2gdHsn0dar626XrUFOJrVb4iC0uc0OfgOO4I6LXdxmp57lUy0cDYKV0rjFE0khrM7DJ35Ldev2Q9nnZXRaPo5c1dxkc+aQbO4OLiOfob7loz3YHQeSirb3IEQ+yUQ+yf28Va+wN+SWTs303oXT131BY3ySV9LEXSRGQlzywOJxxY6rEtUXbssqtNVsWnrNUw3VzW9xI9r/AJpyM83EcsrYt0l0bD2YaQdrGCSWn+SQdxwcWQ7um59kjotW63qezaWyRN0jTVEdwE4LjJ3h+14OR85xHgtWHfrsGvnbOIxhRAMDBGEW0AiKqQRERAFURAEUVQBe3pTU1dpG/wAF2oAHSRjhfGfZkYeYK8NPcPeFDSa0wbxqdUdkWpapl4vNsnhuOOKWIMe3iPnwkNKxDtG7Q4tXMprbbKIUVnozxRNc0B7nYx7hudlr3JPM59d0VaqSYCqKKz/YGadl2pbdpLWTbpdHStpm08jMxt4iHHH6lxotbSWPtLq9R2x0jqaarkc+Nw3khc4nGFh2VMnby69Vg609sGbdpN10xqC/tu+ne+ZJUj+FwyRcIDwMZb7h8VhJVyTzJPqosorS0CrYnZHrK0aMu9yqLw6dsdRTtjYYouLcOytdKglJR5loG5jdOw8uybTWuBPhN/zrVupJLPLqGrfYInxWsub3DHgggcIzzJPPK8ryTrlYxrUWCpg9Bk9NtlEVgNmV2trLbuyWm0rp+Wd9bUO4q6V8fDnO7h9IHuWM6D1O/R+rqK6Ev+Sg93UNaN3Rnn8FjJ35+GE65VarSTXzBk+va6xXPVtXcbA+U0dViVzHx8PBIfaA9SM+9Yzy2TJ8VFmlpaAWxdA6ztOndGaptde+YVFyhLIRHHkHMbm7npzWukyfH9v2KiUVJaYINgAei5KE5OVVkCLYvZLrK06Nr7tNdXTNbU07Y4zGzi3BPP4ha6VHJYyjzLQOUjg+V72+y5xI9MrgnJFkCrt2y4T2i60dxp3FstLM2VpHPY/3LqJlRpPuDKe0LVrtZ6rmuTWubTBjYoI3c2NAGc/+7J96xY8ypkqpFaWgRU8gB18lE+CA3nHrzs1u2jrHZ9Rw1tS+3U8beFsbwGvDA0kFrhnksb1RX9lU+na2PTtuq47oWD5O+Qy4BzvzcRy8VrDKZPiq1Uk97BTzPqiKK0FRREBUREAUVUQBVRVARFVEARVRAERVARVREAREQFURVARFUQERFUBEREBVFVEBVERAEREARFUBEVUQFUVRARVRVARVREAREQFUREAVREBEREBVEVQEVUVQERVRAEVUQBEVQEVURAVRVRAVFEQBERAVRFUBEREARVRAEVUQBEVQEVURAEVUQFREQEVUVQEREQBFVEBVFUQERFUAREQEREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAf//Z")
            _nav_logo_img  = Image.open(io.BytesIO(_nav_logo_data)).resize((46, 46), Image.LANCZOS)
            _nav_logo_ph   = ImageTk.PhotoImage(_nav_logo_img)
            _nav_logo_lbl  = tk.Label(nav, image=_nav_logo_ph, bg="#0a0a0a")
            _nav_logo_lbl.image = _nav_logo_ph
            _nav_logo_lbl.pack(side="left", padx=(16, 4))
        except Exception:
            pass
        tk.Label(nav, text="RESTO BILL", font=("Georgia", 28, "bold"),
                 fg="#FFA500", bg="#0a0a0a").pack(side="left", padx=(4, 6))
 
        _lm = LicenseManager()
        _lic_status, _lic_days = _lm.check_license()
        if _lic_status in ("ok", "warning"):
            badge_text = f"🔑 {_lic_days}d left"
            badge_fg   = "#fbbf24" if _lic_days <= 14 else "#4ade80"
        else:
            badge_text = "⚠ No License"
            badge_fg   = "#f87171"
        tk.Label(nav, text=badge_text, font=("Courier", 9, "bold"),
                 bg=self.C["header"], fg=badge_fg).pack(side="left", padx=6)
 
        right_nav = tk.Frame(nav, bg=self.C["header"])
        right_nav.pack(side="right", padx=10)
 
        btn_kw = dict(font=("Arial", 12, "bold"), bd=0, padx=18, pady=8,
                      cursor="hand2", relief="flat")
 
        tk.Button(right_nav, text="Daily Sales Report",
                  command=self.open_reports,
                  bg=self.C["card"], fg=self.C["text"], **btn_kw).pack(side="left", padx=3)
 
        self._billing_btn = tk.Button(right_nav, text="Billing",
                  command=lambda: self._show_page("billing"),
                  bg=self.C["accent"], fg="white", **btn_kw)
        self._billing_btn.pack(side="left", padx=3)
 
        self._menu_mgmt_btn = tk.Button(right_nav, text="Menu Manager",
                  command=lambda: self._show_page("menu"),
                  bg=self.C["card"], fg=self.C["text"], **btn_kw)
        self._menu_mgmt_btn.pack(side="left", padx=3)

        tk.Button(right_nav, text="🔑 License",
                  command=lambda: LicenseRenewalWindow(parent=self.root),
                  bg=self.C["card"], fg="#fbbf24",
                  font=("Arial", 10, "bold"), bd=0, padx=10, pady=6,
                  cursor="hand2", relief="flat").pack(side="left", padx=2)
 
        self._make_nav_btn(
                  right_nav,
                  image_file="settings.webp",
                  fallback_text="⚙",
                  command=self.open_settings,
                  bg=self.C["card"], fg=self.C["text"],
                  font=("Arial", 13), bd=0, padx=10, pady=4,
                  cursor="hand2", relief="flat").pack(side="left", padx=2) 
        
        self._make_nav_btn(
                  right_nav,
                  image_file="printer2.png",
                  fallback_text="🖨",
                  command=self.open_printer_cfg,
                  bg=self.C["card"], fg=self.C["text"],
                  font=("Arial", 13), bd=0, padx=10, pady=4,
                  cursor="hand2", relief="flat").pack(side="left", padx=2)
        tk.Button(
            right_nav,
            text="📱 Mobile",
            command=lambda: (
                self._mobile_server.show_qr_window(self.root)
                if self._mobile_server
                else messagebox.showwarning(
                    "Unavailable",
                    "Mobile server is not available.\n"
                    "Make sure mobile_order_server.py is in the same folder."
                )
            ),
            bg=self.C["card"], fg="#FFA500",
            font=("Arial", 10, "bold"), bd=0, padx=12, pady=6,
            cursor="hand2", relief="flat",
        ).pack(side="left", padx=2)
        tk.Button(right_nav, text="EXIT",
                  command=self.on_closing,
                  bg=self.C["red"], fg="white",
                  font=("Arial", 10, "bold"), bd=0, padx=16, pady=6,
                  cursor="hand2", relief="flat").pack(side="left", padx=4)
 
        sub = tk.Frame(self.root, bg=self.C["panel"], height=64)
        sub.pack(fill="x")
        sub.pack_propagate(False)

        grp1 = tk.Frame(sub, bg=self.C["card"],
                         highlightbackground=self.C["border"], highlightthickness=1)
        grp1.pack(side="left", padx=(14, 0), pady=10, ipady=4)
        tk.Label(grp1, text="Bill Type", font=("Arial", 9),
                 bg=self.C["card"], fg=self.C["muted"]).pack(side="left", padx=(12, 6))
        self.bill_type_var = tk.StringVar(value="Dine-In")
        bill_type_cb = ttk.Combobox(grp1, textvariable=self.bill_type_var,
                                     values=["Dine-In", "Takeaway", "Delivery"],
                                     width=12, font=("Arial", 11), state="readonly")
        bill_type_cb.pack(side="left", padx=(0, 10))
        bill_type_cb.bind("<<ComboboxSelected>>", self._on_bill_type_change)

        # Open Orders button — plain, next to Bill Type
        grp_orders = tk.Frame(sub, bg=self.C["card"],
                              highlightbackground=self.C["border"], highlightthickness=1)
        grp_orders.pack(side="left", padx=(10, 0), pady=10, ipady=4)
        tk.Button(grp_orders, text="📋  Open Orders",
                  command=self.open_orders_popup,
                  bg=self.C["card"], fg=self.C["accent"],
                  font=("Arial", 11, "bold"), bd=0, padx=14, pady=4,
                  cursor="hand2").pack(side="left", padx=(4, 4))

        self.grp2 = tk.Frame(sub, bg=self.C["card"],
                         highlightbackground=self.C["border"], highlightthickness=1)
        grp2 = self.grp2
        tk.Label(grp2, text="Table No", font=("Arial", 9),
                 bg=self.C["card"], fg=self.C["muted"]).pack(side="left", padx=(12, 6))
        self.table_no_var = tk.StringVar()
        self.table_no_entry = tk.Entry(grp2, textvariable=self.table_no_var,
                                        width=8, font=("Arial", 12),
                                        bg=self.C["bg"], fg=self.C["text"],
                                        insertbackground=self.C["accent"],
                                        highlightbackground=self.C["accent"],
                                        highlightthickness=1, bd=0)
        self.table_no_entry.pack(side="left", ipady=5)
        self.table_no_entry.bind("<Return>", self._on_table_confirm)
        tk.Button(grp2, text="Open  ↵",
                  command=self._on_table_confirm,
                  bg=self.C["accent"], fg="black",
                  font=("Arial", 10, "bold"), bd=0, padx=12, pady=4,
                  cursor="hand2").pack(side="left", padx=(8, 8))

        # Show table group only for Dine-In AND if table feature is enabled
        if self.bill_type_var.get() == "Dine-In" and self.db.get_setting("enable_table", "1") == "1":
            self.grp2.pack(side="left", padx=(10, 0), pady=10, ipady=4)

        # Order Status and Cancel Button - ON THE RIGHT SIDE
        self.grp_order_status = tk.Frame(sub, bg=self.C["card"],
                                         highlightbackground=self.C["border"], highlightthickness=1)
        
        # Order ID label
        tk.Label(self.grp_order_status, text="Order ID", font=("Arial", 9),
                 bg=self.C["card"], fg=self.C["muted"]).pack(side="left", padx=(12, 6))
        
        self.order_status_lbl = tk.Label(self.grp_order_status, text="—",
                                          font=("Arial", 10, "bold"),
                                          bg=self.C["bg"], fg=self.C["muted"],
                                          width=30, anchor="w")
        self.order_status_lbl.pack(side="left", ipady=5, padx=(0, 8))

        # Cancel Order button - styled like "Open ↵" button
        self.cancel_order_btn = tk.Button(self.grp_order_status, text="Cancel",
                                          command=self.cancel_order_popup,
                                          bg=self.C["accent"], fg="black",
                                          font=("Arial", 10, "bold"), bd=0, padx=12, pady=4,
                                          cursor="hand2")
        self.cancel_order_btn.pack(side="left", padx=(0, 8))

        # Initially hidden, shown when order is active
 
        # Show table group only for Dine-In AND if table feature is enabled
        if self.bill_type_var.get() == "Dine-In" and self.db.get_setting("enable_table", "1") == "1":
            self.grp2.pack(side="left", padx=(10, 0), pady=10, ipady=4)
 
        self.order_status_lbl = tk.Label(sub, text="  No active order  ",
                                          font=("Arial", 10, "bold"),
                                          bg=self.C["border"], fg=self.C["muted"],
                                          padx=14, pady=6)
        self.order_status_lbl.pack(side="left", padx=(16, 0), pady=10)

        # Cancel Order button - next to order status, only shown when order is active
        self.cancel_order_btn = tk.Button(sub, text="❌ Cancel Order",
                                          command=self.cancel_order_popup,
                                          bg="#C62828", fg="white",
                                          font=("Arial", 9, "bold"), bd=0, padx=14, pady=6,
                                          cursor="hand2", relief="flat")
        # Don't pack yet - will be packed by _update_status when order is active
 
        # Move this ABOVE page_container.pack(...)
        footer = tk.Frame(self.root, bg=self.C["header"], height=52)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        _lm2 = LicenseManager()
        _s2, _d2 = _lm2.check_license()
        lic_txt = (f"Licensed  ·  Expires {_lm2.get_expiry_date()}"
                   if _s2 in ("ok", "warning") else "No valid license")
        tk.Label(footer,
                 text=f"Krisc_soft © 2026  RestoBill All Rights Reserved  |  {lic_txt}"
                      "  |  F2 Add Item   F5 KOT   F9 Bill   F8 Reprint",
                 font=("Arial", 9), bg=self.C["header"], fg=self.C["muted"]).pack(pady=(4, 0))
 
        _last_run = self.session.get_last_successful_run()
        _last_run_text = (f"Last successful run: {_last_run}" if _last_run
                          else "Last successful run: —")
        tk.Label(footer,
                 text=f"This application has been successfully running {self._session_count + 1} times   |   {_last_run_text}",
                 font=("Arial", 11), bg=self.C["header"], fg="#aaaacc").pack(pady=(0, 4))
 
        self.page_container = tk.Frame(self.root, bg=self.C["bg"])
        self.page_container.pack(fill="both", expand=True)
 
        self._build_billing_page()
        self._build_menu_page()
        self._show_page("billing")
    def _on_mobile_order(self, order: dict):
        def accept(o):
            table  = o.get("table", "Walk-in")
            waiter = o.get("waiter", "")
            note   = o.get("note", "")
 
            # Re-use an existing open order for this table, or create a new one
            existing = self.db.get_open_order_for_table(table)
            if existing:
                order_id = existing["id"]
                order_number = existing["order_number"]
            else:
                order_id, order_number = self.db.create_order(
                    table,
                    "dine_in",
                    waiter=waiter,
                    notes=note,
                )
 
            # Add each item to the order
            for item in o.get("items", []):
                self.db.add_order_item(
                    order_id,
                    item["name"],
                    item.get("category", ""),
                    item["qty"],
                    item["price"],
                )
 
            # Automatically load the incoming order into the UI cart immediately
            self.active_order_id     = order_id
            self.active_order_number = order_number
            self.active_table        = table
            self._suppress_bill_type_reset = True
            try:
                self.bill_type_var.set("Dine-In")
            finally:
                self._suppress_bill_type_reset = False
            self._update_status()
            self.refresh_order_display()
                
            # Always refresh the tables list on the left so the new order appears
            self.refresh_tables()
 
            item_count = sum(i["qty"] for i in o.get("items", []))
            curr       = self.db.get_setting("currency", "Rs.")
            messagebox.showinfo(
                "✔  Order Accepted",
                f"Mobile order added to Open Orders.\n\n"
                f"Table  : {table}\n"
                f"Waiter : {waiter or '—'}\n"
                f"Items  : {item_count}\n"
                f"Total  : {curr} {o.get('total', 0):.2f}\n\n"
                "Order is now in the cart. You can send to KOT or print bill.",
            )
 
        def reject():
            messagebox.showinfo(
                "Order Rejected",
                "The mobile order was rejected and will not be added.",
            )
 
        show_mobile_order_popup(
            self.root, order,
            on_accept=accept,
            on_reject=reject,
        )
    def _show_page(self, name):
        if name == "billing":
            self.menu_page.place_forget()
            self.billing_page.place(relx=0, rely=0, relwidth=1, relheight=1)
            self._billing_btn.config(bg=self.C["accent"], fg="white")
            self._menu_mgmt_btn.config(bg=self.C["card"], fg=self.C["text"])
        else:
            self.billing_page.place_forget()
            self.menu_page.place(relx=0, rely=0, relwidth=1, relheight=1)
            self._billing_btn.config(bg=self.C["card"], fg=self.C["text"])
            self._menu_mgmt_btn.config(bg=self.C["accent"], fg="white")
            self._refresh_menu_mgmt_list()

    def _run_demo_tour(self):
        if not _demo_mode_enabled():
            return

        self.root.after(0, lambda: self._show_page("menu"))
        self.root.after(2200, lambda: self._show_page("billing"))
        self.root.after(4200, self.open_orders_popup)
        self.root.after(7600, self.open_reports)
        self.root.after(12200, self.on_closing)
 
    # ═══════════════════════════════════════════════════════════
    #  BILLING PAGE
    # ═══════════════════════════════════════════════════════════
    def _build_billing_page(self):
        self.billing_page = tk.Frame(self.page_container, bg=self.C["bg"])
        self.billing_page.rowconfigure(0, weight=1)
        self.billing_page.columnconfigure(0, weight=1)
 
        panels = tk.Frame(self.billing_page, bg=self.C["bg"])
        panels.pack(fill="both", expand=True, padx=32, pady=(0, 0))
        
        panels.rowconfigure(0, weight=1)
 
        left = tk.Frame(panels, bg=self.C["bg"])
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 3))
        left.rowconfigure(2, weight=1)
        left.columnconfigure(0, weight=1)
        left.grid_propagate(False)   # ← prevents children from resizing this frame
 
        tk.Label(left, text="ADD ITEM",
                 font=("Arial", 11, "bold"),
                 bg=self.C["bg"], fg=self.C["accent"]).grid(
            row=0, column=0, sticky="w", padx=6, pady=(4, 2))
 
        self.filter_row = tk.Frame(left, bg=self.C["panel"])
        self.filter_row.grid(row=1, column=0, sticky="ew", pady=(0, 2))

        self.cat_var = tk.StringVar(value="All")
        self.cat_label = tk.Label(self.filter_row, text="Category:", font=("Arial", 10),
                 bg=self.C["panel"], fg=self.C["text"])
        self.cat_combo = ttk.Combobox(self.filter_row, textvariable=self.cat_var,
                                       values=self._get_categories(),
                                       width=14, font=("Arial", 10), state="readonly")
        self.cat_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_menu_display())

        self.search_label = tk.Label(self.filter_row, text="Search:", font=("Arial", 10),
                 bg=self.C["panel"], fg=self.C["text"])
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *_: self.refresh_menu_display())
        self.search_entry = tk.Entry(self.filter_row, textvariable=self.search_var,
                 font=("Arial", 13), bg=self.C["border"], fg="White",
                 insertbackground="black", bd=1, relief="solid",
                 highlightthickness=0)

        cat_enabled = self.db.get_setting("enable_category", "1") == "1"
        if cat_enabled:
            self.cat_label.pack(side="left", padx=(10, 4), pady=5)
            self.cat_combo.pack(side="left", padx=(0, 12))
        self.search_label.pack(side="left", padx=(4, 4), pady=5)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(0, 8), ipady=4)

        # Show/hide category sub-frame as a block — search frame never moves
        cat_enabled = self.db.get_setting("enable_category", "1") == "1"
        if not cat_enabled:
            self.cat_label.pack_forget()
            self.cat_combo.pack_forget()
 
        list_outer = tk.Frame(left, bg=self.C["listbg"],
                               highlightbackground=self.C["accent"], highlightthickness=1)
        list_outer.grid(row=2, column=0, sticky="nsew")
        list_outer.rowconfigure(0, weight=1)
        list_outer.columnconfigure(0, weight=1)
 
        self.menu_listbox = tk.Listbox(list_outer,
                                        font=("Courier New", 13),
                                        bg=self.C["listbg"], fg=self.C["text"],
                                        selectbackground=self.C["accent"],
                                        selectforeground="black",
                                        activestyle="none",
                                        highlightthickness=0, bd=0,
                                        exportselection=False)
        self.menu_listbox.grid(row=0, column=0, sticky="nsew")
        tk.Scrollbar(list_outer, orient="vertical",
                     command=self.menu_listbox.yview).grid(row=0, column=1, sticky="ns")
        self.menu_listbox.bind("<Double-1>", lambda e: self.add_selected_item())
 
        # "Add Extra Item" button — bottom of LEFT panel
        tk.Button(left, text="➕  Add Extra Item",
                  command=self._add_extra_item_popup,
                  bg="#B76219", fg="white",
                  font=("Arial", 13, "bold"), bd=0, pady=12,
                  cursor="hand2").grid(row=3, column=0, sticky="ew", pady=(3, 0))

        # ── Arrow button column between the two panels ────────
        panels.columnconfigure(0, weight=11)
        panels.columnconfigure(1, weight=9)

        right = tk.Frame(panels, bg=self.C["bg"])
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        right.grid_propagate(False)   # ← prevents children from resizing this frame
 
        tk.Label(right, text="ORDERS",
                 font=("Arial", 11, "bold"),
                 bg=self.C["bg"], fg=self.C["accent"]).grid(
            row=0, column=0, sticky="w", padx=6, pady=(4, 2))
 
        cart_outer = tk.Frame(right, bg=self.C["listbg"],
                               highlightbackground=self.C["accent"], highlightthickness=1)
        cart_outer.grid(row=1, column=0, sticky="nsew")
        cart_outer.rowconfigure(0, weight=1)
        cart_outer.columnconfigure(0, weight=1)
 
        self.cart_listbox = tk.Listbox(cart_outer,
                                        font=("Courier New", 13),
                                        bg=self.C["listbg"], fg=self.C["text"],
                                        selectbackground="#cc3333",
                                        selectforeground="white",
                                        activestyle="none",
                                        highlightthickness=0, bd=0,
                                        exportselection=False)
        self.cart_listbox.grid(row=0, column=0, sticky="nsew")
        tk.Scrollbar(cart_outer, orient="vertical",
                     command=self.cart_listbox.yview).grid(row=0, column=1, sticky="ns")
        self.cart_listbox.bind("<Button-1>", lambda e: self._edit_cart_item())
 
        tk.Button(right, text="Remove Selected Item",
                  command=self.remove_selected_cart_item,
                  bg="#F54607", fg="white",
                  font=("Arial", 13, "bold"), bd=0, pady=12,
                  cursor="hand2").grid(row=2, column=0, sticky="ew", pady=(3, 0))
 
        # Total label — bottom of RIGHT panel
        self.total_lbl = tk.Label(right, text="Total: Rs. 0.00",
                                   font=("Arial", 18, "bold"),
                                   bg="green", fg="black",
                                   padx=24, pady=12)
        self.total_lbl.grid(row=3, column=0, sticky="ew", pady=(3, 0))
 
        # ── Bottom action bar (same padx=32 as panels) ────────
                # ── Bottom action bar (same padx=32 as panels) ────────
        bar = tk.Frame(self.billing_page, bg=self.C["panel"])
        bar.pack(fill="x", padx=32, pady=(4, 4))

        r1 = tk.Frame(bar, bg=self.C["panel"])
        r1.pack(fill="x", padx=0, pady=(8, 4))

        zone_b = tk.Frame(r1, bg=self.C["panel"])
        zone_b.pack(side="left", padx=(0, 24))

        disc_row = tk.Frame(zone_b, bg=self.C["panel"])
        disc_row.pack(anchor="w", pady=(0, 7))
        tk.Label(disc_row, text="Discount (Rs.)", font=("Arial", 11),
                 bg=self.C["panel"], fg=self.C["muted"]).pack(side="left", padx=(0, 10))
        self.discount_var = tk.StringVar(value="0")
        self.discount_var.trace("w", lambda *_: self.refresh_order_display())
        tk.Entry(disc_row, textvariable=self.discount_var,
                 width=9, font=("Arial", 13),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(side="left", ipady=7)

        pay_row = tk.Frame(zone_b, bg=self.C["panel"])
        pay_row.pack(anchor="w")
        tk.Label(pay_row, text="Payment method", font=("Arial", 11),
                 bg=self.C["panel"], fg=self.C["muted"]).pack(side="left", padx=(0, 10))
        self.payment_var = tk.StringVar(value="cash")
        ttk.Combobox(pay_row, textvariable=self.payment_var,
                     values=["cash", "card", "upi", "wallet"],
                     width=11, font=("Arial", 15), state="readonly").pack(side="left")

        zone_c = tk.Frame(r1, bg=self.C["panel"])
        zone_c.pack(side="right", fill="x", expand=True)

        # 5 styled action buttons with spacing
        for txt, cmd, bg, fg in [
            ("🆕 New Order",     self.new_order,            "#2a2a3a", "#4CAF50"),
            ("🖨 KOT [F5]",      self.print_kot,            "#2a2a3a", "#FFA500"),
            ("🗑 Clear",          self.clear_order,          "#2a2a3a", "#E53935"),
            ("🖨 REPRINT [F8]",  self.reprint_last_receipt, "#2a2a3a", "#2196F3"),
            ("💰 BILL [F9]",     self.generate_bill,        "#FFA500", "#000000"),
        ]:
            tk.Button(zone_c, text=txt, command=cmd,
                      bg=bg, fg=fg,
                      font=("Arial", 12, "bold"), bd=0, pady=10,
                      cursor="hand2",
                      highlightbackground=self.C["border"],
                      highlightthickness=1).pack(side="left", expand=True, fill="x", padx=0)

        tk.Frame(bar, bg=self.C["border"], height=1).pack(fill="x", padx=10, pady=(4, 0))
 
        self.refresh_menu_display()
    def new_order(self):
        """Save the current order in open orders and reset screen for a new order."""
        if self.active_order_id:
            items = self.db.get_order_items(self.active_order_id)
            if items:
                # Order has items — just leave it in open orders and reset screen
                self.active_order_id     = None
                self.active_order_number = None
                self.active_table        = None
                self.table_no_var.set("")
                self._block_discount_trace = True
                self.discount_var.set("0")
                self._block_discount_trace = False
                self._order_item_ids = []
                self.cart_listbox.delete(0, "end")
                curr = self.db.get_setting("currency", "Rs.")
                self.total_lbl.config(text=f"Total: {curr} 0.00")
                self._update_status()
                messagebox.showinfo(
                    "Order Saved",
                    "Current order saved.\n"
                    "Use  📋 Open Orders  to retrieve and bill it.")
                return
            else:
                # Order is empty — just reset silently
                self.active_order_id     = None
                self.active_order_number = None
                self.active_table        = None
                self.table_no_var.set("")
                self._block_discount_trace = True
                self.discount_var.set("0")
                self._block_discount_trace = False
                self._order_item_ids = []
                self.cart_listbox.delete(0, "end")
                curr = self.db.get_setting("currency", "Rs.")
                self.total_lbl.config(text=f"Total: {curr} 0.00")
                self._update_status()
                return

        # No active order — screen is already fresh, nothing to do

    #  MENU MANAGEMENT PAGE
    # ═══════════════════════════════════════════════════════════
    def _build_menu_page(self):
        self.menu_page = tk.Frame(self.page_container, bg=self.C["bg"])

        hdr = tk.Frame(self.menu_page, bg=self.C["accent"], height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📋   MENU MANAGEMENT",
                 font=("Arial", 14, "bold"),
                 bg=self.C["accent"], fg="white").pack(side="left", padx=20)
        tk.Button(hdr, text="← Back to Billing",
                  command=lambda: self._show_page("billing"),
                  bg="#cc6600", fg="white",
                  font=("Arial", 10, "bold"), bd=0, padx=14, pady=6,
                  cursor="hand2").pack(side="right", padx=12)

        content = tk.Frame(self.menu_page, bg=self.C["bg"])
        content.pack(fill="both", expand=True, padx=8, pady=8)
        content.columnconfigure(0, weight=1)
        content.columnconfigure(1, weight=0)
        content.rowconfigure(0, weight=1)

        lf = tk.Frame(content, bg=self.C["panel"])
        lf.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        lf.rowconfigure(0, weight=1)
        lf.columnconfigure(0, weight=1)

        cols = ("Name", "Category", "Price", "Available")
        self._mm_tree = ttk.Treeview(lf, columns=cols, show="headings")
        style = ttk.Style()
        style.configure("mm.Treeview",
                        background=self.C["card"], fieldbackground=self.C["card"],
                        foreground=self.C["text"], rowheight=30,
                        font=("Arial", 10))
        style.configure("mm.Treeview.Heading",
                        background=self.C["border"], foreground=self.C["muted"],
                        font=("Arial", 10, "bold"))
        self._mm_tree.configure(style="mm.Treeview")
        for col, w in [("Name", 240), ("Category", 150), ("Price", 90), ("Available", 90)]:
            self._mm_tree.heading(col, text=col)
            self._mm_tree.column(col, width=w, anchor="center")
        mm_sb = tk.Scrollbar(lf, orient="vertical", command=self._mm_tree.yview)
        self._mm_tree.configure(yscrollcommand=mm_sb.set)
        self._mm_tree.grid(row=0, column=0, sticky="nsew")
        mm_sb.grid(row=0, column=1, sticky="ns")
        self._mm_tree.bind("<ButtonRelease-1>", self._mm_populate_form)

        form = tk.Frame(content, bg=self.C["panel"], padx=24, pady=16, width=360)
        form.grid(row=0, column=1, sticky="nsew")
        form.pack_propagate(False)

        tk.Label(form, text="ADD / EDIT ITEM",
                 font=("Arial", 13, "bold"),
                 bg=self.C["panel"], fg=self.C["accent"]).pack(pady=(0, 10))

        self._mm_name_var  = tk.StringVar()
        self._mm_price_var = tk.StringVar()
        self._mm_desc_var  = tk.StringVar()
        self._mm_cat_var   = tk.StringVar()

        # Create a container frame for all form fields to maintain order
        self._mm_fields_frame = tk.Frame(form, bg=self.C["panel"])
        self._mm_fields_frame.pack(fill="x")

        # Item Name
        tk.Label(self._mm_fields_frame, text="Item Name *", font=("Arial", 9),
                 bg=self.C["panel"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
        tk.Entry(self._mm_fields_frame, textvariable=self._mm_name_var, font=("Arial", 11),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=7)

        # Price
        tk.Label(self._mm_fields_frame, text="Price (Rs.) *", font=("Arial", 9),
                 bg=self.C["panel"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
        tk.Entry(self._mm_fields_frame, textvariable=self._mm_price_var, font=("Arial", 11),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=7)

        # Category frame with dropdown and add button
        self._mm_cat_container = tk.Frame(self._mm_fields_frame, bg=self.C["panel"])
        
        self._mm_cat_label = tk.Label(self._mm_cat_container, text="Category *", font=("Arial", 9),
                 bg=self.C["panel"], fg=self.C["muted"])
        self._mm_cat_label.pack(anchor="w", pady=(8, 2))
        
        # Category row with combobox and add button
        cat_row = tk.Frame(self._mm_cat_container, bg=self.C["panel"])
        cat_row.pack(fill="x")
        
        self._mm_cat_combo = ttk.Combobox(cat_row, textvariable=self._mm_cat_var,
                                           values=self._get_categories()[1:],  # Exclude "All"
                                           font=("Arial", 11), state="normal")
        self._mm_cat_combo.pack(side="left", fill="x", expand=True, ipady=3)
        
        # Add Category button
        tk.Button(cat_row, text="＋", command=self._mm_add_category_popup,
                  bg=self.C["accent"], fg="black",
                  font=("Arial", 12, "bold"), bd=0, padx=8, pady=2,
                  cursor="hand2").pack(side="left", padx=(6, 0))
        
        # Show/hide based on setting
        if self.db.get_setting("enable_category", "1") == "1":
            self._mm_cat_container.pack(fill="x")

        # Description - always last
        self._mm_desc_label = tk.Label(self._mm_fields_frame, text="Description", font=("Arial", 9),
                 bg=self.C["panel"], fg=self.C["muted"])
        self._mm_desc_label.pack(anchor="w", pady=(8, 2))
        tk.Entry(self._mm_fields_frame, textvariable=self._mm_desc_var, font=("Arial", 11),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=7)

        # Separator and buttons
        tk.Frame(form, bg=self.C["border"], height=1).pack(fill="x", pady=14)

        for txt, cmd, bg_col in [
            ("➕  Add New Item",     self._mm_add,    self.C["green"]),
            ("✏   Update Selected", self._mm_update, self.C["blue"]),
            ("🔄  Toggle",self._mm_toggle, "#795548"),
            ("🗑   Delete Selected", self._mm_delete, self.C["red"]),
        ]:
            tk.Button(form, text=txt, command=cmd,
                      bg=bg_col, fg="white",
                      font=("Arial", 11, "bold"),
                      bd=0, pady=10, cursor="hand2",
                      anchor="w", padx=14).pack(fill="x", pady=3)

        tk.Frame(form, bg=self.C["border"], height=1).pack(fill="x", pady=10)
        tk.Button(form, text="🗄  Backup Database",
                  command=self.db.backup_database,
                  bg=self.C["card"], fg=self.C["muted"],
                  font=("Arial", 10), bd=0, pady=7,
                  cursor="hand2").pack(fill="x")
        
    def _mm_add_category_popup(self):
        """Popup to add a new category."""
        dlg = self._make_dialog(360, 220, "Add New Category")
        win = dlg["win"]
        f = dlg["frame"]

        tk.Label(f, text="New Category Name", font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", padx=20, pady=(16, 4))
        
        cat_var = tk.StringVar()
        cat_entry = tk.Entry(f, textvariable=cat_var, font=("Arial", 12),
                             bg=self.C["bg"], fg=self.C["text"],
                             insertbackground=self.C["accent"], bd=0,
                             highlightbackground=self.C["accent"],
                             highlightthickness=1)
        cat_entry.pack(fill="x", padx=20, ipady=8)
        cat_entry.focus_set()

        status_lbl = tk.Label(f, text="", font=("Arial", 9),
                              bg=self.C["card"], fg=self.C["red"])
        status_lbl.pack(anchor="w", padx=20, pady=(8, 0))

        btn_row = tk.Frame(f, bg=self.C["card"], padx=20, pady=16)
        btn_row.pack(fill="x")

        def save_category(event=None):
            new_cat = cat_var.get().strip()
            if not new_cat:
                status_lbl.config(text="Please enter a category name.")
                return
            
            # Check if already exists
            current_cats = self._get_categories()
            if new_cat in current_cats:
                status_lbl.config(text="This category already exists.")
                return
            
            # Add to DEFAULT_CATEGORIES if not there
            if new_cat not in DEFAULT_CATEGORIES:
                DEFAULT_CATEGORIES.append(new_cat)
                DEFAULT_CATEGORIES.sort()
            
            # Update combobox values
            self._mm_cat_combo["values"] = self._get_categories()[1:]  # Exclude "All"
            self._mm_cat_var.set(new_cat)
            self.cat_combo["values"] = self._get_categories()
            
            dlg["close"]()
            messagebox.showinfo("Success", f"Category '{new_cat}' added!")

        cat_entry.bind("<Return>", save_category)

        tk.Button(btn_row, text="Add Category",
                  command=save_category,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 11, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Button(btn_row, text="Cancel",
                  command=dlg["close"],
                  bg=self.C["red"], fg="white",
                  font=("Arial", 11, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")

        win.bind("<Escape>", lambda e: dlg["close"]())
 
    def _refresh_menu_mgmt_list(self):
        self._mm_tree.delete(*self._mm_tree.get_children())
        for item in self.db.fetch_menu():
            self._mm_tree.insert("", "end", values=(
                item["name"], item["category"],
                f"Rs.{item['price']:.0f}",
                "✓ Yes" if item["available"] else "✗ No"))
        self.refresh_menu_display()
        self.cat_combo["values"] = self._get_categories()
 
    def _mm_populate_form(self, event=None):
        sel = self._mm_tree.selection()
        if not sel:
            return
        vals = self._mm_tree.item(sel[0])["values"]
        self._mm_name_var.set(vals[0])
        self._mm_cat_var.set(vals[1])
        self._mm_price_var.set(str(vals[2]).replace("Rs.", ""))
        self._mm_desc_var.set("")
 
    def _mm_add(self):
        n = self._mm_name_var.get().strip()
        c = self._mm_cat_var.get().strip()
        cat_enabled = self.db.get_setting("enable_category", "1") == "1"
        if not c and not cat_enabled:
            c = "General"
        try:
            p = float(self._mm_price_var.get())
        except ValueError:
            messagebox.showerror("Invalid", "Price must be a number."); return
        if not n or (cat_enabled and not c):
            messagebox.showwarning("Required", "Item name and category are required."); return
        try:
            self.db.add_menu_item(n, c, p, self._mm_desc_var.get().strip())
            self._refresh_menu_mgmt_list()
            self._mm_name_var.set(""); self._mm_price_var.set("")
            self._mm_cat_var.set(""); self._mm_desc_var.set("")
            messagebox.showinfo("Added", f"'{n}' added to menu.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not add: {e}")
 
    def _mm_update(self):
        sel = self._mm_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select an item to update."); return
        old_name = self._mm_tree.item(sel[0])["values"][0]
        n = self._mm_name_var.get().strip()
        c = self._mm_cat_var.get().strip()
        try:
            p = float(self._mm_price_var.get())
        except ValueError:
            messagebox.showerror("Invalid", "Price must be a number."); return
        self.db.update_menu_item(old_name, n, c, p, self._mm_desc_var.get().strip())
        self._refresh_menu_mgmt_list()
        messagebox.showinfo("Updated", f"'{n}' updated.")
 
    def _mm_toggle(self):
        sel = self._mm_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select an item to toggle."); return
        name = self._mm_tree.item(sel[0])["values"][0]
        self.db.toggle_item_availability(name)
        self._refresh_menu_mgmt_list()
        # Also refresh the billing page menu so availability reflects immediately
        self.refresh_menu_display()
 
    def _mm_delete(self):
        sel = self._mm_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select an item to delete."); return
        name = self._mm_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Delete", f"Delete '{name}' from menu?"):
            self.db.delete_menu_item(name)
            self._refresh_menu_mgmt_list()
            self._mm_name_var.set(""); self._mm_price_var.set("")
            self._mm_cat_var.set(""); self._mm_desc_var.set("")
 
    # ═══════════════════════════════════════════════════════════
    #  HELPERS
    # ═══════════════════════════════════════════════════════════
    def _get_categories(self):
        cats = ["All"] + DEFAULT_CATEGORIES
        try:
            c = self.db.conn.cursor()
            c.execute("SELECT DISTINCT category FROM menu_items ORDER BY category")
            for row in c.fetchall():
                if row[0] not in cats:
                    cats.append(row[0])
        except Exception:
            pass
        return cats
 
    def _on_bill_type_change(self, event=None):
        new_type = self.bill_type_var.get()
        self.active_order_type = new_type

        # Skip clearing the active order if called programmatically (not by user)
        if not getattr(self, "_suppress_bill_type_reset", False):
            self.active_order_id     = None
            self.active_order_number = None
            self.active_table        = None
            self.table_no_var.set("")
            # Temporarily block discount trace to avoid refresh_order_display
            # firing while active_order_id is None
            self._block_discount_trace = True
            self.discount_var.set("0")
            self._block_discount_trace = False
            self._order_item_ids = []
            self.cart_listbox.delete(0, "end")
            curr = self.db.get_setting("currency", "Rs.")
            self.total_lbl.config(text=f"Total: {curr} 0.00")

            # For Delivery: immediately collect customer details and create the order
            # so the order exists before the cashier starts adding items.
            if new_type == "Delivery":
                def _on_del_confirmed(name, phone, addr):
                    tbl   = f"DEL-{uuid.uuid4().hex[:4].upper()}"
                    notes = f"Phone: {phone} | Address: {addr}"
                    oid, onum = self.db.create_order(
                        tbl, "delivery", customer_name=name, notes=notes)
                    self.active_order_id     = oid
                    self.active_order_number = onum
                    self.active_table        = tbl
                    self._update_status()
                    self.refresh_order_display()
                self._delivery_customer_popup(_on_del_confirmed)

            # For Takeaway: create the order immediately (no extra details needed)
            elif new_type == "Takeaway":
                tbl = f"TKWY-{uuid.uuid4().hex[:4].upper()}"
                oid, onum = self.db.create_order(tbl, "takeaway", customer_name="Walk-in")
                self.active_order_id     = oid
                self.active_order_number = onum
                self.active_table        = tbl
                self._update_status()
                self.refresh_order_display()

        # Table No field is ONLY for Dine-In (and only if enable_table is on)
        table_enabled = self.db.get_setting("enable_table", "1") == "1"
        if new_type == "Dine-In" and table_enabled:
            self.grp2.pack(side="left", padx=(10, 0), pady=10, ipady=4)
        else:
            self.grp2.pack_forget()
            self.table_no_var.set("")   # clear any leftover table number

        self._update_status()
 
    def open_order_if_no_table(self):
        if self.active_order_id:
            return
        table_no = f"BILL-{uuid.uuid4().hex[:4].upper()}"
        oid, onum = self.db.create_order(table_no, "dine_in")
        self.active_order_id, self.active_order_number = oid, onum
        self.active_table = table_no
        self._update_status()
        self.refresh_order_display()
 
    # ═══════════════════════════════════════════════════════════
    #  DELIVERY CUSTOMER INFO POPUP  ← NEW FEATURE
    # ═══════════════════════════════════════════════════════════
    def _delivery_customer_popup(self, on_confirm):
        """
        Show a popup asking for customer Name, Phone, and Address
        for delivery orders. Calls on_confirm(name, phone, address)
        when the user clicks Confirm.
        """
        btype = self.bill_type_var.get()
        icon  = "🛵" if btype == "Delivery" else "🛍"
        dlg = self._make_dialog(420, 420, f"{icon}  Customer Details")
        win = dlg["win"]
        f   = dlg["frame"]
 
        # Header strip
        hdr = tk.Frame(f, bg=self.C["accent"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"Enter Customer Details  ({btype})",
                 font=("Arial", 12, "bold"),
                 bg=self.C["accent"], fg="black").pack()
 
        body = tk.Frame(f, bg=self.C["card"], padx=28, pady=16)
        body.pack(fill="both", expand=True)
 
        name_var  = tk.StringVar()
        phone_var = tk.StringVar()
        addr_var  = tk.StringVar()
 
        btype2 = self.bill_type_var.get()
        addr_label = "Delivery Address *" if btype2 == "Delivery" else "Address (optional)"
        fields = [
            ("Customer Name *", name_var,  False),
            ("Phone Number *",  phone_var, False),
            (addr_label,        addr_var,  False),
        ]
 
        entries = []
        for label, var, _ in fields:
            tk.Label(body, text=label, font=("Arial", 10),
                     bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
            e = tk.Entry(body, textvariable=var, font=("Arial", 12),
                         bg=self.C["bg"], fg=self.C["text"],
                         insertbackground=self.C["accent"],
                         bd=0,
                         highlightbackground=self.C["accent"],
                         highlightthickness=1)
            e.pack(fill="x", ipady=8)
            entries.append(e)
 
        status_lbl = tk.Label(body, text="", font=("Arial", 9),
                               bg=self.C["card"], fg=self.C["red"])
        status_lbl.pack(anchor="w", pady=(6, 0))
 
        btn_row = tk.Frame(f, bg=self.C["card"], padx=28, pady=12)
        btn_row.pack(fill="x")
 
        def confirm(event=None):
            n = name_var.get().strip()
            p = phone_var.get().strip()
            a = addr_var.get().strip()
            if not n:
                status_lbl.config(text="Customer name is required."); return
            if not p:
                status_lbl.config(text="Phone number is required."); return
            btype3 = self.bill_type_var.get()
            if not a and btype3 == "Delivery":
                status_lbl.config(text="Delivery address is required."); return
            dlg["close"]()
            on_confirm(n, p, a)
 
        entries[0].focus_set()
        for e in entries[:-1]:
            e.bind("<Return>", lambda ev, nxt=entries[entries.index(e)+1]: nxt.focus_set())
        entries[-1].bind("<Return>", confirm)
 
        tk.Button(btn_row, text="✔  Confirm Order",
                  command=confirm,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Button(btn_row, text="✕  Cancel",
                  command=dlg["close"],
                  bg=self.C["red"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
        win.bind("<Escape>", lambda e: dlg["close"]())
 
    def _on_table_confirm(self, event=None):
        btype    = self.bill_type_var.get()
        table_no = self.table_no_var.get().strip()
        table_enabled = self.db.get_setting("enable_table", "1") == "1"
 
        order_type_map = {
            "Dine-In":  "dine_in",
            "Takeaway": "takeaway",
            "Delivery": "delivery",
        }
        order_type_db = order_type_map.get(btype, "dine_in")
 
        if btype == "Dine-In":
            if table_enabled and not table_no:
                messagebox.showwarning("Table No", "Please enter a Table Number."); return
            if not table_enabled:
                table_no = f"BILL-{uuid.uuid4().hex[:4].upper()}"
            existing = self.db.get_open_order_for_table(table_no)
            if existing:
                self.active_order_id     = existing["id"]
                self.active_order_number = existing["order_number"]
                self.active_table        = table_no
                self._update_status(); self.refresh_order_display(); return
            oid, onum = self.db.create_order(table_no, order_type_db)
            self.active_order_id, self.active_order_number = oid, onum
            self.active_table = table_no
            self._update_status()
            self.refresh_order_display()
 
        elif btype == "Delivery":
            # Ask customer details upfront for delivery
            def _create_del(n, p, a):
                tbl   = f"DEL-{uuid.uuid4().hex[:4].upper()}"
                notes = f"Phone: {p} | Address: {a}"
                oid, onum = self.db.create_order(tbl, "delivery",
                                                 customer_name=n, notes=notes)
                self.active_order_id, self.active_order_number = oid, onum
                self.active_table = tbl
                self._update_status()
                self.refresh_order_display()
            self._delivery_customer_popup(_create_del)
            return
 
        else:  # Takeaway
            tbl    = f"TKWY-{uuid.uuid4().hex[:4].upper()}"
            oid, onum = self.db.create_order(tbl, order_type_db,
                                              customer_name=table_no or "Walk-in")
            self.active_order_id, self.active_order_number = oid, onum
            self.active_table = tbl
            self._update_status()
            self.refresh_order_display()
 
    def _update_status(self):
        if self.active_order_id:
            order_obj  = self.db.get_order(self.active_order_id)
            token_num  = dict(order_obj).get("token_number", 0) if order_obj else 0
            token_part = f"  🎫 #{token_num:03d}" if token_num else ""
            table_enabled = self.db.get_setting("enable_table", "1") == "1"
            if table_enabled:
                badge = f"  {self.active_table}  |  {self.bill_type_var.get()}{token_part}  |  {self.active_order_number}  "
            else:
                badge = f"  {self.bill_type_var.get()}{token_part}  |  {self.active_order_number}  "
            self.order_status_lbl.config(text=badge, bg=self.C["accent"], fg="black")
            # Show cancel button next to order status
            self.cancel_order_btn.pack(side="left", padx=(8, 16), pady=10, after=self.order_status_lbl)
        else:
            self.order_status_lbl.config(
                text="  No active order  ",
                bg=self.C["border"], fg=self.C["muted"])
            # Hide cancel button
            self.cancel_order_btn.pack_forget()
 
    def refresh_menu_display(self):
        self.menu_listbox.delete(0, "end")
        self._menu_items_cache = []
        cat_enabled = self.db.get_setting("enable_category", "1") == "1"
        cat   = self.cat_var.get() if cat_enabled else "All"
        query = self.search_var.get().strip().lower()
        try:
            c = self.db.conn.cursor()
            if cat == "All":
                c.execute("SELECT name, price, category, available FROM menu_items ORDER BY name")
            else:
                c.execute("SELECT name, price, category, available FROM menu_items WHERE category=? ORDER BY name", (cat,))
            items = c.fetchall()
        except Exception:
            items = []
        if query:
            items = [i for i in items if query in i["name"].lower()]
        curr = self.db.get_setting("currency", "Rs.")
        for item in items:
            if item["available"]:
                cat_tag = f"  [{item['category']}]" if cat_enabled else ""
                line = f"  {item['name']:<28}  {curr}{item['price']:.2f}{cat_tag}"
            else:
                line = f"  ✗ {item['name']:<26}  (unavailable)"
            self.menu_listbox.insert("end", line)
            self._menu_items_cache.append(dict(item))
        self.cat_combo["values"] = self._get_categories()
 
    def add_selected_item(self):
        sel = self.menu_listbox.curselection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select an item from the menu list."); return
        item = self._menu_items_cache[sel[0]]
        if not item["available"]:
            messagebox.showwarning("Unavailable", f"'{item['name']}' is not available."); return
        if not self.active_order_id:
            btype = self.bill_type_var.get()
            if btype == "Dine-In":
                if self.db.get_setting("enable_table", "1") == "1":
                    messagebox.showwarning("No Order",
                        "Enter a Table No and press Enter / 'Open Table' to start an order."); return
                else:
                    self.open_order_if_no_table()
            else:
                # Delivery/Takeaway order should already exist from bill-type selection.
                # If somehow it doesn't, guard gracefully.
                messagebox.showwarning("No Order",
                    "Please select the order type first to start an order."); return
        self._qty_popup(item["name"], item["price"], item["category"])
 
    def _qty_popup(self, item_name, default_price, category):
        dlg  = self._make_dialog(320, 310, "Add Item")
        win  = dlg["win"]
        f    = dlg["frame"]
        curr = self.db.get_setting("currency", "Rs.")
 
        hdr = tk.Frame(f, bg=self.C["accent"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=item_name,
                 font=("Arial", 13, "bold"),
                 bg=self.C["accent"], fg="black",
                 wraplength=280).pack(padx=12)
        if self.db.get_setting("enable_category", "1") == "1":
            tk.Label(hdr, text=f"Category: {category}",
                     font=("Arial", 9),
                     bg=self.C["accent"], fg="#333").pack()
 
        body = tk.Frame(f, bg=self.C["card"], padx=24, pady=14)
        body.pack(fill="both", expand=True)
 
        qty_row = tk.Frame(body, bg=self.C["card"])
        qty_row.pack(fill="x", pady=(0, 10))
        tk.Label(qty_row, text="Quantity :", width=14, anchor="w",
                 font=("Arial", 11, "bold"),
                 bg=self.C["card"], fg=self.C["muted"]).pack(side="left")
        qty_var = tk.StringVar(value="1")
        qty_e = tk.Entry(qty_row, textvariable=qty_var,
                         width=8, font=("Arial", 13, "bold"),
                         bg=self.C["bg"], fg=self.C["accent"],
                         insertbackground=self.C["accent"],
                         justify="center", bd=0,
                         highlightbackground=self.C["accent"],
                         highlightthickness=1)
        qty_e.pack(side="left", ipady=6)
        qty_e.focus_set()
        qty_e.select_range(0, "end")
 
        price_row = tk.Frame(body, bg=self.C["card"])
        price_row.pack(fill="x", pady=(0, 10))
        tk.Label(price_row, text=f"Price ({curr}) :", width=14, anchor="w",
                 font=("Arial", 11, "bold"),
                 bg=self.C["card"], fg=self.C["muted"]).pack(side="left")
        price_var = tk.StringVar(value=str(default_price))
        tk.Entry(price_row, textvariable=price_var,
                 width=8, font=("Arial", 13, "bold"),
                 bg=self.C["bg"], fg=self.C["accent"],
                 insertbackground=self.C["accent"],
                 justify="center", bd=0,
                 highlightbackground=self.C["accent"],
                 highlightthickness=1).pack(side="left", ipady=6)
 
        tk.Label(body,
                 text=f"Actual price: {curr}{default_price:.2f}  (edit if needed)",
                 font=("Arial", 8), bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
 
        btn_row = tk.Frame(f, bg=self.C["card"], pady=10, padx=24)
        btn_row.pack(fill="x")
 
        def confirm(event=None):
            try:
                q = int(qty_var.get())
                p = float(price_var.get())
                if q <= 0 or p < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Invalid",
                    "Enter a valid quantity (>=1) and price.", parent=win)
                return
            # Check if a KOT was already printed — warn cashier to re-print
            kot_already_printed = bool(
                self.db.get_order_items(self.active_order_id)
                and any(i["kot_printed"] == 1
                        for i in self.db.get_order_items(self.active_order_id))
            )
            self.db.add_order_item(self.active_order_id, item_name, category, q, p)
            self.refresh_order_display()
            dlg["close"]()
            if kot_already_printed:
                messagebox.showwarning(
                    "\u26a0\ufe0f  Re-print KOT",
                    f"'{item_name}' added after KOT was already sent to kitchen.\n\n"
                    "Please print a new KOT so the kitchen sees this item."
                )
 
        tk.Button(btn_row, text="Add to Orders",
                  command=confirm,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 11, "bold"), bd=0, pady=8,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 6))
 
        tk.Button(btn_row, text="Cancel",
                  command=dlg["close"],
                  bg=self.C["red"], fg="white",
                  font=("Arial", 11, "bold"), bd=0, pady=8,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
        qty_e.bind("<Return>", confirm)
        win.bind("<Escape>", lambda e: dlg["close"]())
 
    def _add_extra_item_popup(self):
        if not self.active_order_id:
            btype = self.bill_type_var.get()
            if btype == "Dine-In":
                if self.db.get_setting("enable_table", "1") == "1":
                    messagebox.showwarning("No Order", "Please open a table first."); return
                else:
                    self.open_order_if_no_table()
            else:
                messagebox.showwarning("No Order",
                    "Please select the order type first to start an order."); return
        curr = self.db.get_setting("currency", "Rs.")
        dlg  = self._make_dialog(340, 330, "Add Extra Item")
        win  = dlg["win"]
        f    = dlg["frame"]
 
        tk.Label(f, text="Item Name *", font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", padx=20, pady=(14, 2))
        name_var = tk.StringVar()
        name_e = tk.Entry(f, textvariable=name_var, font=("Arial", 13),
                          bg=self.C["bg"], fg=self.C["text"],
                          insertbackground=self.C["accent"], bd=0,
                          highlightbackground=self.C["accent"],
                          highlightthickness=1)
        name_e.pack(fill="x", padx=20, ipady=7)
        name_e.focus_set()
 
        qty_price_row = tk.Frame(f, bg=self.C["card"])
        qty_price_row.pack(fill="x", padx=20, pady=(10, 0))
 
        qty_col = tk.Frame(qty_price_row, bg=self.C["card"])
        qty_col.pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Label(qty_col, text="Quantity *", font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(0, 2))
        qty_var = tk.StringVar(value="1")
        tk.Entry(qty_col, textvariable=qty_var, font=("Arial", 13),
                 bg=self.C["bg"], fg=self.C["accent"],
                 insertbackground=self.C["accent"],
                 justify="center", bd=0,
                 highlightbackground=self.C["accent"],
                 highlightthickness=1).pack(fill="x", ipady=7)
 
        price_col = tk.Frame(qty_price_row, bg=self.C["card"])
        price_col.pack(side="left", expand=True, fill="x")
        tk.Label(price_col, text=f"Price ({curr}) *", font=("Arial", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(0, 2))
        price_var = tk.StringVar(value="0")
        tk.Entry(price_col, textvariable=price_var, font=("Arial", 13),
                 bg=self.C["bg"], fg=self.C["accent"],
                 insertbackground=self.C["accent"],
                 justify="center", bd=0,
                 highlightbackground=self.C["accent"],
                 highlightthickness=1).pack(fill="x", ipady=7)
 
        status_lbl = tk.Label(f, text="", font=("Arial", 9),
                               bg=self.C["card"], fg=self.C["red"])
        status_lbl.pack(anchor="w", padx=20, pady=(6, 0))
 
        btn_row = tk.Frame(f, bg=self.C["card"], padx=20, pady=12)
        btn_row.pack(fill="x")
 
        def save(event=None):
            name = name_var.get().strip()
            if not name:
                status_lbl.config(text="Item name is required."); return
            try:
                q = int(qty_var.get())
                p = float(price_var.get())
                if q <= 0: raise ValueError
                if p < 0:  raise ValueError
            except ValueError:
                status_lbl.config(text="Enter valid quantity (>=1) and price (>=0)."); return
            kot_already_printed = bool(
                self.db.get_order_items(self.active_order_id)
                and any(i["kot_printed"] == 1
                        for i in self.db.get_order_items(self.active_order_id))
            )
            self.db.add_order_item(self.active_order_id, name, "Extra", q, p)
            self.refresh_order_display()
            dlg["close"]()
            if kot_already_printed:
                messagebox.showwarning(
                    "\u26a0️  Re-print KOT",
                    f"'{name}' added after KOT was already sent to kitchen.\n\n"
                    "Please print a new KOT so the kitchen sees this item."
                )

        tk.Button(btn_row, text="Save",
                  command=save,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Button(btn_row, text="Cancel",
                  command=dlg["close"],
                  bg=self.C["red"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
        name_e.bind("<Return>", save)
        win.bind("<Escape>", lambda e: dlg["close"]())
 
    def refresh_order_display(self):
        self.cart_listbox.delete(0, "end")
        self._order_item_ids = []
        curr = self.db.get_setting("currency", "Rs.")
        if not self.active_order_id:
            self.total_lbl.config(text=f"Total: {curr} 0.00")
            return
        items    = self.db.get_order_items(self.active_order_id)
        subtotal = 0
        for item in items:
            line = (f"  {item['quantity']}x  {item['item_name'][:22]:<22}"
                    f"  {curr}{item['total']:.2f}")
            self.cart_listbox.insert("end", line)
            self._order_item_ids.append(item["id"])
            subtotal += item["total"]
        disc  = self._get_discount()
 
        sgst_pct = float(self.db.get_setting("sgst_percent", "0"))
        cgst_pct = float(self.db.get_setting("cgst_percent", "0"))
        tax_p    = float(self.db.get_setting("tax_percent",  "0"))
        svc_p    = float(self.db.get_setting("service_charge", "0"))
 
        combined_gst = sgst_pct + cgst_pct if (sgst_pct + cgst_pct) > 0 else tax_p
        taxable_base = round((subtotal - disc) / (1 + combined_gst / 100), 2) if combined_gst > 0 else (subtotal - disc)
        tax   = round(taxable_base * combined_gst / 100, 2)
        svc   = round(taxable_base * svc_p / 100, 2)
        total = round(taxable_base + tax + svc, 2)
        self.total_lbl.config(text=f"Total: {curr} {total:.2f}")
 
    def _get_discount(self):
        try: return float(self.discount_var.get())
        except Exception: return 0.0
 
    def remove_selected_cart_item(self):
        sel = self.cart_listbox.curselection()
        if not sel:
            messagebox.showwarning("Select", "Select an item in the cart to remove."); return
        self.db.remove_order_item(self._order_item_ids[sel[0]])
        self.refresh_order_display()
 
    def _edit_cart_item(self):
        sel = self.cart_listbox.curselection()
        if not sel: return
        db_id = self._order_item_ids[sel[0]]
        dlg = self._make_dialog(280, 230, "Edit Quantity")
        f   = dlg["frame"]
        qty_var = tk.StringVar(value="1")
        tk.Label(f, text="New Quantity (0 = remove):", font=("Arial", 9),
                 bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", padx=16, pady=(16, 2))
        qty_e = tk.Entry(f, textvariable=qty_var, font=("Arial", 12),
                          bg=self.C["bg"], fg=self.C["text"],
                          insertbackground=self.C["accent"], bd=0)
        qty_e.pack(fill="x", padx=16, ipady=5); qty_e.focus_set()
        def confirm(event=None):
            try:
                q = int(qty_var.get())
            except ValueError:
                messagebox.showerror("Invalid", "Enter a valid number.", parent=dlg["win"]); return
            if q <= 0: self.db.remove_order_item(db_id)
            else: self.db.update_order_item_qty(db_id, q)
            self.refresh_order_display(); dlg["close"]()
        qty_e.bind("<Return>", confirm)
        btn_r = tk.Frame(f, bg=self.C["card"], padx=16, pady=8)
        btn_r.pack(fill="x")
        tk.Button(btn_r, text="Save", command=confirm,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 11, "bold"), bd=0, pady=8,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 6))
        tk.Button(btn_r, text="Cancel", command=dlg["close"],
                  bg=self.C["red"], fg="white",
                  font=("Arial", 11, "bold"), bd=0, pady=8,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
    def clear_order(self):
        if not self.active_order_id: return
        all_items = self.db.get_order_items(self.active_order_id)
        if not all_items:
            return

        if not messagebox.askyesno("Clear Order", f"Clear cart and save order to Open Orders?"):
            return

        # Only remove items that were NOT sent to kitchen
        new_items = [i for i in all_items if i["kot_printed"] == 0]
        kot_items = [i for i in all_items if i["kot_printed"] == 1]

        for item in new_items:
            self.db.remove_order_item(item["id"])

        # If there are KOT printed items, keep the order open in DB
        # If nothing was KOT printed, delete the order entirely
        if not kot_items:
            order = self.db.get_order(self.active_order_id)
            if order and order["order_type"] == "dine_in":
                self.db.set_table_status(order["table_number"], "free")
            c = self.db.conn.cursor()
            c.execute("DELETE FROM orders WHERE id=?", (self.active_order_id,))
            self.db.conn.commit()

        # Reset screen
        self.active_order_id = None
        self.active_order_number = None
        self.active_table = None
        self.table_no_var.set("")
        self._block_discount_trace = True
        self.discount_var.set("0")
        self._block_discount_trace = False
        self._order_item_ids = []
        self.cart_listbox.delete(0, "end")
        curr = self.db.get_setting("currency", "Rs.")
        self.total_lbl.config(text=f"Total: {curr} 0.00")
        self._update_status()
    def add_to_order(self, item_name, price, category):
        if self.active_order_id:
            self.db.add_order_item(self.active_order_id, item_name, category, 1, price)
            self.refresh_order_display()
 
    def print_kot(self):
        if not self.active_order_id:
            messagebox.showwarning("No Order", "No active order!"); return
        items = self.db.get_unprinted_kot_items(self.active_order_id)
        if not items:
            messagebox.showinfo("KOT", "No new items to send to kitchen."); return
        order = self.db.get_order(self.active_order_id)
        lines = self.formatter.format_kot(self.active_order_id, items, dict(order))
        self._show_kot_preview(lines, len(items))
 
    def _show_kot_preview(self, lines, item_count):
        """Show KOT in a preview popup with a Print button — same as bill preview."""
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])
        W, H = 420, 530
        x = (win.winfo_screenwidth()  - W) // 2
        y = (win.winfo_screenheight() - H) // 2
        win.geometry(f"{W}x{H}+{x}+{y}")
        win.grab_set()
 
        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=self.C["card"])
        inner.pack(fill="both", expand=True)
 
        # Title bar
        tbar = tk.Frame(inner, bg="#FFA500", pady=8)
        tbar.pack(fill="x")
        order_obj = self.db.get_order(self.active_order_id) if self.active_order_id else None
        token_num = dict(order_obj).get("token_number", 0) if order_obj else 0
        title_txt = f"🍳  KOT  —  Token #{token_num:03d}" if token_num else "🍳  KITCHEN ORDER TICKET"
        tk.Label(tbar, text=title_txt,
                 font=("Arial", 13, "bold"),
                 bg="#FFA500", fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                  bg="#FFA500", fg="black", font=("Arial", 11, "bold"),
                  bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)
 
        # KOT text
        txt_frame = tk.Frame(inner, bg=self.C["card"])
        txt_frame.pack(fill="both", expand=True, padx=6, pady=6)
        txt = tk.Text(txt_frame, font=("Courier New", 11),
                      bg=self.C["card"], fg=self.C["text"],
                      bd=0, padx=12, pady=10)
        sb = tk.Scrollbar(txt_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(fill="both", expand=True)
        txt.insert("1.0", "\n".join(lines))
        txt.configure(state="disabled")
 
        # Buttons
        btn_frame = tk.Frame(inner, bg=self.C["card"])
        btn_frame.pack(fill="x", padx=6, pady=(0, 6))
 
        cfg = self.printer.get_config()
        is_thermal = cfg.get("method") == "Thermal Printer"
 
        def send_to_kitchen():
            self.db.mark_kot_printed(self.active_order_id)
            win.destroy()
            self.printer.print_lines(lines, "KOT")

            order_obj  = self.db.get_order(self.active_order_id)
            order_type = dict(order_obj).get("order_type", "dine_in") if order_obj else "dine_in"

            if order_type != "dine_in":
                # Counter service: reset screen, order stays in Open Orders for billing
                self.active_order_id     = None
                self.active_order_number = None
                self.active_table        = None
                self.table_no_var.set("")
                self._block_discount_trace = True
                self.discount_var.set("0")
                self._block_discount_trace = False
                self._order_item_ids = []
                self.cart_listbox.delete(0, "end")
                curr = self.db.get_setting("currency", "Rs.")
                self.total_lbl.config(text=f"Total: {curr} 0.00")
                self._update_status()
                if not is_thermal:
                    messagebox.showinfo("KOT Sent",
                        f"{item_count} item(s) sent to kitchen.\n\n"
                        "Screen reset for next customer.\n"
                        "Use  📋 Open Orders  to bill this order.")
 
        btn_label = "🖨  Print (Thermal)" if is_thermal else "🖨  Print (Browser)"
        tk.Button(btn_frame, text=btn_label,
                  command=send_to_kitchen,
                  bg="#FFA500", fg="black",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 4))
        tk.Button(btn_frame, text="Close",
                  command=win.destroy,
                  bg=self.C["border"], fg=self.C["text"],
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
 
    def reprint_last_receipt(self):
        if not self._last_bill_lines:
            messagebox.showinfo("No Receipt", "No receipt printed in this session."); return
        self._show_bill_preview(self._last_bill_lines, self._last_bill_obj)
 
    def generate_bill(self):
        if not self.active_order_id:
            messagebox.showwarning("No Order", "Please open a table / order first."); return
        items = self.db.get_order_items(self.active_order_id)
        if not items:
            messagebox.showwarning("Empty Cart", "No items in the cart!"); return
 
        # Ask for customer details before billing for Delivery and Takeaway
        order = self.db.get_order(self.active_order_id)
        otype = order["order_type"]
        has_details = bool(order["notes"]) or (
            bool(order["customer_name"]) and order["customer_name"] != "Walk-in"
        )
        # Delivery details already collected when adding first item
        # Takeaway — no customer details needed
 
        self._finalise_bill()
 
    def _save_delivery_info_and_bill(self, cust_name, cust_phone, cust_addr):
        """Save customer info to order then finalise bill."""
        now   = datetime.now().isoformat()
        notes = f"Phone: {cust_phone} | Address: {cust_addr}"
        c = self.db.conn.cursor()
        c.execute("UPDATE orders SET customer_name=?, notes=?, updated_at=? WHERE id=?",
                  (cust_name, notes, now, self.active_order_id))
        self.db.conn.commit()
        self._finalise_bill()
    def _finalise_bill(self):
        if not self.active_order_id:
            return
        items = self.db.get_order_items(self.active_order_id)
        if not items:
            messagebox.showwarning("Empty Cart", "No items in the cart!"); return

        order_row  = self.db.get_order(self.active_order_id) or {}
        extra_info = {
            "customer_name": dict(order_row).get("customer_name", "") if order_row else "",
            "notes":         dict(order_row).get("notes", "")         if order_row else "",
            "token_number":  dict(order_row).get("token_number", 0)   if order_row else 0,
        }
        bill, err = self.db.create_bill(self.active_order_id,
                                        self._get_discount(),
                                        self.payment_var.get())
        if err:
            messagebox.showerror("Billing Error", err); return

        items_data = json.loads(bill["items_json"])
        lines = self.formatter.format_bill(dict(bill), items_data, extra_info=extra_info)
        self._last_bill_lines = lines
        self._last_bill_obj   = bill
        try:
            with open(LAST_BILL_FILE, "w", encoding="utf-8") as f:
                json.dump({"lines": lines, "bill": dict(bill)}, f, indent=2)
        except Exception:
            pass

        # Reset screen BEFORE showing preview
        self.active_order_id = None
        self.active_order_number = None
        self.active_table = None
        self.table_no_var.set("")
        self._block_discount_trace = True
        self.discount_var.set("0")
        self._block_discount_trace = False
        self._order_item_ids = []
        self.cart_listbox.delete(0, "end")
        curr = self.db.get_setting("currency", "Rs.")
        self.total_lbl.config(text=f"Total: {curr} 0.00")
        self._update_status()

        self._show_bill_preview(lines, bill)

    def _show_bill_preview(self, lines, bill):
        curr = self.db.get_setting("currency", "Rs.")
        # ── Rounded popup with accent border ─────────────────
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])   # accent = border color
        W, H = 500, 660
        x = (win.winfo_screenwidth()  - W) // 2
        y = (win.winfo_screenheight() - H) // 2
        win.geometry(f"{W}x{H}+{x}+{y}")
        win.grab_set()
 
        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=self.C["card"])
        inner.pack(fill="both", expand=True)
 
        tbar = tk.Frame(inner, bg=self.C["accent"], pady=8)
        tbar.pack(fill="x")
        tk.Label(tbar, text="BILL RECEIPT", font=("Arial", 14, "bold"),
                bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                bg=self.C["accent"], fg="black", font=("Arial", 11, "bold"),
                bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)
 
        txt_frame = tk.Frame(inner, bg=self.C["card"])
        txt_frame.pack(fill="both", expand=True, padx=6, pady=6)
        txt = tk.Text(txt_frame, font=("Courier New", 10),
                    bg=self.C["card"], fg=self.C["text"], bd=0, padx=12, pady=10)
        sb = tk.Scrollbar(txt_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(fill="both", expand=True)
        btn_inner = tk.Frame(txt_frame, bg=self.C["card"])
        btn_inner.pack(fill="x", pady=(0, 6), padx=12)
        def _print_bill():
            self.printer.print_lines(lines, "Bill")
            win.destroy()
 
        tk.Button(btn_inner, text="🖨  Print Bill",
                command=_print_bill,
                bg=self.C["green"], fg="white",
                font=("Arial", 11, "bold"), bd=0, pady=8,
                cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 4))
        tk.Button(btn_inner, text="Close", command=win.destroy,
                bg=self.C["border"], fg=self.C["text"],
                font=("Arial", 11, "bold"), bd=0, pady=8,
                cursor="hand2").pack(side="left", expand=True, fill="x")
        txt.insert("1.0", "\n".join(lines))
        txt.configure(state="disabled")
 
    def refresh_tables(self): pass
    def select_table(self, *a): pass
    def _new_table_order(self, *a): pass
    def new_non_table_order(self): pass
    def _update_order_header(self): self._update_status()
    def set_customer_info(self): pass
    def on_order_item_dbl_click(self, event): pass
    def open_menu_mgmt(self): self._show_page("menu")
 
    def open_orders_popup(self):
        """Show all open orders — click any to switch to it for billing."""
        orders = self.db.get_all_open_orders()
 
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])
        W, H = 660, 500
        self._center_win(win, W, H)
        win.grab_set()
 
        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=self.C["bg"])
        inner.pack(fill="both", expand=True)
 
        # Title bar
        tbar = tk.Frame(inner, bg=self.C["accent"], pady=8)
        tbar.pack(fill="x")
        tk.Label(tbar, text="📋  Open Orders  —  Select to Bill",
                 font=("Arial", 13, "bold"),
                 bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                  bg=self.C["accent"], fg="black", font=("Arial", 11, "bold"),
                  bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)
 
        # Search bar
        search_frame = tk.Frame(inner, bg=self.C["bg"], pady=6)
        search_frame.pack(fill="x", padx=10)
        tk.Label(search_frame, text="🔍", font=("Arial", 12),
                 bg=self.C["bg"], fg=self.C["muted"]).pack(side="left", padx=(0, 6))
        search_var = tk.StringVar()
        tk.Entry(search_frame, textvariable=search_var,
                 font=("Arial", 11), bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"],
                 highlightbackground=self.C["accent"], highlightthickness=1,
                 bd=0).pack(side="left", ipady=6, fill="x", expand=True)
 
        if not orders:
            tk.Label(inner, text="No open orders at the moment.",
                     font=("Arial", 13), bg=self.C["bg"],
                     fg=self.C["muted"]).pack(expand=True)
            return
 
        # Scrollable list
        list_outer = tk.Frame(inner, bg=self.C["bg"])
        list_outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(list_outer, bg=self.C["bg"], highlightthickness=0)
        sb = tk.Scrollbar(list_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=8, pady=4)
        frame = tk.Frame(canvas, bg=self.C["bg"])
        win_id = canvas.create_window((0, 0), window=frame, anchor="nw")
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
 
        type_labels = {"dine_in": "Dine-In", "takeaway": "Takeaway", "delivery": "Delivery"}
        all_orders  = list(orders)
 
        def build_cards(filter_text=""):
            for w in frame.winfo_children():
                w.destroy()
            q = filter_text.strip().lower()
            shown = 0
            for order in all_orders:
                o       = dict(order)
                token   = o.get("token_number", 0) or 0
                otype   = type_labels.get(o["order_type"], o["order_type"])
                cname   = o.get("customer_name", "") or ""
                tbl     = o.get("table_number", "") or ""
                created = o.get("created_at", "")[:16].replace("T", "  ")
                searchable = f"{token:03d} {cname} {tbl} {otype} {o['order_number']}".lower()
                if q and q not in searchable:
                    continue
                shown += 1
 
                c = self.db.conn.cursor()
                c.execute("SELECT SUM(quantity) FROM order_items WHERE order_id=?", (o["id"],))
                qty_row    = c.fetchone()
                item_count = int(qty_row[0]) if qty_row and qty_row[0] else 0
 
                card = tk.Frame(frame, bg=self.C["card"],
                                highlightbackground=self.C["border"],
                                highlightthickness=1)
                card.pack(fill="x", pady=3, padx=4)
 
                token_f = tk.Frame(card, bg="#FFA500", width=80)
                token_f.pack(side="left", fill="y")
                token_f.pack_propagate(False)
                tk.Label(token_f,
                         text=f"#{token:03d}" if token else "—",
                         font=("Georgia", 20, "bold"),
                         bg="#FFA500", fg="black").pack(expand=True)
 
                info = tk.Frame(card, bg=self.C["card"], padx=14, pady=10)
                info.pack(side="left", fill="both", expand=True)
                row1 = tk.Frame(info, bg=self.C["card"])
                row1.pack(anchor="w")
                tk.Label(row1, text=otype, font=("Arial", 12, "bold"),
                         bg=self.C["card"], fg=self.C["accent"]).pack(side="left")
                if tbl and not any(tbl.startswith(p) for p in ("TKWY-","DEL-","BILL-")):
                    tk.Label(row1, text=f"   Table: {tbl}",
                             font=("Arial", 11),
                             bg=self.C["card"], fg=self.C["text"]).pack(side="left")
                if cname and cname != "Walk-in":
                    tk.Label(row1, text=f"   👤 {cname}",
                             font=("Arial", 10),
                             bg=self.C["card"], fg=self.C["muted"]).pack(side="left")
                tk.Label(info,
                         text=f"{o['order_number']}   ·   {item_count} item(s)   ·   {created}",
                         font=("Arial", 9),
                         bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w", pady=(2, 0))
 
                def _select(order_id=o["id"], order_number=o["order_number"],
                            table=tbl, w=win):
                    # Set active order FIRST so any trace callbacks
                    # (discount_var etc.) find a valid order
                    self.active_order_id     = order_id
                    self.active_order_number = order_number
                    self.active_table        = table
                    # Now sync bill type dropdown — suppress the full reset
                    self._suppress_bill_type_reset = True
                    try:
                        ot = self.db.get_order(order_id)
                        if ot:
                            type_map = {"dine_in": "Dine-In", "takeaway": "Takeaway", "delivery": "Delivery"}
                            self.bill_type_var.set(type_map.get(dict(ot)["order_type"], "Dine-In"))
                    finally:
                        self._suppress_bill_type_reset = False
                    self._update_status()
                    self.refresh_order_display()
                    w.destroy()
 
                tk.Button(card, text="Select\n→ Bill",
                          command=_select,
                          bg=self.C["green"], fg="white",
                          font=("Arial", 11, "bold"), bd=0,
                          padx=18, cursor="hand2").pack(side="right", fill="y")
 
            if shown == 0:
                tk.Label(frame, text="No orders match your search.",
                         font=("Arial", 11), bg=self.C["bg"],
                         fg=self.C["muted"]).pack(pady=20)
            canvas.configure(scrollregion=canvas.bbox("all"))
 
        build_cards()
        search_var.trace("w", lambda *_: build_cards(search_var.get()))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))
 
    def open_settings(self):
        # ── Rounded popup with yellow (accent) border ─────────
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])   # yellow border visible at edges
        W, H = 520, 640
        self._center_win(win, W, H)
        win.grab_set()

        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner_bg = tk.Frame(border, bg=self.C["bg"])
        inner_bg.pack(fill="both", expand=True)

        tbar = tk.Frame(inner_bg, bg=self.C["accent"], pady=8)
        tbar.pack(fill="x")
        tk.Label(tbar, text="⚙  Restaurant Settings",
                 font=("Arial", 14, "bold"),
                 bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                  bg=self.C["accent"], fg="black", font=("Arial", 11, "bold"),
                  bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)

        canvas    = tk.Canvas(inner_bg, bg=self.C["bg"], highlightthickness=0)
        scrollbar = tk.Scrollbar(inner_bg, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        frame    = tk.Frame(canvas, bg=self.C["bg"], padx=30, pady=20)
        frame_id = canvas.create_window((0, 0), window=frame, anchor="nw")
        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_configure(e):
            canvas.itemconfig(frame_id, width=e.width)
        frame.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))

        fields = [
            ("Restaurant Name", "restaurant_name"),
            ("Address",         "address"),
            ("Phone",           "phone"),
            ("GST Number",      "gst_number"),
            ("Service Charge %","service_charge"),
            ("Currency Symbol", "currency"),
            ("Footer Message",  "footer_message"),
            ("Mobile Order URL (e.g. Vercel)", "mobile_url_override"),
        ]
        entries = {}
        for label, key in fields:
            tk.Label(frame, text=label + ":", font=("Arial", 10),
                     bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
            var = tk.StringVar(value=self.db.get_setting(key))
            tk.Entry(frame, textvariable=var, font=("Arial", 11),
                     bg=self.C["card"], fg=self.C["text"],
                     insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=6)
            entries[key] = var

        tk.Frame(frame, bg=self.C["border"], height=1).pack(fill="x", pady=(14, 4))
        tk.Label(frame, text="GST Settings",
                 font=("Arial", 10, "bold"),
                 bg=self.C["bg"], fg=self.C["accent"]).pack(anchor="w", pady=(0, 2))
        tk.Label(frame,
                 text="Enter GST % to auto-split into SGST+CGST, or set SGST+CGST to compute GST.",
                 font=("Arial", 8), bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w", pady=(0, 6))

        gst_var  = tk.StringVar(value=self.db.get_setting("tax_percent",  "5.0"))
        sgst_var = tk.StringVar(value=self.db.get_setting("sgst_percent", "2.5"))
        cgst_var = tk.StringVar(value=self.db.get_setting("cgst_percent", "2.5"))
        entries["tax_percent"]  = gst_var
        entries["sgst_percent"] = sgst_var
        entries["cgst_percent"] = cgst_var

        _syncing = [False]

        def _sync_from_gst(*_):
            if _syncing[0]: return
            _syncing[0] = True
            try:
                g    = float(gst_var.get())
                half = round(g / 2, 2)
                sgst_var.set(str(half))
                cgst_var.set(str(round(g - half, 2)))
            except ValueError:
                pass
            finally:
                _syncing[0] = False

        def _sync_from_sgst_cgst(*_):
            if _syncing[0]: return
            _syncing[0] = True
            try:
                total_gst = round(float(sgst_var.get()) + float(cgst_var.get()), 2)
                gst_var.set(str(total_gst))
            except ValueError:
                pass
            finally:
                _syncing[0] = False

        tk.Label(frame, text="GST %  (auto-splits into SGST + CGST):", font=("Arial", 10),
                 bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
        tk.Entry(frame, textvariable=gst_var, font=("Arial", 11),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=6)
        gst_var.trace("w", _sync_from_gst)

        tk.Label(frame, text="SGST %:", font=("Arial", 10),
                 bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
        tk.Entry(frame, textvariable=sgst_var, font=("Arial", 11),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=6)
        sgst_var.trace("w", _sync_from_sgst_cgst)

        tk.Label(frame, text="CGST %:", font=("Arial", 10),
                 bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w", pady=(8, 2))
        tk.Entry(frame, textvariable=cgst_var, font=("Arial", 11),
                 bg=self.C["card"], fg=self.C["text"],
                 insertbackground=self.C["accent"], bd=0).pack(fill="x", ipady=6)
        cgst_var.trace("w", _sync_from_sgst_cgst)

        tk.Label(frame,
                 text="  Item prices are GST-inclusive. Bill shows pre-tax base + SGST + CGST = Total.",
                 font=("Arial", 8), bg=self.C["bg"], fg="#555566").pack(anchor="w", pady=(4, 0))

        tk.Frame(frame, bg=self.C["border"], height=1).pack(fill="x", pady=(14, 8))

        table_var = tk.IntVar(value=1 if self.db.get_setting("enable_table", "1") == "1" else 0)
        tk.Checkbutton(frame,
                       text="Enable Table Number  (uncheck for hotels without tables)",
                       variable=table_var,
                       font=("Arial", 11),
                       bg=self.C["bg"], fg=self.C["text"],
                       selectcolor=self.C["card"],
                       activebackground=self.C["bg"],
                       activeforeground=self.C["accent"]).pack(anchor="w")

        cat_var = tk.IntVar(value=1 if self.db.get_setting("enable_category", "1") == "1" else 0)
        tk.Checkbutton(frame,
                       text="Enable Categories  (uncheck for hotels without categories)",
                       variable=cat_var,
                       font=("Arial", 11),
                       bg=self.C["bg"], fg=self.C["text"],
                       selectcolor=self.C["card"],
                       activebackground=self.C["bg"],
                       activeforeground=self.C["accent"]).pack(anchor="w", pady=(6, 0))

        def save_settings():
            for key, var in entries.items():
                self.db.set_setting(key, var.get().strip())
            self.db.set_setting("enable_table",    "1" if table_var.get() else "0")
            self.db.set_setting("enable_category", "1" if cat_var.get() else "0")

            # Update table visibility
            if table_var.get() and self.bill_type_var.get() == "Dine-In":
                self.grp2.pack(side="left", padx=(10, 0), pady=10, ipady=4)
            else:
                self.grp2.pack_forget()
                self.table_no_var.set("")

            # Update category visibility in billing page
            if cat_var.get():
                self.cat_label.pack(before=self.search_label,
                                    side="left", padx=(10, 4), pady=5)
                self.cat_combo.pack(before=self.search_label,
                                    side="left", padx=(0, 12))
            else:
                self.cat_var.set("All")
                self.cat_label.pack_forget()
                self.cat_combo.pack_forget()

            # Update category visibility in menu management page
            self._mm_cat_container.pack_forget()
            if cat_var.get():
                self._mm_cat_container.pack(fill="x")

            self.refresh_menu_display()
            messagebox.showinfo("Saved", "Settings saved!", parent=win)
            win.destroy()

        tk.Frame(frame, bg=self.C["border"], height=1).pack(fill="x", pady=(14, 8))
        btn_row = tk.Frame(frame, bg=self.C["bg"])
        btn_row.pack(fill="x", pady=(0, 16))
        tk.Button(btn_row, text="💾  Save Settings", command=save_settings,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 8))
        tk.Button(btn_row, text="Cancel", command=win.destroy,
                  bg=self.C["red"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(side="left", expand=True, fill="x")
        
    def _save_report_pdf(self, date_str, rows):
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Missing Library",
                "reportlab is not installed.\nRun: pip install reportlab"); return
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        import tkinter.filedialog as fd
 
        path = fd.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            initialfile=f"Sales_Report_{date_str}.pdf",
            title="Save Report as PDF")
        if not path:
            return
        curr  = self.db.get_setting("currency", "Rs.")
        rname = self.db.get_setting("restaurant_name", "Restaurant")
        doc   = SimpleDocTemplate(path, pagesize=A4,
                                leftMargin=40, rightMargin=40,
                                topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        story  = []
        story.append(Paragraph(rname.upper(), styles["Title"]))
        story.append(Paragraph(f"Daily Sales Report — {date_str}", styles["Heading2"]))
        story.append(Spacer(1, 12))
 
        if not rows:
            story.append(Paragraph("No sales found for this date.", styles["Normal"]))
        else:
            headers = ["Bill No", "Table", "Type", "Subtotal", "Discount", "Tax", "Total", "Payment"]
            data    = [headers]
            grand   = 0
            for row in rows:
                data.append([
                    row["bill_number"][:16],
                    row["table_number"],
                    row["order_type"],
                    f"{curr}{row['subtotal']:.2f}",
                    f"{curr}{row['discount']:.2f}",
                    f"{curr}{row['tax']:.2f}",
                    f"{curr}{row['total']:.2f}",
                    row["payment_method"].upper(),
                ])
                grand += row["total"]
            data.append(["", "", "GRAND TOTAL", "", "", "", f"{curr}{grand:.2f}", ""])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#FF6B35")),
                ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,0), 9),
                ("BACKGROUND", (0,-1),(-1,-1), colors.HexColor("#FFA500")),
                ("FONTNAME",   (0,-1),(-1,-1), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.HexColor("#f9f9f9"), colors.white]),
                ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
                ("FONTSIZE",   (0,1), (-1,-1), 8),
                ("ALIGN",      (3,0), (-1,-1), "RIGHT"),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ]))
            story.append(t)
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"Total Bills: {len(rows)}    Grand Total: {curr}{grand:.2f}", styles["Normal"]))
        doc.build(story)
        messagebox.showinfo("PDF Saved", f"Report saved to:\n{path}")
 
    def _append_report_to_excel(self, date_str, sales_rows):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
 
            curr       = self.db.get_setting("currency", "Rs.")
            rname      = self.db.get_setting("restaurant_name", "Restaurant")
            gst_number = self.db.get_setting("gst_number", "")
            tax_pct    = float(self.db.get_setting("tax_percent", "0"))
            svc_pct    = float(self.db.get_setting("service_charge", "0"))
            sgst_pct   = float(self.db.get_setting("sgst_percent", "0"))
            cgst_pct   = float(self.db.get_setting("cgst_percent", "0"))
            tax_enabled = tax_pct > 0
            has_svc     = svc_pct > 0
            has_gst     = bool(gst_number.strip())
 
            item_map = {}
            for sale_row in sales_rows:
                try:
                    items_data = json.loads(sale_row["items_json"]) if "items_json" in sale_row.keys() else []
                except Exception:
                    items_data = []
                for it in items_data:
                    key = (it.get("item_name","?"), it.get("category",""))
                    if key not in item_map:
                        item_map[key] = {"qty": 0, "rate": it.get("price", 0), "amount": 0}
                    item_map[key]["qty"]    += it.get("quantity", 0)
                    item_map[key]["amount"] += it.get("total", 0)
 
            if not item_map:
                c = self.db.conn.cursor()
                c.execute("""SELECT b.items_json FROM bills b WHERE DATE(b.created_at)=?""", (date_str,))
                for brow in c.fetchall():
                    try:
                        for it in json.loads(brow["items_json"]):
                            key = (it.get("item_name","?"), it.get("category",""))
                            if key not in item_map:
                                item_map[key] = {"qty": 0, "rate": it.get("price",0), "amount": 0}
                            item_map[key]["qty"]    += it.get("quantity", 0)
                            item_map[key]["amount"] += it.get("total", 0)
                    except Exception:
                        pass
 
            item_rows = [
                {"name": k[0], "category": k[1],
                 "qty": v["qty"], "rate": v["rate"], "amount": v["amount"]}
                for k, v in item_map.items()
            ]
 
            month_key  = date_str[:7]
            sheet_name = datetime.strptime(month_key, "%Y-%m").strftime("%b %Y")
            date_label = datetime.strptime(date_str, "%Y-%m-%d").strftime("%A, %d %B %Y")
            _ensure_hidden_dir()
 
            if os.path.exists(SALES_EXCEL_FILE):
                wb = load_workbook(SALES_EXCEL_FILE)
            else:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
 
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and date_label in str(cell):
                            return
            else:
                ws = wb.create_sheet(title=sheet_name)
 
            ORANGE = "FF6B35"; DARK = "1E1E2E"; COL_HDR = "2D2D3E"
            LIGHT  = "FFF8F0"; WHITE = "FFFFFF"; GREY = "F0F0F0"
            thin   = Side(style="thin", color="CCCCCC")
            bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
            al_c   = Alignment(horizontal="center", vertical="center")
            al_r   = Alignment(horizontal="right",  vertical="center")
            al_l   = Alignment(horizontal="left",   vertical="center")
            def fill(c): return PatternFill("solid", start_color=c)
 
            COL_LAST = 6
            is_first = (ws.max_row == 1 and ws.cell(1,1).value is None)
 
            if is_first:
                for col, w in zip("ABCDEF", [5, 28, 16, 7, 13, 14]):
                    ws.column_dimensions[col].width = w
 
            if is_first:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_LAST)
                c = ws.cell(1, 1, rname.upper())
                c.font = Font(name="Georgia", size=15, bold=True, color=ORANGE)
                c.fill = fill(DARK); c.alignment = al_c
                ws.row_dimensions[1].height = 28
 
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COL_LAST)
                sub = f"Monthly Sales Report  —  {sheet_name}"
                if has_gst:
                    sub += f"     GST: {gst_number.strip()}"
                c = ws.cell(2, 1, sub)
                c.font = Font(name="Arial", size=9, italic=True, color="999999")
                c.fill = fill(DARK); c.alignment = al_c
                ws.row_dimensions[2].height = 16
                next_row = 4
            else:
                next_row = ws.max_row + 2
 
            ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=COL_LAST)
            c = ws.cell(next_row, 1, f"  {date_label}")
            c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            c.fill = fill(ORANGE); c.alignment = al_l
            ws.row_dimensions[next_row].height = 22
            next_row += 1
 
            for col, h in enumerate(["#", "Item Name", "Category", "Qty",
                                      f"Rate ({curr})", f"Amount ({curr})"], 1):
                c = ws.cell(next_row, col, h)
                c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
                c.fill = fill(COL_HDR)
                c.alignment = al_l if col == 2 else al_c
                c.border = bdr
            ws.row_dimensions[next_row].height = 17
            next_row += 1
 
            day_subtotal = 0
            for sno, item in enumerate(item_rows, 1):
                bg = LIGHT if sno % 2 == 0 else WHITE
                vals   = [sno, item["name"], item["category"], item["qty"], item["rate"], item["amount"]]
                aligns = [al_c, al_l, al_c, al_c, al_r, al_r]
                for col, (val, al) in enumerate(zip(vals, aligns), 1):
                    c = ws.cell(next_row, col, val)
                    c.font = Font(name="Arial", size=9)
                    c.fill = fill(bg); c.alignment = al; c.border = bdr
                    if col in (5, 6): c.number_format = "#,##0.00"
                ws.row_dimensions[next_row].height = 15
                day_subtotal += item["amount"]
                next_row += 1
 
            def summary_row(label, value, bold=False, bg_col=GREY, fg_col="333333"):
                ws.merge_cells(start_row=next_row, start_column=1,
                               end_row=next_row, end_column=COL_LAST - 1)
                lc = ws.cell(next_row, 1, label)
                lc.font = Font(name="Arial", size=9, bold=bold, color=fg_col)
                lc.fill = fill(bg_col); lc.alignment = al_r; lc.border = bdr
                vc = ws.cell(next_row, COL_LAST, value if value != "" else "")
                vc.font = Font(name="Arial", size=9, bold=bold, color=fg_col)
                vc.fill = fill(bg_col); vc.alignment = al_r; vc.border = bdr
                if isinstance(value, (int, float)):
                    vc.number_format = "#,##0.00"
                ws.row_dimensions[next_row].height = 15
 
            summary_row("Subtotal", day_subtotal)
            next_row += 1
            running = day_subtotal
 
            combined_gst = sgst_pct + cgst_pct if (sgst_pct + cgst_pct) > 0 else tax_pct
            if combined_gst > 0:
                pre_tax  = round(day_subtotal / (1 + combined_gst / 100), 2)
                tax_total = round(day_subtotal - pre_tax, 2)
                if sgst_pct > 0 and cgst_pct > 0:
                    sgst_total = round(pre_tax * sgst_pct / 100, 2)
                    cgst_total = round(pre_tax * cgst_pct / 100, 2)
                    summary_row(f"SGST  ({sgst_pct}%)", sgst_total)
                    next_row += 1
                    summary_row(f"CGST  ({cgst_pct}%)", cgst_total)
                    next_row += 1
                else:
                    summary_row(f"GST  ({tax_pct}%)", tax_total)
                    next_row += 1
                running = pre_tax + tax_total
 
            if has_svc:
                svc_amt = round(day_subtotal * svc_pct / 100, 2)
                summary_row(f"Service Charge  ({svc_pct}%)", svc_amt)
                next_row += 1
                running += svc_amt
 
            if has_gst:
                summary_row(f"GST No: {gst_number.strip()}", "—")
                next_row += 1
 
            ws.merge_cells(start_row=next_row, start_column=1,
                           end_row=next_row, end_column=COL_LAST - 1)
            lc = ws.cell(next_row, 1, f"Day Total  —  {date_label}")
            lc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            lc.fill = fill(ORANGE); lc.alignment = al_r; lc.border = bdr
            vc = ws.cell(next_row, COL_LAST, running)
            vc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            vc.fill = fill(ORANGE); vc.alignment = al_r
            vc.number_format = "#,##0.00"; vc.border = bdr
            ws.row_dimensions[next_row].height = 20
 
            wb.save(SALES_EXCEL_FILE)
 
        except Exception as e:
            print(f"[Excel] Error: {e}")
 
    def open_reports(self):
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])   # accent border
        W, H = 680, 580
        self._center_win(win, W, H)
        win.grab_set()
 
        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=self.C["bg"])
        inner.pack(fill="both", expand=True)
 
        tbar = tk.Frame(inner, bg=self.C["accent"], pady=6)
        tbar.pack(fill="x")
        tk.Label(tbar, text="Daily Sales Report",
                font=("Arial", 13, "bold"),
                bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                bg=self.C["accent"], fg="black", font=("Arial", 11, "bold"),
                bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)
 
        top = tk.Frame(inner, bg=self.C["bg"], pady=6)
        top.pack(fill="x", padx=12)
        tk.Label(top, text="Date:", font=("Arial", 10),
                bg=self.C["bg"], fg=self.C["muted"]).pack(side="left")
        date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(top, textvariable=date_var, width=12, font=("Arial", 11),
                bg=self.C["card"], fg=self.C["text"],
                insertbackground=self.C["accent"], bd=0).pack(side="left", padx=6, ipady=4)
 
        self._last_report_data = []
 
        report_txt = tk.Text(inner, font=("Courier New", 9),
                            bg=self.C["card"], fg=self.C["text"], bd=0, padx=8, pady=8)
        report_txt.pack(fill="both", expand=True, padx=6, pady=3)
        report_txt.configure(state="disabled")
 
        def generate():
            date_str = date_var.get().strip()
            rows = self.db.get_daily_sales(date_str)
            self._last_report_data = rows
            if rows:
                self._append_report_to_excel(date_str, rows)
            report_txt.configure(state="normal")
            report_txt.delete("1.0", "end")
            curr  = self.db.get_setting("currency", "Rs.")
            rname = self.db.get_setting("restaurant_name", "Restaurant")
            W2 = 60
 
            def ins(text=""):
                report_txt.insert("end", text + "\n")
 
            ins("=" * W2)
            ins(rname.upper().center(W2))
            ins(f"DAILY SALES REPORT  --  {date_str}".center(W2))
            ins("=" * W2)
 
            if not rows:
                ins("  No sales found for this date.")
                report_txt.configure(state="disabled"); return
 
            # ── Fetch all bill items for this date ─────────────
            c = self.db.conn.cursor()
            c.execute("""SELECT b.items_json, b.order_type
                         FROM bills b WHERE DATE(b.created_at)=?""", (date_str,))
            bill_rows = c.fetchall()
 
            # Aggregate items sold
            item_map = {}   # {item_name: {qty, total}}
            for br in bill_rows:
                try:
                    items_data = json.loads(br["items_json"])
                except Exception:
                    items_data = []
                for it in items_data:
                    name = it.get("item_name", "?")
                    if name not in item_map:
                        item_map[name] = {"qty": 0, "price": it.get("price", 0), "total": 0}
                    item_map[name]["qty"]   += it.get("quantity", 0)
                    item_map[name]["total"] += it.get("total", 0)
 
            # ── Items sold section ──────────────────────────────
            ins()
            ins("  ITEMS SOLD")
            ins("-" * W2)
            ins(f"  {'ITEM':<28} {'QTY':>5}  {'PRICE':>8}  {'AMOUNT':>9}")
            ins("-" * W2)
            items_total = 0
            for name, v in sorted(item_map.items()):
                line = f"  {name[:28]:<28} {v['qty']:>5}  {curr}{v['price']:>6.2f}  {curr}{v['total']:>7.2f}"
                ins(line)
                items_total += v["total"]
            ins("-" * W2)
            ins(f"  {'Items Subtotal':<40} {curr}{items_total:>7.2f}")
            ins()
 
            # ── Totals by order type ────────────────────────────
            by_type   = {}
            by_method = {}
            grand     = 0
            for row in rows:
                ot = row["order_type"]
                by_type[ot]   = by_type.get(ot, 0) + row["total"]
                pm = row["payment_method"]
                by_method[pm] = by_method.get(pm, 0) + row["total"]
                grand        += row["total"]
 
            ins("  TOTAL BY ORDER TYPE")
            ins("-" * W2)
            type_labels = {"dine_in": "Dine-In", "takeaway": "Takeaway", "delivery": "Delivery"}
            for ot, amt in by_type.items():
                label = type_labels.get(ot, ot.replace("_", " ").title())
                ins(f"    {label:<20} {curr}{amt:.2f}")
            ins()
 
            ins("  TOTAL BY PAYMENT METHOD")
            ins("-" * W2)
            for method, amt in by_method.items():
                ins(f"    {method.upper():<20} {curr}{amt:.2f}")
            ins()
 
            ins("=" * W2)
            ins(f"  Total Bills : {len(rows)}")
            ins(f"  GRAND TOTAL : {curr}{grand:.2f}".center(W2))
            ins("=" * W2)
            report_txt.configure(state="disabled")
 
        tk.Button(top, text="Generate", command=generate,
                bg=self.C["accent"], fg="black",
                font=("Arial", 10, "bold"), bd=0, padx=12, pady=4,
                cursor="hand2").pack(side="left", padx=4)

        tk.Button(top, text="📄 Save PDF",
                command=lambda: self._save_report_pdf(
                    date_var.get().strip(), self._last_report_data),
                bg="#E53935", fg="white",
                font=("Arial", 10, "bold"), bd=0, padx=12, pady=4,
                cursor="hand2").pack(side="left", padx=2)

        tk.Button(top, text="📊 Open Excel",
                command=lambda: os.startfile(SALES_EXCEL_FILE)
                    if os.path.exists(SALES_EXCEL_FILE) and sys.platform == "win32"
                    else messagebox.showinfo("Excel", f"File location:\n{SALES_EXCEL_FILE}"),
                bg=self.C["green"], fg="white",
                font=("Arial", 10, "bold"), bd=0, padx=12, pady=4,
                cursor="hand2").pack(side="left", padx=2)
        
        tk.Button(top, text="🚫 Cancelled Orders",
                command=self.open_cancelled_orders_report,
                bg="#FF5722", fg="white",
                font=("Arial", 10, "bold"), bd=0, padx=12, pady=4,
                cursor="hand2").pack(side="left", padx=2)
 
        generate()

    def open_cancelled_orders_report(self):
        """Show report of cancelled orders."""
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])
        W, H = 720, 580
        self._center_win(win, W, H)
        win.grab_set()

        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=self.C["bg"])
        inner.pack(fill="both", expand=True)

        tbar = tk.Frame(inner, bg=self.C["accent"], pady=6)
        tbar.pack(fill="x")
        tk.Label(tbar, text="🚫  Cancelled Orders Report",
                font=("Arial", 13, "bold"),
                bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                bg=self.C["accent"], fg="black", font=("Arial", 11, "bold"),
                bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)

        # Date filters
        filter_frame = tk.Frame(inner, bg=self.C["bg"], pady=8)
        filter_frame.pack(fill="x", padx=12)
        
        tk.Label(filter_frame, text="From:", font=("Arial", 10),
                bg=self.C["bg"], fg=self.C["muted"]).pack(side="left", padx=(0, 4))
        from_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(filter_frame, textvariable=from_var, width=12, font=("Arial", 11),
                bg=self.C["card"], fg=self.C["text"],
                insertbackground=self.C["accent"], bd=0).pack(side="left", padx=(0, 12), ipady=4)
        
        tk.Label(filter_frame, text="To:", font=("Arial", 10),
                bg=self.C["bg"], fg=self.C["muted"]).pack(side="left", padx=(0, 4))
        to_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(filter_frame, textvariable=to_var, width=12, font=("Arial", 11),
                bg=self.C["card"], fg=self.C["text"],
                insertbackground=self.C["accent"], bd=0).pack(side="left", padx=(0, 12), ipady=4)
        
        # Treeview for cancelled orders
        tree_frame = tk.Frame(inner, bg=self.C["bg"])
        tree_frame.pack(fill="both", expand=True, padx=12, pady=8)
        
        cols = ("Time", "Order #", "Type", "Table", "Items", "Amount", "Reason", "By")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=16)
        
        style = ttk.Style()
        style.configure("cancelled.Treeview",
                        background=self.C["card"], fieldbackground=self.C["card"],
                        foreground=self.C["text"], rowheight=28,
                        font=("Arial", 9))
        style.configure("cancelled.Treeview.Heading",
                        background=self.C["border"], foreground=self.C["muted"],
                        font=("Arial", 9, "bold"))
        tree.configure(style="cancelled.Treeview")
        
        col_widths = [70, 120, 80, 80, 50, 80, 150, 80]
        for col, w in zip(cols, col_widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center")
        
        tree.column("Reason", anchor="w")
        
        vsb = tk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)
        
        # Summary labels
        summary_frame = tk.Frame(inner, bg=self.C["bg"], pady=8)
        summary_frame.pack(fill="x", padx=12)
        
        total_lbl = tk.Label(summary_frame, text="Total Cancelled: 0 orders | Rs. 0.00",
                            font=("Arial", 11, "bold"),
                            bg=self.C["bg"], fg=self.C["accent"])
        total_lbl.pack(side="left")
        
        def load_data():
            tree.delete(*tree.get_children())
            rows = self.db.get_cancelled_orders_report(from_var.get(), to_var.get())
            
            total_count = 0
            total_amount = 0
            
            for row in rows:
                row = dict(row)
                time_str = datetime.fromisoformat(row["cancelled_at"]).strftime("%H:%M")
                
                tree.insert("", "end", values=(
                    time_str,
                    row["order_number"][:15],
                    row["order_type"].replace("_", " ").title(),
                    row["table_number"][:10],
                    row["total_items"],
                    f"Rs.{row['subtotal']:.0f}",
                    row["reason"][:30],
                    row["cancelled_by"][:12]
                ))
                
                total_count += 1
                total_amount += row["subtotal"]
            
            total_lbl.config(text=f"Total Cancelled: {total_count} orders | Rs. {total_amount:.2f}")
        
        tk.Button(filter_frame, text="Load Report", command=load_data,
                bg=self.C["accent"], fg="black",
                font=("Arial", 10, "bold"), bd=0, padx=14, pady=4,
                cursor="hand2").pack(side="left", padx=4)
        
        # Export button
        def export_cancelled():
            if not tree.get_children():
                messagebox.showinfo("No Data", "No cancelled orders to export.")
                return
            
            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
                initialfile=f"Cancelled_Orders_{from_var.get()}_to_{to_var.get()}.txt")
            if not path:
                return
            
            try:
                with open(path, "w", encoding="utf-8") as f:
                    f.write("CANCELLED ORDERS REPORT\n")
                    f.write(f"Period: {from_var.get()} to {to_var.get()}\n")
                    f.write("=" * 80 + "\n\n")
                    
                    for item in tree.get_children():
                        vals = tree.item(item)["values"]
                        f.write(f"Time: {vals[0]}\n")
                        f.write(f"Order: {vals[1]}\n")
                        f.write(f"Type: {vals[2]}\n")
                        f.write(f"Table: {vals[3]}\n")
                        f.write(f"Items: {vals[4]}\n")
                        f.write(f"Amount: {vals[5]}\n")
                        f.write(f"Reason: {vals[6]}\n")
                        f.write(f"Cancelled By: {vals[7]}\n")
                        f.write("-" * 40 + "\n")
                
                messagebox.showinfo("Exported", f"Report saved to:\n{path}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not export: {e}")
        
        tk.Button(filter_frame, text="📄 Export", command=export_cancelled,
                bg=self.C["card"], fg=self.C["text"],
                font=("Arial", 10, "bold"), bd=0, padx=14, pady=4,
                cursor="hand2").pack(side="left", padx=4)
        
        load_data()

    def open_printer_cfg(self):
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])   # accent border
        self._center_win(win, 424, 324)
        win.grab_set()
 
        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=self.C["bg"])
        inner.pack(fill="both", expand=True)
 
        tbar = tk.Frame(inner, bg=self.C["accent"], pady=8)
        tbar.pack(fill="x")
        tk.Label(tbar, text="Printer Settings",
                 font=("Arial", 13, "bold"),
                 bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                  bg=self.C["accent"], fg="black", font=("Arial", 11, "bold"),
                  bd=0, padx=10, cursor="hand2", relief="flat").pack(side="right", padx=6)
        f = tk.Frame(inner, bg=self.C["bg"], padx=30, pady=20)
        f.pack(fill="both", expand=True)
        tk.Label(f, text="Print Method:", font=("Arial", 10),
                 bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w")
        cfg = self.printer.get_config()
        method_var = tk.StringVar(value=cfg.get("method", "Browser"))
        ttk.Combobox(f, textvariable=method_var,
                     values=["Browser", "Thermal Printer"],
                     font=("Arial", 10), width=28, state="readonly").pack(fill="x", pady=(0, 16))
        tk.Label(f, text="Printer Name (Thermal only):", font=("Arial", 10),
                 bg=self.C["bg"], fg=self.C["muted"]).pack(anchor="w")
        printer_var = tk.StringVar(value=cfg.get("printer_name", ""))
        try:
            printers = [p[2] for p in __import__("win32print").EnumPrinters(2)] if WINDOWS_PRINT_AVAILABLE else []
        except Exception:
            printers = []
        ttk.Combobox(f, textvariable=printer_var, values=printers,
                     font=("Arial", 10), width=28).pack(fill="x", pady=(0, 16))
        def save():
            self.printer.save_config({"method": method_var.get(), "printer_name": printer_var.get()})
            messagebox.showinfo("Saved", "Printer settings saved!", parent=win)
            win.destroy()
        tk.Button(inner, text="Save", command=save,
                  bg=self.C["green"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, pady=10,
                  cursor="hand2").pack(fill="x", padx=30, pady=(0, 16))
 
    def _make_dialog(self, width, height, title=""):
        """
        Creates a borderless Toplevel with:
        - accent-colored outer border (3px)
        - card-colored inner area
        - no raw square OS background visible
        """
        win = tk.Toplevel(self.root)
        win.overrideredirect(True)
        win.configure(bg=self.C["accent"])   # accent = border color
        win.resizable(False, False)
        self._center_win(win, width, height)
        win.grab_set()
 
        # Outer accent border
        border = tk.Frame(win, bg=self.C["accent"], padx=3, pady=3)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
 
        # Title bar
        tbar = tk.Frame(border, bg=self.C["accent"], pady=8)
        tbar.pack(fill="x")
        tk.Label(tbar, text=title, font=("Arial", 11, "bold"),
                 bg=self.C["accent"], fg="black").pack(side="left", padx=14)
        tk.Button(tbar, text="✕", command=win.destroy,
                  bg=self.C["accent"], fg="black",
                  font=("Arial", 10, "bold"), bd=0, padx=8,
                  cursor="hand2", relief="flat").pack(side="right", padx=4)
 
        # Inner content frame
        frame = tk.Frame(border, bg=self.C["card"])
        frame.pack(fill="both", expand=True)
 
        return {"win": win, "frame": frame, "close": win.destroy}
 
    def _center_win(self, win, w, h):
        win.update_idletasks()
        x = (win.winfo_screenwidth()  - w) // 2
        y = (win.winfo_screenheight() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")
 
    def on_closing(self):
            self.session.end_session()
            if hasattr(self, "_mobile_server") and self._mobile_server:
                self._mobile_server.stop()
            self.root.destroy()
 
 
# ═══════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ═══════════════════════════════════════════════════════════════
class LoginWindow:
    C = {
        "bg":     "#0a0a0a",
        "panel":  "#111111",
        "card":   "#1a1a1a",
        "accent": "#FF6B35",
        "text":   "#F5F5F5",
        "muted":  "#666666",
    }
 
    def __init__(self):
        self.db   = AuthDB()
        self.root = tk.Tk()
        self.root.title("RESTO BILL — Login")
        self.root.state("zoomed")
        self.root.configure(bg=self.C["bg"])
        self.root.overrideredirect(True)
 
        self._set_window_icon()
        self._build_ui()
        if _demo_mode_enabled():
            self.root.after(1200, self._demo_login)
        self.root.mainloop()
 
    def _set_window_icon(self):
        for name in ("icon.ico", "icon.png"):
            path = resource_path(name)
            if os.path.exists(path):
                try:
                    img = tk.PhotoImage(file=path)
                    self.root.iconphoto(True, img)
                    return
                except Exception:
                    pass
        try:
            ico = resource_path("icon.ico")
            if os.path.exists(ico):
                self.root.iconbitmap(ico)
        except Exception:
            pass
 
    def _build_ui(self):
        bg_frame = tk.Frame(self.root, bg=self.C["bg"])
        bg_frame.place(relx=0, rely=0, relwidth=1, relheight=1)
 
        brand = tk.Frame(bg_frame, bg="#110800")
        brand.place(relx=0, rely=0, relwidth=0.55, relheight=1)
 
        tk.Label(brand, text="🍽️", font=("Segoe UI Emoji", 90),
                 bg="#110800", fg=self.C["accent"]).place(relx=0.5, rely=0.35, anchor="center")
        tk.Label(brand, text="RESTO BILL",
                 font=("Georgia", 52, "bold"),
                 bg="#110800", fg=self.C["accent"]).place(relx=0.5, rely=0.48, anchor="center")
        tk.Label(brand, text="Restaurant Billing System",
                 font=("Georgia", 16),
                 bg="#110800", fg="#664422").place(relx=0.5, rely=0.57, anchor="center")
        tk.Label(brand, text="Manage tables  ·  Take orders  ·  Print bills",
                 font=("Segoe UI", 11),
                 bg="#110800", fg="#444444").place(relx=0.5, rely=0.63, anchor="center")
        tk.Label(brand, text="Fast  ·  Simple  ·  Reliable",
                 font=("Segoe UI", 10),
                 bg="#110800", fg="#333322").place(relx=0.5, rely=0.72, anchor="center")
 
        login_panel = tk.Frame(bg_frame, bg=self.C["card"])
        login_panel.place(relx=0.55, rely=0, relwidth=0.45, relheight=1)
 
        self.container = tk.Frame(login_panel, bg=self.C["card"])
        self.container.place(relx=0.5, rely=0.5, anchor="center", width=340)
        self.show_login()
 
    def clear(self):
        for w in self.container.winfo_children():
            w.destroy()
 
    def show_login(self):
        self.clear()
        tk.Label(self.container, text="Welcome Back",
                 font=("Georgia", 24, "bold"),
                 bg=self.C["card"], fg=self.C["text"]).pack(pady=(0, 4))
        tk.Label(self.container, text="Sign in to continue",
                 font=("Segoe UI", 11),
                 bg=self.C["card"], fg=self.C["muted"]).pack(pady=(0, 30))
 
        for field, show in [("Username", ""), ("Password", "●")]:
            tk.Label(self.container, text=field,
                     font=("Segoe UI", 10),
                     bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
            e = tk.Entry(self.container, font=("Segoe UI", 13),
                         bg="#222222", fg=self.C["text"],
                         insertbackground=self.C["accent"],
                         bd=0, show=show if show else "")
            e.pack(fill="x", ipady=10, pady=(4, 14))
            if field == "Username":
                self.username = e
            else:
                self.password = e
                e.bind("<Return>", lambda _: self.login())
 
        tk.Button(self.container, text="LOGIN",
                  command=self.login,
                  bg=self.C["accent"], fg="white",
                  font=("Georgia", 13, "bold"),
                  bd=0, pady=12, cursor="hand2",
                  activebackground="#cc5522").pack(fill="x", pady=(6, 16))
 
        tk.Button(self.container, text="Create new account →",
                  command=self.show_signup,
                  bg=self.C["card"], fg=self.C["accent"],
                  font=("Segoe UI", 10), bd=0,
                  cursor="hand2").pack()
 
        tk.Button(self.container, text="Exit",
                  command=self.root.destroy,
                  bg=self.C["card"], fg=self.C["muted"],
                  font=("Segoe UI", 9), bd=0,
                  cursor="hand2").pack(pady=(10, 0))

    def _demo_login(self):
        try:
            user = os.environ.get("RESTOBILL_DEMO_USER", "demo")
            pwd = os.environ.get("RESTOBILL_DEMO_PASS", "demo123")
            self.username.delete(0, tk.END)
            self.username.insert(0, user)
            self.password.delete(0, tk.END)
            self.password.insert(0, pwd)
            self.root.after(700, self.login)
        except Exception:
            pass

    def show_signup(self):
        self.clear()
        tk.Label(self.container, text="Create Account",
                 font=("Georgia", 22, "bold"),
                 bg=self.C["card"], fg=self.C["text"]).pack(pady=(0, 4))
        tk.Label(self.container, text="Register to get started",
                 font=("Segoe UI", 10),
                 bg=self.C["card"], fg=self.C["muted"]).pack(pady=(0, 24))
 
        for field, show in [("Username", ""), ("Password", "●")]:
            tk.Label(self.container, text=field,
                     font=("Segoe UI", 10),
                     bg=self.C["card"], fg=self.C["muted"]).pack(anchor="w")
            e = tk.Entry(self.container, font=("Segoe UI", 13),
                         bg="#222222", fg=self.C["text"],
                         insertbackground=self.C["accent"],
                         bd=0, show=show if show else "")
            e.pack(fill="x", ipady=10, pady=(4, 14))
            if field == "Username":
                self.username = e
            else:
                self.password = e
 
        tk.Button(self.container, text="SIGN UP",
                  command=self.signup,
                  bg="#4CAF50", fg="white",
                  font=("Georgia", 13, "bold"),
                  bd=0, pady=12, cursor="hand2").pack(fill="x", pady=(6, 16))
 
        tk.Button(self.container, text="← Back to login",
                  command=self.show_login,
                  bg=self.C["card"], fg=self.C["accent"],
                  font=("Segoe UI", 10), bd=0,
                  cursor="hand2").pack()
 
        tk.Button(self.container, text="Exit",
                  command=self.root.destroy,
                  bg=self.C["card"], fg=self.C["muted"],
                  font=("Segoe UI", 9), bd=0,
                  cursor="hand2").pack(pady=(10, 0))
 
    def signup(self):
        user = self.username.get().strip()
        pwd  = self.password.get().strip()
        if not user or not pwd:
            messagebox.showerror("Error", "All fields required"); return
        if self.db.signup(user, pwd):
            messagebox.showinfo("Success", "Account created! Please login.")
            self.show_login()
        else:
            messagebox.showerror("Error", "Username already exists")
 
    def login(self):
        user = self.username.get().strip()
        pwd  = self.password.get().strip()
        if not user or not pwd:
            messagebox.showwarning("Required", "Please enter username and password."); return
        row = self.db.login(user, pwd)
        if not row:
            messagebox.showerror("Login Failed", "Invalid username or password."); return
 
        lm = LicenseManager()
        status, _ = lm.check_license()
        if status not in ("ok", "warning"):
            self.root.withdraw()
            _show_blocked_window(reason=status)
            status2, _ = lm.check_license()
            if status2 not in ("ok", "warning"):
                self.root.destroy()
                return
            self.root.deiconify()
            return
 
        self.container.destroy()
        app = RestaurantApp(self.root, username=user)
 
 
# ═══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    lm = LicenseManager()
    status, _ = lm.check_license()
 
    if status == "not_activated":
        LicenseWindow()
    elif status in ("ok", "warning"):
        show_splash()
    else:
        _show_blocked_window(reason=status)
